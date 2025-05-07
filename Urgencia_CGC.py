import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from statistics import mean, median, mode
import re

# Configurações iniciais
chrome_options = Options()
chrome_options.add_argument("--headless")  # Remove se quiser ver o navegador
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--window-size=1920x1080")

# Caminho para o driver do Chrome (baixe o correspondente à sua versão)
driver_path = "chromedriver.exe"  # Altere conforme necessário
service = Service(driver_path)


# Função para verificar se um item é unitário
def is_unitario(texto):
    texto = texto.lower()
    padroes_descartar = [
        r'\d+\s*[x×]\s*\d+',  # Padrão para "2x", "3 x", etc.
        r'kit\b',
        r'conjunto\b',
        r'\d+\s*unidades?',
        r'pack\b',
        r'\d+\s*peças?',
        r'combo\b'
    ]

    for padrao in padroes_descartar:
        if re.search(padrao, texto):
            return False
    return True


# Função para pesquisar no Google Shopping e obter preços
def pesquisar_google_shopping(produto):
    driver = webdriver.Chrome(service=service, options=chrome_options)
    try:
        driver.get("https://www.google.com/shopping")
        time.sleep(2)

        # Localizar e preencher a barra de pesquisa
        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(produto)
        search_box.send_keys(Keys.RETURN)
        time.sleep(3)

        # Coletar os resultados
        resultados = []
        items = driver.find_elements(By.CSS_SELECTOR, "div.sh-dgr__content")

        for item in items[:6]:  # Pegar apenas os 6 primeiros
            try:
                # Extrair título
                titulo = item.find_element(By.CSS_SELECTOR, "h3.tAxDx").text

                # Verificar se é unitário
                if not is_unitario(titulo):
                    continue

                # Extrair preço
                preco_element = item.find_element(By.CSS_SELECTOR, "span.a8Pemb")
                preco_text = preco_element.text.replace('R$', '').replace('.', '').replace(',', '.').strip()
                preco = float(preco_text)

                # Extrair link
                link_element = item.find_element(By.CSS_SELECTOR, "a.shntl")
                link = link_element.get_attribute("href")

                resultados.append({"preco": preco, "link": link, "titulo": titulo})

            except Exception as e:
                print(f"Erro ao processar item: {e}")
                continue

        return resultados

    except Exception as e:
        print(e)

# Função principal
def processar_planilha(caminho_arquivo):
    # Carregar a planilha
    wb = load_workbook(caminho_arquivo)
    ws_principal = wb.active  # Supondo que a primeira planilha é a principal

    # Criar ou acessar a planilha de resultados
    if "Resultados Google" not in wb.sheetnames:
        ws_resultados = wb.create_sheet("Resultados Google")
        ws_resultados.append(["ID Produto", "Preço 1", "Link 1", "Preço 2", "Link 2",
                              "Preço 3", "Link 3", "Preço 4", "Link 4", "Preço 5", "Link 5",
                              "Preço 6", "Link 6", "Média", "Mediana", "Moda"])
    else:
        ws_resultados = wb["Resultados Google"]

    # Configurar formatação condicional (vermelho claro)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Processar linhas a partir da linha 11 (considerando que a linha 12 é cabeçalho)
    for row in range(11, ws_principal.max_row + 1):
        produto = ws_principal[f'C{row}'].value

        if not produto:
            continue

        print(f"Processando: {produto}")

        # Pesquisar no Google Shopping
        resultados = pesquisar_google_shopping(produto)

        if not resultados:
            print(f"Nenhum resultado válido encontrado para {produto}")
            continue

        # Extrair preços e links
        precos = [item['preco'] for item in resultados]
        links = [item['link'] for item in resultados]

        # Calcular estatísticas
        media = mean(precos)
        mediana = median(precos)
        try:
            moda = mode(precos)
        except:
            moda = "N/A"  # Pode não haver moda se todos os valores forem únicos

        # Adicionar dados à planilha de resultados
        dados_linha = [row]  # ID baseado no número da linha
        for i in range(6):
            if i < len(resultados):
                dados_linha.extend([resultados[i]['preco'], resultados[i]['link']])
            else:
                dados_linha.extend(["", ""])

        dados_linha.extend([media, mediana, moda])
        ws_resultados.append(dados_linha)

        # Adicionar média à planilha principal (nova coluna)
        coluna_media = ws_principal.max_column + 1
        ws_principal.cell(row=12, column=coluna_media, value="Média Google")
        ws_principal.cell(row=row, column=coluna_media, value=media)

        # Verificar se o valor unitário é 25% maior que a média
        valor_unitario = ws_principal[f'D{row}'].value  # Supondo que a coluna D tem o valor unitário
        if valor_unitario and isinstance(valor_unitario, (int, float)):
            if valor_unitario > media * 1.25:
                # Aplicar formatação condicional a toda a linha
                for col in range(1, ws_principal.max_column + 1):
                    ws_principal.cell(row=row, column=col).fill = red_fill

    # Salvar as alterações
    wb.save(caminho_arquivo)
    print("Processamento concluído!")


if __name__ == "__main__":
    processar_planilha()