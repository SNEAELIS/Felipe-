import sys, os
import re
import json
import traceback

import pandas as pd
import pdfplumber

from datetime import datetime
from functools import wraps

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from collections import ChainMap

#==================================================
# PAYROLL PARSER (PDF --> EXCEL)  
# =================================================

def normalize_text(text):
    """Normalize text to UTF-8-safe str, remove R$ currency markers, and normalize whitespace."""
    if text is None:
        return ""
    if isinstance(text, bytes):
        text = text.decode("utf-8", errors="replace")

    text = str(text).strip()
    text = text.replace("R$", "").replace("r$", "")
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text

def clean_header(text):
    text = normalize_text(text)
    return re.sub(r"[^A-Za-zÀ-ÿ0-9 ]", "", text).strip()

def debug_pdf(pdf_path, page_range=1):
    """
    Comprehensive debugger for PDF structure - Fixed version
    
    Args:
        pdf_path: Path to the PDF file
        page_range: Single page (int), range as tuple (start, end), or range as string "start-end"
                   Examples: 1, (1, 5), "1-5"
    """
    pages_to_debug = []
    if isinstance(page_range, int):
        pages_to_debug = [page_range]
    elif isinstance(page_range, tuple) and len(page_range) == 2:
        pages_to_debug = list(range(page_range[0], page_range[1] + 1))
    elif isinstance(page_range, str) and '-' in page_range:
        parts = page_range.split('-')
        if len(parts) == 2:
            try:
                start, end = int(parts[0].strip()), int(parts[1].strip())
                pages_to_debug = list(range(start, end + 1))
            except ValueError:
                print(f"❌ Invalid range format: {page_range}. Use 'start-end' format.")
                return
    else:
        print(f"❌ Invalid page_range: {page_range}. Use int, tuple (start, end), or string 'start-end'.")
        return
    
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        valid_pages = [p for p in pages_to_debug if 1 <= p <= total_pages]
        if not valid_pages:
            print(f"❌ No valid pages in range. PDF has {total_pages} pages.")
            return
        if len(valid_pages) < len(pages_to_debug):
            invalid = [p for p in pages_to_debug if p not in valid_pages]
            print(f"⚠️  Skipping invalid pages: {invalid}")
        
        for page_num in valid_pages:
            _debug_pdf_page(pdf, pdf_path, page_num, total_pages)

def _debug_pdf_page(pdf, pdf_path, page_num, total_pages):
    """Helper function to debug a single page"""
    print("="*80)
    print(f"🔍 PDF DEBUGGER - Analyzing: {pdf_path.split('\\')[-1]}")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Page {page_num}/{total_pages}")
    print("="*80)
    
    print(f"\n📄 DOCUMENT INFO:")
    print(f"   Total pages: {total_pages}")
    print(f"   Metadata: {pdf.metadata if pdf.metadata else 'No metadata'}")
    
    page = pdf.pages[page_num - 1]
    
    print(f"\n📃 PAGE {page_num} INFO:")
    print(f"   Width: {page.width}")
    print(f"   Height: {page.height}")
    print(f"   Crop box: {page.cropbox}")
    
    print(f"\n📝 RAW TEXT EXTRACTION:")
    print("-"*60)
    text = page.extract_text()
    if text:
        print(f"✓ Text extracted ({len(text)} characters, {len(text.split())} words)")
        print(f"\n📏 TEXT BY LINES:")
        print("-"*60)
        lines = page.extract_text_lines()
        if lines:
            print(f"✓ Found {len(lines)} lines")
            print("\nAll lines:")
            for i, line in enumerate(lines[:], 1):
                text_line = line.get('text', '').strip()
                if text_line:
                    print(f"   {i:3d}. {text_line[:100]}")
                    if len(text_line) > 100:
                        print(f"       ... (total {len(text_line)} chars)")
        else:
            print("❌ No lines found")
        
        print(f"\n🔤 WORDS WITH POSITIONS:")
        print("-"*60)
        words = page.extract_words()
        if words:
            print(f"✓ Found {len(words)} words")
            print("\nFirst 15 words with coordinates:")
            for i, word in enumerate(words[:15], 1):
                print(f"   {i:3d}. '{word['text']}' at (x:{word['x0']:.1f}, y:{word['top']:.1f})")
        else:
            print("❌ No words found")
        
        print(f"\n📊 TABLES DETECTION:")
        print("-"*60)
        
        strategies = ['lines', 'lines_strict']
        all_tables = []
        
        for strategy in strategies:
            try:
                print(f"\nTrying strategy: '{strategy}'")
                table_settings = {
                    "vertical_strategy": strategy, 
                    "horizontal_strategy": strategy,
                    "snap_tolerance": 4,
                }
                tables = page.extract_tables(table_settings)
                
                if not tables:
                    print(f"   ❌ No tables found with '{strategy}' strategy")
                    continue
                
                non_empty_tables = []
                for t in tables:
                    if t and len(t) > 1:
                        non_empty_tables.append(t)
                
                if not non_empty_tables:
                    print(f"   ℹ️  Found {len(tables)} table(s), but all are empty or only headers")
                    continue
                
                print(f"   ✓ Found {len(non_empty_tables)} non‑empty table(s)")
                
                for idx, table in enumerate(non_empty_tables, start=1):
                    rows = len(table)
                    cols = len(table[0]) if table else 0
                    print(f"\n   Table {idx} (strategy '{strategy}'):")
                    print(f"      Dimensions: {rows} rows × {cols} columns")
                    
                    preview_limit = min(rows,1000000000)
                    print(f"      Preview (first {preview_limit} rows):")
                    for r in range(preview_limit):
                        row_display = [str(cell) if cell is not None else 'None' for cell in table[r]]
                        print(f"         Row {r+1}: {row_display}")
                    
                    if rows > 3:
                        print(f"         ... and {rows-3} more rows")
                    
                    all_tables.append((strategy, idx, table))
                    
            except Exception as e:
                print(f"   ⚠️  Error with '{strategy}': {e}")
        
        if all_tables:
            print("\n" + "="*60)
            print("📋 SUMMARY OF ALL TABLES FOUND:")
            print("-"*60)
            for strategy, idx, table in all_tables:
                rows = len(table)
                cols = len(table[0]) if table else 0
                print(f"   • Strategy '{strategy}' → Table {idx}: {rows}×{cols}")
            
            print("\n💡 To extract a particular table, use:")
            print("   extract_table_rows_by_strategy(pdf_path, page_num=page, table_num=<index>, strategy='<strategy>')")
        else:
            print("\n❌ No tables detected with any strategy")
            print("   💡 Try these alternative settings:")
            print("      - Use 'text' strategy with custom tolerances")
            print("      - Use explicit table boundaries if known")
            print("      - Check if content is truly tabular")
        
        print(f"\n🔍 SEARCH FOR COMMON PATTERNS:")
        print("-"*60)
        if text:
            patterns = {
                'Dates (dd/mm/yyyy)': r'\d{2}/\d{2}/\d{4}',
                'Dates (mm/dd/yyyy)': r'\d{2}/\d{2}/\d{4}',
                'Dates (dd-mm-yyyy)': r'\d{2}-\d{2}-\d{4}',
                'Numbers with decimals': r'\d+\.\d{2}',
                'Email addresses': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
                'Phone numbers': r'\(?\d{2,3}\)?[\s.-]?\d{4,5}[\s.-]?\d{4}',
                'CPF (Brazil)': r'\d{3}\.\d{3}\.\d{3}-\d{2}',
                'CNPJ (Brazil)': r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}',
                'Currency (R$)': r'R\$\s*\d+[\.,]\d{2}',
            }
            
            found_patterns = []
            for name, pattern in patterns.items():
                matches = re.findall(pattern, text)
                if matches:
                    found_patterns.append((name, matches[:5]))
                    print(f"✓ {name}: Found {len(matches)} matches")
                    print(f"   Examples: {matches[:3]}")
            
            if not found_patterns:
                print("ℹ️  No common patterns found in text")
        else:
            print("⚠️  No text to search for patterns")
        
        print(f"\n🖼️  IMAGES:")
        print("-"*60)
        images = page.images
        if images:
            print(f"✓ Found {len(images)} images:")
            for i, img in enumerate(images[:5], 1):
                print(f"   Image {i}: {img.get('width', 'N/A')}x{img.get('height', 'N/A')}")
        else:
            print("ℹ️  No images found")
        
        print("\n" + "="*80)
        print("📋 EXTRACTION SUMMARY:")
        print("="*80)
        print(f"✓ Text extracted: {'✅ Yes' if text else '❌ No'}")
        print(f"✓ Tables found: {'✅ Yes' if all_tables else '❌ No'}")
        print(f"✓ Images found: {'✅ Yes' if images else '❌ No'}")
        print(f"✓ Total words: {len(words) if words else 0}")
        print(f"✓ Total characters: {len(text) if text else 0}")
        
        print("\n💡 RECOMMENDATIONS:")
        if not text:
            print("   ⚠️  No text extracted → Page is likely scanned images. Need OCR!")
        elif all_tables:
            print("   📊 Tables detected! Use page.extract_tables() with custom settings")
        else:
            print("   📝 Text-only page → Use page.extract_text() and parse manually")
        
        if text:
            print("\n📌 NEXT STEPS SUGGESTION:")
            print("   Based on the content, you might want to:")
            if any(patterns in str(found_patterns) for patterns in ['Dates', 'Currency']):
                print("   - Extract specific data using regex patterns")
            if all_tables:
                print("   - Extract tables into structured data")
            print("   - Split text by sections or delimiters")
            print("   - Create rules to identify and extract key information")
        
        print("\n" + "="*80)
        return {
            'has_text': bool(text),
            'has_tables': bool(all_tables),
            'word_count': len(words) if words else 0,
            'char_count': len(text) if text else 0,
            'page_height': page.height,
            'page_width': page.width
        }

def extract_lines_from_base_to_declaro(pdf, page_num=1):
    """
    Extract text lines from the page starting at the line beginning with 'BASE'
    and continuing until the first line that contains 'DECLARO'.
    """
    if page_num < 1 or page_num > len(pdf.pages):
        raise ValueError(f"Page {page_num} is out of range. Document has {len(pdf.pages)} pages.")

    page = pdf.pages[page_num - 1]
    lines = page.extract_text_lines()
    if not lines:
        return []

    section_lines = []
    in_section = False
    start_prefix = 'BASE'
    stop_word = 'DECLARO'

    for line in lines:
        text_line = line.get('text', '').strip()
        if not text_line:
            continue

        if not in_section:
            if text_line.upper().startswith(start_prefix.upper()):
                in_section = True
                section_lines.append(text_line)
                if stop_word.upper() in text_line.upper():
                    break
        else:
            section_lines.append(text_line)
            if stop_word.upper() in text_line.upper():
                break
    
    pattern = r'([^:]+): (R\$ [\d,.]+)'
    kv = {}

    for ln in section_lines:
        if ':' in ln:
            matches = re.findall(pattern, ln)
            for key, val in matches:
                key = key.strip()
                num = val.replace('R$ ', '').strip()
                if key.upper().startswith('BASE') or key.upper() == 'VALOR LÍQUIDO':
                    if key in kv:
                        kv[key] += ' | ' + num
                    else:
                        kv[key] = num

    if not kv:
        return pd.DataFrame()

    return pd.DataFrame([kv])

def save_to_excel(data, excel_path):
    """
    Save DataFrame(s) to Excel with formatting
    """
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        if isinstance(data, dict):
            for sheet_name, df in data.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    format_sheet(writer, sheet_name)
        else:
            df = data
            if not df.empty:
                df.to_excel(writer, sheet_name='Data', index=False)
                format_sheet(writer, 'Data')

def format_sheet(writer, sheet_name):
    """Apply formatting to a specific sheet in the workbook"""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    if worksheet.max_row > 0:
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

def extract_table_rows_by_strategy(pdf, page_num=1, table_num=None, start_row=1, end_row=5, strategy='lines'):
    """
    Extract a range of rows from a table on a given PDF page using a specified strategy.
    """
    if page_num:
        if page_num < 1 or page_num > len(pdf.pages):
            raise ValueError(f"Page {page_num} is out of range. Document has {len(pdf.pages)} pages.")

        page = pdf.pages[page_num - 1]
        table_settings = {
            "vertical_strategy": strategy,
            "horizontal_strategy": strategy,
            "snap_tolerance": 4,
            "join_tolerance": 4,
        }

        tables = page.extract_tables(table_settings)

        if not tables:
            raise ValueError(f"No tables found on page {page_num} using strategy '{strategy}'.")

        non_empty_tables = [t for t in tables if t and len(t) > 1]
        if not non_empty_tables:
            raise ValueError(f"No non-empty tables found on page {page_num} using strategy '{strategy}'.")
    
        if table_num < 1 or table_num > len(non_empty_tables):
            raise ValueError(f"Table {table_num} not found on page {page_num}. Found {len(non_empty_tables)} non-empty tables.")
        
        t = table_num

        while t <= len(tables):
            table = tables[t - 1]

            if t <= 1:
                selected_rows = table[start_row - 2:end_row]
            else:
                selected_rows = table[:]

            if not selected_rows:
                t += 1
            else:
                break

        if strategy == 'lines':
            value_per_header_dicts = []
            for row in selected_rows:
                row_dict = {}
                for item in row:
                    if item is None:
                        continue
                    normalized_item = normalize_text(item)
                    parts = normalized_item.split(':', 1)
                    left = clean_header(parts[0])
                    right = normalize_text(parts[-1])
                    if left:
                        row_dict[left] = right
                value_per_header_dicts.append(row_dict)

            merged_value_per_header_dicts = dict(ChainMap(*reversed(value_per_header_dicts)))
            return pd.DataFrame([merged_value_per_header_dicts])
        
        elif strategy == 'lines_strict':
            headers = [clean_header(h) if h is not None else None for h in selected_rows[0]]
            data_row = selected_rows[t - 1]

            raw_col_data = []
            for item in data_row:
                if item is not None:
                    parts = item.split('\n')
                    raw_col_data.append([normalize_text(p) for p in parts if p.strip() != ''] if parts else [])
                else:
                    raw_col_data.append([])

            max_rows = max(len(col) for col in raw_col_data) if raw_col_data else 0
            if max_rows == 0:
                return pd.DataFrame()
            
            df_dict = {}
            for header, col_items in zip(headers, raw_col_data):
                if header is None:
                    continue
                if header == 'Descontos':
                    padded_items = [''] * (max_rows - len(col_items)) + col_items
                else:
                    padded_items = col_items + [''] * (max_rows - len(col_items))
                df_dict[header] = padded_items

            return pd.DataFrame(df_dict)

def extract_rows_1_to_5_using_lines_strategy(pdf, page_num=1, table_num_start=1):
    """
    Extract rows 1‑5 using the 'lines' strategy.
    """
    max_tables_to_try = 5

    for t in range(table_num_start, table_num_start + max_tables_to_try):
        try:
            if page_num < 1 or page_num > len(pdf.pages):
                raise ValueError(f"Page {page_num} out of range.")

            page = pdf.pages[page_num - 1]
            table_settings = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
            }
            tables = page.extract_tables(table_settings)
            if not tables:
                continue

            non_empty_tables = [tb for tb in tables if tb and len(tb) > 1]
            if t > len(non_empty_tables):
                continue

            table = non_empty_tables[t - 1]
            selected_rows = table[0:5]

            if not selected_rows:
                continue

            value_per_header_dicts = []
            for row in selected_rows:
                row_dict = {}
                for item in row:
                    if item is None:
                        continue
                    normalized_item = normalize_text(item)
                    parts = normalized_item.split(':', 1)
                    if len(parts) == 2:
                        left = clean_header(parts[0])
                        right = normalize_text(parts[1])
                        if left:
                            row_dict[left] = right
                if row_dict:
                    value_per_header_dicts.append(row_dict)

            if not value_per_header_dicts:
                continue

            merged = dict(ChainMap(*reversed(value_per_header_dicts)))

            expected = ['Nome', 'Matrícula', 'Departamento', 'Cargo', 'CBO',
                        'Referência', 'Data admissão', 'Pagamento', 'PIS', 'CPF', 'Salário Base']
            found = [col for col in expected if col in merged]
            if len(found) >= 3:
                df = pd.DataFrame([merged])
                return df

        except Exception as e:
            print(f"Page {page_num}: Table {t} (lines) failed: {e}")
            continue

    print(f"⚠️  Page {page_num}: No valid 'lines' table found.")
    return pd.DataFrame()

def extract_rows_5_and_6_using_lines_strict_strategy(pdf, page_num=1, table_num_start=1):
    """
    Extract rows 6‑7 using the 'lines_strict' strategy.
    """
    max_tables_to_try = 5

    for t in range(table_num_start, table_num_start + max_tables_to_try):
        try:
            if page_num < 1 or page_num > len(pdf.pages):
                raise ValueError(f"Page {page_num} out of range.")

            page = pdf.pages[page_num - 1]
            table_settings = {
                "vertical_strategy": "lines_strict",
                "horizontal_strategy": "lines_strict",
                "snap_tolerance": 3,
                "join_tolerance": 3,
            }
            tables = page.extract_tables(table_settings)
            if not tables:
                continue

            non_empty_tables = [tb for tb in tables if tb and len(tb) > 1]
            if t > len(non_empty_tables):
                continue

            table = non_empty_tables[t - 1]
            selected_rows = table[5:7]

            if len(selected_rows) < 2:
                continue

            headers = [clean_header(h) if h is not None else None for h in selected_rows[0]]
            data_row = selected_rows[1]

            raw_col_data = []
            for item in data_row:
                if item is not None:
                    parts = str(item).split('\n')
                    cleaned = [normalize_text(p) for p in parts if p.strip() != '']
                    raw_col_data.append(cleaned)
                else:
                    raw_col_data.append([])

            max_rows = max(len(col) for col in raw_col_data) if raw_col_data else 0
            if max_rows == 0:
                continue

            df_dict = {}
            for header, col_items in zip(headers, raw_col_data):
                if header is None:
                    continue
                if header == 'Descontos':
                    padded = [''] * (max_rows - len(col_items)) + col_items
                else:
                    padded = col_items + [''] * (max_rows - len(col_items))
                df_dict[header] = padded

            df = pd.DataFrame(df_dict)

            expected = ['Evento', 'Descrição', 'Índice', 'Proventos', 'Descontos']
            found = [col for col in expected if col in df.columns]
            if len(found) >= 3 and not df.empty:
                print(f"✓ Page {page_num}: Found valid 'lines_strict' table at index {t}")
                return df

        except Exception as e:
            print(f"Page {page_num}: Table {t} (lines_strict) failed: {e}")
            continue

    print(f"⚠️  Page {page_num}: No valid 'lines_strict' table found.")
    return pd.DataFrame()

def combine_page_data(rows_1_5: pd.DataFrame, rows_6_7: pd.DataFrame, remaining_data: pd.DataFrame, 
                      page_num: int | None = None, columns: list = None) -> pd.DataFrame:
    """Combine data for one page, repeat single-row parts, and enforce column order."""
    if rows_6_7.empty:
        combined = pd.concat(
            [rows_1_5.reset_index(drop=True), remaining_data.reset_index(drop=True)],
            axis=1,
        )
    else:
        combined = pd.concat(
            [
            rows_1_5.reset_index(drop=True),
            rows_6_7.reset_index(drop=True),
            remaining_data.reset_index(drop=True)
            ],
            axis=1
        )
        combined = combined.fillna('')
        print(f'{combined[["Descrição"]]}\n')
    if page_num is not None and not combined.empty:
        combined.insert(0, 'Page', page_num)

    if columns is not None:
        for col in columns:
            if col not in combined.columns:
                combined[col] = ''
        combined = combined[columns]

    return combined

def extract_pages_to_dataframe(pdf_path, page_numbers=None, table_num=1) -> pd.DataFrame:
    """Extract and combine data for multiple PDF pages with constant headers."""
    DESIRED_COLUMNS = [
        'Page', 'Nome', 'Matrícula', 'Departamento', 'Cargo', 'CBO', 'Referência',
        'Data admissão', 'Pagamento', 'PIS', 'CPF', 'Salário Base',
        'Evento', 'Descrição', 'Índice', 'Proventos', 'Descontos',
        'BASE INSS', 'BASE IRRF', 'BASE FGTS', 'Valor líquido', 'BASE INSS EMPRESA'     
    ]

    # OPEN PDF EXACTLY ONCE FOR PAYROLL
    with pdfplumber.open(pdf_path) as pdf:
        if page_numbers is None:
            page_numbers = list(range(1, len(pdf.pages) + 1))
        elif isinstance(page_numbers, int):
            page_numbers = [page_numbers]

        all_pages = []
        for _, page_num in enumerate(page_numbers):
            print(f"\n{'='*80}")
            print(f"PROCESSING PAGE {page_num}")
            print(f"{'='*80}")
            
            # Pass the opened `pdf` object down to the helpers
            rows_1_5 = extract_rows_1_to_5_using_lines_strategy(pdf, page_num=page_num)
            rows_6_7 = extract_rows_5_and_6_using_lines_strict_strategy(pdf, page_num=page_num)
            remaining_data = extract_lines_from_base_to_declaro(pdf, page_num=page_num)

            page_df = combine_page_data(rows_1_5, rows_6_7, remaining_data, page_num=page_num, columns=DESIRED_COLUMNS)
            if not page_df.empty:
                all_pages.append(page_df)

    if not all_pages:
        return pd.DataFrame(columns=DESIRED_COLUMNS)

    return pd.concat(all_pages, ignore_index=True)

def preview_table_data(pdf_path, page_num=1, table_num=1):
    """Preview a specific table for debugging"""
    with pdfplumber.open(pdf_path) as pdf:
        if page_num > len(pdf.pages):
            print(f"⚠️  Page {page_num} doesn't exist")
            return
        
        page = pdf.pages[page_num - 1]
        
        table_settings = {
            "vertical_strategy": "lines_strict",
            "horizontal_strategy": "lines_strict",
            "snap_tolerance": 3,
        }
        
        tables = page.extract_tables(table_settings)
        
        if not tables or table_num > len(tables):
            print(f"❌ Table {table_num} not found on page {page_num}")
            return
        
        table = tables[table_num - 1]
        
        print(f"\n📊 PREVIEW - Page {page_num}, Table {table_num}")
        print("="*80)
        
        if table:
            print(f"Dimensions: {len(table)} rows x {len(table[0]) if table else 0} columns")
            print("\nFull table:")
            print("-"*80)
            
            for row_idx, row in enumerate(table, 1):
                print(f"Row {row_idx:3d}: {row}")

#==================================================
# TIMESHEET PARSER (PDF --> EXCEL)  
# =================================================

def find_worker_boundaries(pdf):
    """
    Find all worker entries by treating each page that contains the timesheet
    marker as the START of a new worker's section. The section extends until
    the page before the next marker (or end of PDF).
    """
    workers = []
    marker = "Espelho de ponto"
    total_pages = len(pdf.pages)

    marker_pages = []
    for page_num, page in enumerate(pdf.pages, start=1):
        text = page.extract_text() or ""
        if marker in text:
            marker_pages.append((page_num, text))

    for i, (start_page, text) in enumerate(marker_pages):
        end_page = (marker_pages[i + 1][0] - 1) if i + 1 < len(marker_pages) else total_pages
        name = f"Worker_{start_page}"
        workers.append((name, start_page, end_page))

    return workers 
   
def convert_hours_to_decimal(hours_str):
    """Convert HH:MM format to decimal hours"""
    if ':' in hours_str:
        parts = hours_str.split(':')
        if len(parts) == 2:
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours + minutes / 60.0
    return float(hours_str)

def extract_worker_data(pdf, start_page, end_page):
    """
    Extract worker data using the text lines approach.
    Removed the unused `all_lines` extraction which was causing massive slowdowns.
    """
    worker_data = {}
    try:
        if start_page < 1 or end_page > len(pdf.pages) or start_page > end_page:
            raise ValueError(f"Invalid page range: {start_page}-{end_page}")

        first_page = pdf.pages[start_page - 1]
        text = first_page.extract_text() or ""
        personal_data = extract_personal_data(text)
        worker_data.update(personal_data)

        consolidation_data = extract_table_data(pdf, start_page, end_page)
        worker_data.update(consolidation_data)

        worker_data['page_start'] = start_page
        worker_data['page_end'] = end_page

        return worker_data
    except Exception as e:
        exc_type, exc_value, exc_tb = sys.exc_info()
        print(f"Error type: {exc_type.__name__}")
        print(f"Line number: {exc_tb.tb_lineno}")
        print(f"   ❌ Error processing worker: {e}")
        return {}

def extract_personal_data(text):
    """Extract personal information from text."""
    data = {}

    patterns = {
        'name': r'(?:Nome|Funcionário):\s*\d+\s*-\s*([A-Za-zÀ-ÿ\s]+)',
        'employee_id': r'(?:Matrícula|ID)[:\s]+(\d+)',
        'department': r'(?:Departamento|Setor)[:\s]+([A-Za-zÀ-ÿ\s]+)',
        'position': r'(?:Cargo|Função)[:\s]+([A-Za-zÀ-ÿ\s]+)',
        'cpf': r'CPF[:\s]+([\d.]+)',
    }

    for field, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data[field] = match.group(1).strip()

    month_match = re.search(r'Espelho de ponto de ([\w]+)', text)
    if month_match:
        data['month'] = month_match.group(1)

    name = data.get('name', '')
    name_match = re.search(r'[A-ZÀ-Ý][a-zà-ÿ]', name)
    if name_match:
        data['name'] = name[:name_match.start()].strip()
    else:
        data['name'] = name.strip()

    position = data.get('position', '')
    if position:
        parts = position.split()
        data['position'] = ' '.join(parts[:-1]) if len(parts) > 1 else position
    else:
        data['position'] = ''

    return data

def extract_table_data(pdf, start_page, end_page):
    """Extract data using strategy 'lines' directly"""
    target_mapping = {
        'Dias Trab.': 'actual_workdays',
        'H. Trabalhadas': 'worked_hours',
        'Dias de Trab. previstos': 'expected_workdays',
        'H. Previstas': 'expected_hours'
    }
    
    data = {}
    
    for page_num in range(start_page - 1, end_page):
        if page_num < len(pdf.pages):
            page = pdf.pages[page_num]
            tables = page.extract_tables()
            
            for table in tables:
                for row_idx, row in enumerate(table):
                    clean_row = [cell.strip() if cell else "" for cell in row]
                    row_text = " ".join(clean_row)
                    if all(target in row_text for target in target_mapping):
                        if row_idx + 1 < len(table):
                            next_row = [c.strip() if c else "" for c in table[row_idx + 1]]
                            for col_idx, header in enumerate(clean_row):
                                for target_header, dict_key in target_mapping.items():
                                    if target_header in header and col_idx < len(next_row):
                                        data[dict_key] = next_row[col_idx]
                                        
                        for cell in clean_row:
                            if '\n' in cell:
                                parts = cell.split('\n')
                                header_part = parts[0].strip()
                                val_part = parts[1].strip()
                                for target_header, dict_key in target_mapping.items():
                                    if target_header in header_part:
                                        data[dict_key] = val_part
                        break
    return data

def process_all_workers(pdf_path, output_excel_path, page_numbers=None):
    """
    Process all worker pages in the PDF and save to Excel.
    Opens the PDF exactly ONE time for maximum performance.
    """
    print("🔍 Finding worker boundaries...")
    
    # OPEN THE PDF EXACTLY ONCE FOR TIMESHEETS
    with pdfplumber.open(pdf_path) as pdf:
        
        if page_numbers is None:
            workers = find_worker_boundaries(pdf)
        else:
            if isinstance(page_numbers, int):
                page_numbers = [page_numbers]
            workers = [(f"Worker_{page_num}", page_num, page_num) for page_num in page_numbers]

        print(f"✅ Found {len(workers)} workers")

        all_workers_data = []

        for idx, (worker_name, start_page, end_page) in enumerate(workers, 1):
            print(f"\n📋 Processing worker {idx}: {worker_name} (Pages {start_page}-{end_page})")

            try:
                worker_data = extract_worker_data(pdf, start_page, end_page)

                if worker_data:
                    all_workers_data.append(worker_data)
                    print(f"   ✅ Extracted: {len(worker_data)} fields")
                else:
                    print("   ❌ No data extracted")

            except Exception as e:
                exc_type, exc_value, exc_tb = sys.exc_info()
                print(f"Error type: {exc_type.__name__}")
                print(f"Line number: {exc_tb.tb_lineno}")
                print(f"   ❌ Error processing worker {idx}: {e}")
                continue

    df = pd.DataFrame(all_workers_data)

    if not df.empty:
        priority_cols = ['name', 'employee_id', 'department', 'position',
                         'actual_workdays', 'expected_workdays',
                         'worked_hours', 'expected_hours',
                         'page_start', 'page_end']

        available_cols = [col for col in priority_cols if col in df.columns]
        other_cols = [col for col in df.columns if col not in priority_cols]
        df = df[available_cols + other_cols]

        df.to_excel(output_excel_path, index=False, engine='openpyxl')
        print(f"\n✅ Saved data for {len(df)} workers to {output_excel_path}")
    else:
        print("❌ No data extracted")

    return df

#==================================================
# Run the enhanced parser 
# =================================================
if __name__ == "__main__":
    os.system('cls' if os.name == 'nt' else 'clear')
    
    payroll_pdf_path = None#r"C:\Users\felipe.rsouza\Downloads\04._Contracheques_06.2026.pdf"
    timesheet_pdf_path = r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Teste001\account_MESP\G&E\10._Folhas_de_ponto_07.2026.pdf"
    excel_path = r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Teste001\account_MESP\G&E\Medição_G&E.xlsx"
    
    dataframes_to_save = {}
    
    # Extract payroll data
    if payroll_pdf_path:
        payroll_df = extract_pages_to_dataframe(payroll_pdf_path, page_numbers=None, table_num=1)
    else:
        payroll_df = pd.DataFrame()
    if not payroll_df.empty:
        dataframes_to_save['Payroll'] = payroll_df
        print(f"✅ Payroll data: {len(payroll_df)} rows extracted")
    else:
        print("❌ No payroll data extracted.")
    
    # Extract timesheet data
    timesheet_df = process_all_workers(pdf_path=timesheet_pdf_path, output_excel_path=excel_path)
    if not timesheet_df.empty:
        dataframes_to_save['Timesheet'] = timesheet_df
        print(f"✅ Timesheet data: {len(timesheet_df)} rows extracted")
    else:
        print("❌ No timesheet data extracted.")
    
    # Save all DataFrames to Excel with separate sheets
    if dataframes_to_save:
        save_to_excel(dataframes_to_save, excel_path)
        print(f"✅ Saved all data to {excel_path} with {len(dataframes_to_save)} sheets")
    else:
        print("❌ No data extracted to save.")