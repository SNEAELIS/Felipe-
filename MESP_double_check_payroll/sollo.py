import sys, os
import re
import json
import traceback

import pandas as pd
import pdfplumber

from datetime import datetime
from functools import wraps

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from collections import ChainMap

#==================================================
# PAYROLL PARSER (PDF --> EXCEL)  
# =================================================

def normalize_text(text):
    """Normalize text to UTF-8-safe str, remove R$ currency markers, and normalize whitespace."""
    if text is None:
        return ""
    if isinstance(text, bytes):
        text = text.decode("utf-8", errors="replace")

    text = str(text).strip()
    text = text.replace("R$", "").replace("r$", "")
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text

def debug_pdf(pdf_path, page_range=1):
    """
    Comprehensive debugger for PDF structure - Fixed version
    
    Args:
        pdf_path: Path to the PDF file
        page_range: Single page (int), range as tuple (start, end), or range as string "start-end"
                   Examples: 1, (1, 5), "1-5"
    """
    pages_to_debug = []
    if isinstance(page_range, int):
        pages_to_debug = [page_range]
    elif isinstance(page_range, tuple) and len(page_range) == 2:
        pages_to_debug = list(range(page_range[0], page_range[1] + 1))
    elif isinstance(page_range, str) and '-' in page_range:
        parts = page_range.split('-')
        if len(parts) == 2:
            try:
                start, end = int(parts[0].strip()), int(parts[1].strip())
                pages_to_debug = list(range(start, end + 1))
            except ValueError:
                print(f"❌ Invalid range format: {page_range}. Use 'start-end' format.")
                return
    else:
        print(f"❌ Invalid page_range: {page_range}. Use int, tuple (start, end), or string 'start-end'.")
        return
    
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        valid_pages = [p for p in pages_to_debug if 1 <= p <= total_pages]
        if not valid_pages:
            print(f"❌ No valid pages in range. PDF has {total_pages} pages.")
            return
        if len(valid_pages) < len(pages_to_debug):
            invalid = [p for p in pages_to_debug if p not in valid_pages]
            print(f"⚠️  Skipping invalid pages: {invalid}")
        
        for page_num in valid_pages:
            _debug_pdf_page(pdf, pdf_path, page_num, total_pages)

def _debug_pdf_page(pdf, pdf_path, page_num, total_pages):
    """Helper function to debug a single page"""
    print("="*80)
    print(f"🔍 PDF DEBUGGER - Analyzing: {pdf_path.split('\\')[-1]}")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Page {page_num}/{total_pages}")
    print("="*80)
    
    print(f"\n📄 DOCUMENT INFO:")
    print(f"   Total pages: {total_pages}")
    print(f"   Metadata: {pdf.metadata if pdf.metadata else 'No metadata'}")
    
    page = pdf.pages[page_num - 1]
    
    print(f"\n📃 PAGE {page_num} INFO:")
    print(f"   Width: {page.width}")
    print(f"   Height: {page.height}")
    print(f"   Crop box: {page.cropbox}")
    
    print(f"\n📝 RAW TEXT EXTRACTION:")
    print("-"*60)
    text = page.extract_text()
    if text:
        print(f"✓ Text extracted ({len(text)} characters, {len(text.split())} words)")
        print(f"\n📏 TEXT BY LINES:")
        print("-"*60)
        lines = page.extract_text_lines()
        if lines:
            print(f"✓ Found {len(lines)} lines")
            print("\nAll lines:")
            for i, line in enumerate(lines[:], 1):
                text_line = line.get('text', '').strip()
                if text_line:
                    print(f"   {i:3d}. {text_line[:100]}")
                    if len(text_line) > 100:
                        print(f"       ... (total {len(text_line)} chars)")
        else:
            print("❌ No lines found")
        
        print(f"\n🔤 WORDS WITH POSITIONS:")
        print("-"*60)
        words = page.extract_words()
        if words:
            print(f"✓ Found {len(words)} words")
            print("\nFirst 15 words with coordinates:")
            for i, word in enumerate(words[:15], 1):
                print(f"   {i:3d}. '{word['text']}' at (x:{word['x0']:.1f}, y:{word['top']:.1f})")
        else:
            print("❌ No words found")
        
        print(f"\n📊 TABLES DETECTION:")
        print("-"*60)
        
        strategies = ['lines', 'lines_strict']
        all_tables = []
        
        for strategy in strategies:
            try:
                print(f"\nTrying strategy: '{strategy}'")
                table_settings = {
                    "vertical_strategy": strategy, 
                    "horizontal_strategy": strategy,
                    "snap_tolerance": 4,
                }
                tables = page.extract_tables(table_settings)
                
                if not tables:
                    print(f"   ❌ No tables found with '{strategy}' strategy")
                    continue
                
                non_empty_tables = []
                for t in tables:
                    if t and len(t) > 1:
                        non_empty_tables.append(t)
                
                if not non_empty_tables:
                    print(f"   ℹ️  Found {len(tables)} table(s), but all are empty or only headers")
                    continue
                
                print(f"   ✓ Found {len(non_empty_tables)} non‑empty table(s)")
                
                for idx, table in enumerate(non_empty_tables, start=1):
                    rows = len(table)
                    cols = len(table[0]) if table else 0
                    print(f"\n   Table {idx} (strategy '{strategy}'):")
                    print(f"      Dimensions: {rows} rows × {cols} columns")
                    
                    preview_limit = min(rows,1000000000)
                    print(f"      Preview (first {preview_limit} rows):")
                    for r in range(preview_limit):
                        row_display = [str(cell) if cell is not None else 'None' for cell in table[r]]
                        print(f"         Row {r+1}: {row_display}")
                    
                    if rows > 3:
                        print(f"         ... and {rows-3} more rows")
                    
                    all_tables.append((strategy, idx, table))
                    
            except Exception as e:
                print(f"   ⚠️  Error with '{strategy}': {e}")
        
        if all_tables:
            print("\n" + "="*60)
            print("📋 SUMMARY OF ALL TABLES FOUND:")
            print("-"*60)
            for strategy, idx, table in all_tables:
                rows = len(table)
                cols = len(table[0]) if table else 0
                print(f"   • Strategy '{strategy}' → Table {idx}: {rows}×{cols}")
            
            print("\n💡 To extract a particular table, use:")
            print("   extract_table_rows_by_strategy(pdf_path, page_num=page, table_num=<index>, strategy='<strategy>')")
        else:
            print("\n❌ No tables detected with any strategy")
            print("   💡 Try these alternative settings:")
            print("      - Use 'text' strategy with custom tolerances")
            print("      - Use explicit table boundaries if known")
            print("      - Check if content is truly tabular")
        
        print(f"\n🔍 SEARCH FOR COMMON PATTERNS:")
        print("-"*60)
        if text:
            patterns = {
                'Dates (dd/mm/yyyy)': r'\d{2}/\d{2}/\d{4}',
                'Dates (mm/dd/yyyy)': r'\d{2}/\d{2}/\d{4}',
                'Dates (dd-mm-yyyy)': r'\d{2}-\d{2}-\d{4}',
                'Numbers with decimals': r'\d+\.\d{2}',
                'Email addresses': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
                'Phone numbers': r'\(?\d{2,3}\)?[\s.-]?\d{4,5}[\s.-]?\d{4}',
                'CPF (Brazil)': r'\d{3}\.\d{3}\.\d{3}-\d{2}',
                'CNPJ (Brazil)': r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}',
                'Currency (R$)': r'R\$\s*\d+[\.,]\d{2}',
            }
            
            found_patterns = []
            for name, pattern in patterns.items():
                matches = re.findall(pattern, text)
                if matches:
                    found_patterns.append((name, matches[:5]))
                    print(f"✓ {name}: Found {len(matches)} matches")
                    print(f"   Examples: {matches[:3]}")
            
            if not found_patterns:
                print("ℹ️  No common patterns found in text")
        else:
            print("⚠️  No text to search for patterns")
        
        print(f"\n🖼️  IMAGES:")
        print("-"*60)
        images = page.images
        if images:
            print(f"✓ Found {len(images)} images:")
            for i, img in enumerate(images[:5], 1):
                print(f"   Image {i}: {img.get('width', 'N/A')}x{img.get('height', 'N/A')}")
        else:
            print("ℹ️  No images found")
        
        print("\n" + "="*80)
        print("📋 EXTRACTION SUMMARY:")
        print("="*80)
        print(f"✓ Text extracted: {'✅ Yes' if text else '❌ No'}")
        print(f"✓ Tables found: {'✅ Yes' if all_tables else '❌ No'}")
        print(f"✓ Images found: {'✅ Yes' if images else '❌ No'}")
        print(f"✓ Total words: {len(words) if words else 0}")
        print(f"✓ Total characters: {len(text) if text else 0}")
        
        print("\n💡 RECOMMENDATIONS:")
        if not text:
            print("   ⚠️  No text extracted → Page is likely scanned images. Need OCR!")
        elif all_tables:
            print("   📊 Tables detected! Use page.extract_tables() with custom settings")
        else:
            print("   📝 Text-only page → Use page.extract_text() and parse manually")
        
        if text:
            print("\n📌 NEXT STEPS SUGGESTION:")
            print("   Based on the content, you might want to:")
            if any(patterns in str(found_patterns) for patterns in ['Dates', 'Currency']):
                print("   - Extract specific data using regex patterns")
            if all_tables:
                print("   - Extract tables into structured data")
            print("   - Split text by sections or delimiters")
            print("   - Create rules to identify and extract key information")
        
        print("\n" + "="*80)
        return {
            'has_text': bool(text),
            'has_tables': bool(all_tables),
            'word_count': len(words) if words else 0,
            'char_count': len(text) if text else 0,
            'page_height': page.height,
            'page_width': page.width
        }

def parse_payroll_currency(value):
    """Normalize currency values to a decimal string like 5000.00."""
    if value is None:
        return ""

    text = normalize_text(value)
    if not text:
        return ""

    match = re.search(r"[-+]?\d+(?:[.,]\d+)?", text)
    if not match:
        return ""

    number_text = match.group(0).replace(".", "").replace(",", ".")
    return f"{float(number_text):.2f}"


def extract_labeled_value(text, label):
    """Extract a numeric value after a label even when the label appears mid-line."""
    pattern = re.compile(rf"{re.escape(label)}\s*:\s*([+-]?\d+(?:[.,]\d+)?)", re.IGNORECASE)
    match = pattern.search(text)
    if match:
        return parse_payroll_currency(match.group(1))

    fallback_pattern = re.compile(rf"{re.escape(label)}\s*([+-]?\d+(?:[.,]\d+)?)", re.IGNORECASE)
    match = fallback_pattern.search(text)
    if match:
        return parse_payroll_currency(match.group(1))

    return ""


def extract_payroll_rows_from_text_lines(lines):
    """Create a worker-level table from PDF lines using the Func marker as a boundary."""
    rows = []
    current_worker = None

    for entry in lines:
        text = entry.get("text", "") if isinstance(entry, dict) else str(entry)
        normalized_text = normalize_text(text)
        if not normalized_text:
            continue

        func_match = re.search(r"Func:\s*(.+)", normalized_text, re.IGNORECASE)
        if func_match:
            if current_worker is not None:
                rows.append(current_worker)

            raw_text = normalize_text(func_match.group(1))
            sep_match = re.match(r"^(\d+)\s*(.+)$", raw_text)

            if sep_match:
                enroll_number = sep_match.group(1)
                clean_name = sep_match.group(2).strip()
                current_worker = {
                    "MATRICULA": enroll_number,
                    "NOME": clean_name
                }
                name = current_worker.get('NOME', '')
                name_match = re.search(r'[A-ZÀ-Ý][a-zà-ÿ]', name)
                if name_match:
                    current_worker['NOME'] = name[:name_match.start()].strip()

            else:
                current_worker = {"NOME": raw_text}

            continue

        if current_worker is None:
            continue

        upper_text = normalized_text.upper()

        if "PROVENTOS" in upper_text and not current_worker.get("PROVENTOS"):
            current_worker["PROVENTOS"] = extract_labeled_value(normalized_text, "PROVENTOS")

        if ("SALÁRIO" in upper_text or "SALARIO" in upper_text) and not current_worker.get("SALÁRIO BASE"):
            current_worker["SALÁRIO BASE"] = (
                extract_labeled_value(normalized_text, "SALÁRIO")
                or extract_labeled_value(normalized_text, "SALARIO")
            )

        if ("SITUAÇÃO" in upper_text or "SITUACAO" in upper_text) and not current_worker.get("SITUAÇÃO"):
            match = re.search(r"(?:SITUAÇÃO|SITUACAO)\s*:\s*([^\s]+(?:\s+[^\s]+)*)$", normalized_text, re.IGNORECASE)
            if match:
                current_worker["SITUAÇÃO"] = match.group(1).strip()
            else:
                match = re.search(r"(?:SITUAÇÃO|SITUACAO)\s*:\s*(.+)", normalized_text, re.IGNORECASE)
                if match:
                    current_worker["SITUAÇÃO"] = match.group(1).strip()

    if current_worker is not None:
        rows.append(current_worker)

    columns = ["MATRICULA", "NOME", "PROVENTOS", "SALÁRIO BASE", "SITUAÇÃO"]
    df = pd.DataFrame(rows, columns=columns)
    df = df.drop_duplicates(subset=["MATRICULA", "NOME"], keep="first")
    return df.fillna("")


def extract_payroll_rows_from_pdf(pdf_path, page_numbers=None):
    """Extract worker rows from a PDF using the Func marker as the separation point."""
    with pdfplumber.open(pdf_path) as pdf:
        if page_numbers is None:
            page_numbers = list(range(1, len(pdf.pages) + 1))
        elif isinstance(page_numbers, int):
            page_numbers = [page_numbers]

        frames = []
        for page_num in page_numbers:
            if page_num < 1 or page_num > len(pdf.pages):
                continue

            page = pdf.pages[page_num - 1]
            lines = page.extract_text_lines() or []
            if not lines:
                continue

            page_df = extract_payroll_rows_from_text_lines(lines)
            if not page_df.empty:
                frames.append(page_df)

    if not frames:
        return pd.DataFrame(columns=["NOME", "PROVENTOS", "SALÁRIO BASE"])

    return pd.concat(frames, ignore_index=True)

def save_to_excel(data, excel_path):
    """
    Save DataFrame(s) to Excel with formatting
    """
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        if isinstance(data, dict):
            for sheet_name, df in data.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    format_sheet(writer, sheet_name)
        else:
            df = data
            if not df.empty:
                df.to_excel(writer, sheet_name='Data', index=False)
                format_sheet(writer, 'Data')

def format_sheet(writer, sheet_name):
    """Apply formatting to a specific sheet in the workbook"""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    if worksheet.max_row > 0:
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

#==================================================
# TIMESHEET PARSER (PDF --> EXCEL)  
# =================================================

def find_worker_boundaries(pdf):
    """
    Find all worker entries by treating each page that contains the timesheet
    marker as the START of a new worker's section. The section extends until
    the page before the next marker (or end of PDF).
    """
    workers = []
    marker = "CARTÃO PONTO"
    total_pages = len(pdf.pages)

    marker_pages = []
    for page_num, page in enumerate(pdf.pages, start=1):
        text = page.extract_text() or ""
        if marker in text:
            marker_pages.append((page_num, text))

    for i, (start_page, text) in enumerate(marker_pages):
        end_page = (marker_pages[i + 1][0] - 1) if i + 1 < len(marker_pages) else total_pages
        name = f"Worker_{start_page}"
        workers.append((name, start_page, end_page))

    return workers 
   
def convert_hours_to_decimal(hours_str):
    """Convert HH:MM format to decimal hours"""
    if ':' in hours_str:
        parts = hours_str.split(':')
        if len(parts) == 2:
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours + minutes / 60.0
    return float(hours_str)

def extract_worker_data(pdf, start_page, end_page):
    """
    Extract worker data using the text lines approach.
    Removed the unused `all_lines` extraction which was causing massive slowdowns.
    """
    worker_data = {}
    try:
        if start_page < 1 or end_page > len(pdf.pages) or start_page > end_page:
            raise ValueError(f"Invalid page range: {start_page}-{end_page}")

        first_page = pdf.pages[start_page - 1]
        text = first_page.extract_text() or ""
        personal_data = extract_personal_data(text)
        worker_data.update(personal_data)

        consolidation_data = extract_table_data(pdf, start_page, end_page)
        worker_data.update(consolidation_data)

        worker_data['page_start'] = start_page
        worker_data['page_end'] = end_page

        return worker_data
    except Exception as e:
        exc_type, exc_value, exc_tb = sys.exc_info()
        print(f"Error type: {exc_type.__name__}")
        print(f"Line number: {exc_tb.tb_lineno}")
        print(f"   ❌ Error processing worker: {e}")
        return {}

def extract_personal_data(text):
    """Extract personal information and summary values from the timesheet text."""
    data = {}

    name_match = re.search(r'(?:Nome|Funcionário|Funcionária)\s*[:\-]?\s*(.+)', text, re.IGNORECASE)
    if name_match:
        data['Nome'] = normalize_text(name_match.group(1))

    employee_id_match = re.search(r'(?:Matrícula|Matricula|ID)\s*[:\-]?\s*(\d+)', text, re.IGNORECASE)
    if employee_id_match:
        data['Matrícula'] = employee_id_match.group(1).strip()

    days_worked_match = re.search(r'(?:Dias trabalhados|Dias Trab\.|Dias de Trab\.|Dias de Trabalho)\s*[:\-]?\s*(\d+)', text, re.IGNORECASE)
    if days_worked_match:
        data['Dias trabalhados'] = days_worked_match.group(1).strip()

    absences_match = re.search(r'(?:Faltas|Ausências|Ausencias)\s*[:\-]?\s*(\d+)', text, re.IGNORECASE)
    if absences_match:
        data['Faltas'] = absences_match.group(1).strip()

    if 'Nome' not in data:
        data['Nome'] = ''
    if 'Matrícula' not in data:
        data['Matrícula'] = ''
    if 'Dias trabalhados' not in data:
        data['Dias trabalhados'] = ''
    if 'Faltas' not in data:
        data['Faltas'] = ''

    return data

def extract_table_data(pdf, start_page, end_page):
    """Extract data using strategy 'lines' directly"""
    target_mapping = {
        'Dias Trab.': 'actual_workdays',
        'H. Trabalhadas': 'worked_hours',
        'Dias de Trab. previstos': 'expected_workdays',
        'H. Previstas': 'expected_hours'
    }
    
    data = {}
    
    for page_num in range(start_page - 1, end_page):
        if page_num < len(pdf.pages):
            page = pdf.pages[page_num]
            tables = page.extract_tables()
            
            for table in tables:
                for row_idx, row in enumerate(table):
                    clean_row = [cell.strip() if cell else "" for cell in row]
                    row_text = " ".join(clean_row)
                    if all(target in row_text for target in target_mapping):
                        if row_idx + 1 < len(table):
                            next_row = [c.strip() if c else "" for c in table[row_idx + 1]]
                            for col_idx, header in enumerate(clean_row):
                                for target_header, dict_key in target_mapping.items():
                                    if target_header in header and col_idx < len(next_row):
                                        data[dict_key] = next_row[col_idx]
                                        
                        for cell in clean_row:
                            if '\n' in cell:
                                parts = cell.split('\n')
                                header_part = parts[0].strip()
                                val_part = parts[1].strip()
                                for target_header, dict_key in target_mapping.items():
                                    if target_header in header_part:
                                        data[dict_key] = val_part
                        break
    return data

def save_payroll_and_timesheet_to_excel(payroll_df, timesheet_df, excel_path):
    """Save both parser outputs to the same workbook with separate sheets."""
    dataframes_to_save = {}
    if not payroll_df.empty:
        dataframes_to_save['Payroll'] = payroll_df
    if not timesheet_df.empty:
        dataframes_to_save['Timesheet'] = timesheet_df

    if dataframes_to_save:
        save_to_excel(dataframes_to_save, excel_path)
        print(f"✅ Saved {len(dataframes_to_save)} sheets to {excel_path}")
    else:
        print("❌ No data extracted to save.")


def process_all_workers(pdf_path, output_excel_path, page_numbers=None):
    """
    Process all worker pages in the PDF and save to Excel.
    Opens the PDF exactly ONE time for maximum performance.
    """
    print("🔍 Finding worker boundaries...")
    
    # OPEN THE PDF EXACTLY ONCE FOR TIMESHEETS
    with pdfplumber.open(pdf_path) as pdf:
        
        if page_numbers is None:
            workers = find_worker_boundaries(pdf)
        else:
            if isinstance(page_numbers, int):
                page_numbers = [page_numbers]
            workers = [(f"Worker_{page_num}", page_num, page_num) for page_num in page_numbers]

        print(f"✅ Found {len(workers)} workers")

        all_workers_data = []

        for idx, (worker_name, start_page, end_page) in enumerate(workers, 1):
            print(f"\n📋 Processing worker {idx}: {worker_name} (Pages {start_page}-{end_page})")

            try:
                worker_data = extract_worker_data(pdf, start_page, end_page)

                if worker_data:
                    all_workers_data.append(worker_data)
                    print(f"   ✅ Extracted: {len(worker_data)} fields")
                else:
                    print("   ❌ No data extracted")

            except Exception as e:
                exc_type, exc_value, exc_tb = sys.exc_info()
                print(f"Error type: {exc_type.__name__}")
                print(f"Line number: {exc_tb.tb_lineno}")
                print(f"   ❌ Error processing worker {idx}: {e}")
                continue

    df = pd.DataFrame(all_workers_data)

    if not df.empty:
        priority_cols = ['Nome', 'Matrícula', 'Dias trabalhados', 'Faltas']

        available_cols = [col for col in priority_cols if col in df.columns]
        other_cols = [col for col in df.columns if col not in priority_cols]
        df = df[available_cols + other_cols]

        df.to_excel(output_excel_path, index=False, engine='openpyxl')
        print(f"\n✅ Saved data for {len(df)} workers to {output_excel_path}")
    else:
        print("❌ No data extracted")

    return df

#==================================================
# Run the enhanced parser 
# =================================================
if __name__ == "__main__":
    os.system('cls' if os.name == 'nt' else 'clear')

    payroll_pdf_path = r"C:\Users\felipe.rsouza\Downloads\4_Folha_de_Pagamento.pdf"
    timesheet_pdf_path = r"C:\Users\felipe.rsouza\Downloads\3_Folha_de_ponto.pdf"
    excel_path = r"C:\Users\felipe.rsouza\Downloads\Medição_SOLLO.xlsx"

    payroll_df = extract_payroll_rows_from_pdf(pdf_path=payroll_pdf_path, page_numbers=None)
    timesheet_df = process_all_workers(pdf_path=timesheet_pdf_path, output_excel_path=excel_path)

    save_payroll_and_timesheet_to_excel(payroll_df, timesheet_df, excel_path)