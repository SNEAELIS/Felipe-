import os
import re
import sys
import shutil
import warnings
from datetime import datetime
from dataclasses import dataclass, fields
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

warnings.filterwarnings(
    "ignore",
    message=r"^Workbook contains no default style, apply openpyxl's default$",
    category=UserWarning,
)

# ==================================================
# Classe com lógica de negócios para cálculos orçamentários
# ==================================================
class BudgetCalculator:
    """Business logic for budget calculations"""
    
    def __init__(self):
        self.changes_detected = False
    
    def _to_float(self, value) -> float:
        import math
        if value is None:
            return 0.0
        if isinstance(value, str) and value.strip() == '':
            return 0.0
        try:
            val = float(value)
            return 0.0 if math.isnan(val) else val
        except (ValueError, TypeError):
            return 0.0
    
    def _check_change(self, calculated: float, comp_data: float, item: str) -> float:
        comp_float = self._to_float(comp_data)
        calc_float = self._to_float(calculated)

        if (calc_float is None) and (comp_float is not None and comp_float > 0.0):
            print(f'Calculado inválido para item {item} ({calculated}); mantendo comparação: {comp_float}')
            return comp_float

        elif calc_float == comp_float:
            print(f'Dados do item {item} sem alteração')
            return comp_float
        
        elif calc_float != comp_float:
            print(f'‼️ Dados do item {item} alterados: {comp_float} -> {calc_float}')
            self.changes_detected = True
            return calc_float
    
    def item_a(self, comp_data, args: list) -> float:
        safe_args = [self._to_float(arg) for arg in args]
        result = sum(safe_args)
        return self._check_change(result, comp_data, item='A')
    
    def item_b(self, comp_data, credito_disp) -> float:
        credito_disp = self._to_float(credito_disp)
        return self._check_change(credito_disp, comp_data, item='B')
    
    def item_c(self, comp_data, cdto_indisponivel, dsza_p_empenhada) -> float:
        if self._to_float(dsza_p_empenhada) != 0:
            result = self._to_float(cdto_indisponivel) - self._to_float(dsza_p_empenhada)
            return self._check_change(result, comp_data, item='C')
        else:
            result = self._to_float(cdto_indisponivel)
            return self._check_change(result, comp_data, item='C')
        
    def item_d(self, comp_data, dspz_p_empenhada) -> float:
        result = self._to_float(dspz_p_empenhada)        
        return self._check_change(result, comp_data, item='D')
    
    def item_e(self, comp_data, dstq_concedido, dspz_empenhada) -> float:
        result = self._to_float(dstq_concedido) + self._to_float(dspz_empenhada)
        return self._check_change(result, comp_data, item='E')
    
    def item_f(self, comp_data, dspz_pagas) -> float:
        dspz_pagas = self._to_float(dspz_pagas)
        return self._check_change(dspz_pagas, comp_data, item='F')

# ==================================================
# Cria o filtro na planilha final
# ==================================================
def create_filter_column(df):
    def process_row(row):
        try:
            col0 = re.sub(r'\.0$', '', str(row.iloc[0])).strip() if 'nan' not in str(row.iloc[0]).lower() else ''
        except Exception as e:
            print(f"Aviso: erro ao ler coluna 0 da planilha de comparação ({type(e).__name__}: {e}) — linha ignorada.")
            col0 = ''
        try:
            col1_raw = re.sub(r'\.0$', '', str(row.iloc[1])).strip()
            col1 = re.sub(r'\s+', '', col1_raw.split('-')[0]) if '-' in col1_raw else col1_raw
            col1 = col1 if 'nan' not in col1.lower() else ''
        except Exception as e:
            print(f"Aviso: erro ao ler coluna 1 da planilha de comparação ({type(e).__name__}: {e}) — linha ignorada.")
            col1 = ''
        try:
            col2_raw = re.sub(r'\.0$', '', str(row.iloc[2])).strip()
            col2 = col2_raw[1] if len(col2_raw) > 1 else col2_raw[:-1]
            col2 = col2 if 'nan' not in col2.lower() else ''
        except Exception as e:
            print(f"Aviso: erro ao ler coluna 2 da planilha de comparação ({type(e).__name__}: {e}) — linha ignorada.")
            col2 = ''
        return col0 + col1 + col2
    
    return df.iloc[:, [1, 2, 3]].apply(process_row, axis=1)

@dataclass
class BudgetData:
    dotacao_atual: Optional[float] = None 
    dstq_concedido: Optional[float] = None 
    credito_disp: Optional[float] = None 
    cred_indisp: Optional[float] = None 
    dspz_p_empenhada: Optional[float] = None 
    dsza_empenhada: Optional[float] = None 
    dspz_pagas: Optional[float] = None

@dataclass
class ComparisonData:
    ploa: Optional[float] = None
    dotacao_atual: Optional[float] = None
    dotacao_disponivel: Optional[float] = None
    contingenciado: Optional[float] = None
    pre_empenhado: Optional[float] = None
    empenhado: Optional[float] = None
    pago: Optional[float] = None

def get_destination_with_today_date(base_dir: str) -> Path:
    """Create destination path with today's date in the given directory"""
    directory = Path(base_dir)
    today_date = datetime.now().strftime("%d.%m.%y")
    dest_path = directory / f"Execução Orçamento {today_date}.xlsx"
    counter = 1
    original_path = dest_path
    while dest_path.exists():
        dest_path = directory / f"Execução Orçamento {today_date} ({counter}).xlsx"
        counter += 1
    return dest_path

def save_updates_without_breaking_formulas(dest_path: str, exact_columns: dict, df_final_clean: pd.DataFrame, output_dir: str) -> tuple[Path, int]:
    """Save updates to Excel while preserving all formulas. Returns (saved_file_path, number_of_updates)."""
    # Cria o arquivo de destino no diretório de saída desejado
    file_path = get_destination_with_today_date(base_dir=output_dir)
    sheet_name = 'Dados'
    
    # Copia o arquivo modelo (que contém fórmulas e formatações)
    shutil.copy(dest_path, file_path)
    
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    # Mapeia colunas pelos cabeçalhos
    header_row = 3
    col_indices = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value:
            cell_str = str(cell_value)
            if 'Dotação Atual' in cell_str and '(A)' in cell_str:
                col_indices['A'] = col_idx
            elif 'Dotação Disponível' in cell_str and '(B)' in cell_str:
                col_indices['B'] = col_idx
            elif 'Contingenciado/Bloqueado SOF' in cell_str:
                col_indices['C'] = col_idx
            elif 'Pré-Empenhado' in cell_str and '(D)' in cell_str:
                col_indices['D'] = col_idx
            elif 'Empenhado/Descentralizado' in cell_str and '(E)' in cell_str:
                col_indices['E'] = col_idx
            elif 'Pago' in cell_str and '(F)' in cell_str:
                col_indices['F'] = col_idx  
    
    # Mapeamento filter_col -> linha do Excel
    row_to_filter = {}
    for excel_row in range(header_row + 1, ws.max_row + 1):
        try:
            col0 = ws.cell(row=excel_row, column=2).value
            col1 = ws.cell(row=excel_row, column=3).value
            col2 = ws.cell(row=excel_row, column=4).value
            
            col0_str = re.sub(r'\.0$', '', str(col0)).strip() if col0 and 'nan' not in str(col0).lower() else ''
            col1_raw = re.sub(r'\.0$', '', str(col1)).strip() if col1 else ''
            col1_clean = re.sub(r'\s+', '', col1_raw.split('-')[0]) if '-' in col1_raw else col1_raw
            col1_clean = col1_clean if 'nan' not in col1_clean.lower() else ''
            col2_raw = re.sub(r'\.0$', '', str(col2)).strip() if col2 else ''
            col2_clean = col2_raw[1] if len(col2_raw) > 1 else col2_raw[:-1]
            col2_clean = col2_clean if 'nan' not in col2_clean.lower() else ''
            
            filter_value = col0_str + col1_clean + col2_clean
            if filter_value and len(filter_value) == 6:
                row_to_filter[excel_row] = filter_value
        except:
            continue
    
    updates_count = 0
    for excel_row, filter_value in row_to_filter.items():
        matching_rows = df_final_clean[df_final_clean['filter_col'] == filter_value]
        if matching_rows.empty:
            continue
        row_data = matching_rows.iloc[0]
        for col_letter, col_idx in col_indices.items():
            df_col_name = exact_columns[col_letter]
            if df_col_name in row_data and pd.notna(row_data[df_col_name]):
                new_value = row_data[df_col_name]
                current_value = ws.cell(row=excel_row, column=col_idx).value
                if current_value != new_value:
                    ws.cell(row=excel_row, column=col_idx, value=new_value)
                    updates_count += 1
    
    wb.save(file_path)
    print(f"\n✅ {updates_count} cells updated")
    print(f"✅ File saved: {file_path}")
    return file_path, updates_count

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Atualizador de Orçamento - CGOFC")
        self.geometry("800x550")
        self.minsize(700, 450)
        self.resizable(True, True)
        
        # Variáveis para caminhos dos arquivos
        self.source_path = tk.StringVar()
        self.dest_path = tk.StringVar()
        
        self.create_widgets()
    
    def create_widgets(self):
        # Frame para seleção de arquivos
        file_frame = tk.LabelFrame(self, text="Seleção de Arquivos", padx=10, pady=10)
        file_frame.pack(fill="x", padx=10, pady=5)
        file_frame.columnconfigure(1, weight=1)  # Entry expands; button column stays fixed
        
        # Arquivo fonte
        tk.Label(file_frame, text="Arquivo fonte:").grid(row=0, column=0, sticky="w")
        tk.Entry(file_frame, textvariable=self.source_path).grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        tk.Button(file_frame, text="Procurar", command=self.browse_source).grid(row=0, column=2, padx=5)
        tk.Label(file_frame, text="Exemplo: 1.1 - MESP-Ano 2025-UG 180077 - Por resultado primário - C*.xlsx", 
                 fg="gray", font=("Arial", 8)).grid(row=1, column=1, sticky="w", padx=5)
        
        # Arquivo para comparação
        tk.Label(file_frame, text="Arquivo para comparação:").grid(row=2, column=0, sticky="w", pady=(10,0))
        tk.Entry(file_frame, textvariable=self.dest_path).grid(row=2, column=1, sticky="ew", padx=5, pady=(10,0))
        tk.Button(file_frame, text="Procurar", command=self.browse_dest).grid(row=2, column=2, padx=5, pady=(10,0))
        tk.Label(file_frame, text="Exemplo: Execução Orçamento *.xlsx", 
                 fg="gray", font=("Arial", 8)).grid(row=3, column=1, sticky="w", padx=5)
        
        # Frame de botões de ação
        action_frame = tk.Frame(self)
        action_frame.pack(fill="x", padx=10, pady=5)
        tk.Button(action_frame, text="Processar", command=self.process, bg="#4CAF50", fg="white", height=2, width=15).pack(side="left", padx=5)
        tk.Button(action_frame, text="Fechar", command=self.destroy, bg="#f44336", fg="white", height=2, width=15).pack(side="right", padx=5)
        
        # Área de log (saída de prints)
        log_frame = tk.LabelFrame(self, text="Log de Processamento", padx=5, pady=5)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, width=80, height=15)
        self.log_text.pack(fill="both", expand=True)
        
        # Redirecionar stdout para o widget de log
        sys.stdout = TextRedirector(self.log_text)
    
    def browse_source(self):
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo fonte",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if filename:
            self.source_path.set(filename)
    
    def browse_dest(self):
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo de comparação (modelo)",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if filename:
            self.dest_path.set(filename)
    
    def process(self):
        source = self.source_path.get().strip()
        dest = self.dest_path.get().strip()
        
        if not source or not os.path.exists(source):
            messagebox.showerror("Erro", "Selecione um arquivo fonte válido.")
            return
        if not dest or not os.path.exists(dest):
            messagebox.showerror("Erro", "Selecione um arquivo de comparação válido.")
            return
        
        # Desabilitar botões durante o processamento
        for widget in self.winfo_children():
            if isinstance(widget, tk.Button):
                widget.config(state="disabled")
        self.update()
        
        try:
            # Executa o processamento principal
            saved_file, updates = self.run_processing(source, dest)
            messagebox.showinfo("Processamento Concluído", 
                                f"Arquivo salvo em:\n{saved_file}\n\nNúmero de células alteradas: {updates}")
        except Exception as e:
            messagebox.showerror("Erro no Processamento", str(e))
        finally:
            for widget in self.winfo_children():
                if isinstance(widget, tk.Button):
                    widget.config(state="normal")
    
    def run_processing(self, source_path_str, dest_path_str):
        """Lógica principal adaptada para GUI"""
        source_path = Path(source_path_str)
        dest_path = Path(dest_path_str)
        
        # Carregar dados
        df_data_source = pd.read_excel(source_path, header=2)
        df_data_final = pd.read_excel(dest_path, sheet_name='Dados', header=2)
        
        # Mapeamento das colunas no arquivo fonte
        column_mapping = {
            13: 'DOTACAO ATUALIZADA',
            18: 'DESTAQUE CONCEDIDO',
            19: 'CREDITO DISPONIVEL',
            20: 'CREDITO INDISPONIVEL',
            22: 'DESPESAS PRE-EMPENHADAS A EMPENHAR',
            23: 'DESPESAS EMPENHADAS',
            28: 'DESPESAS PAGAS'
        }
        df_data_source_clean = df_data_source.copy()
        df_data_source_clean.rename(columns=column_mapping, inplace=True)
        
        # Criar coluna de filtro
        df_data_source_clean['filter_col'] = df_data_source_clean.iloc[:, [0, 2, 5]].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()
        
        df_final_clean = df_data_final.copy()
        df_final_clean['filter_col'] = create_filter_column(df_data_final)
        
        filter_final_df_Set = set(df_final_clean['filter_col'])
        
        print(f"DEBUG: Unique filters in Destiny: {len(filter_final_df_Set)}")
        print(f"DEBUG: Unique filters in Source: {len(set(df_data_source_clean['filter_col']))}")
        
        source_columns = [
            'DOTACAO ATUALIZADA',
            'DESTAQUE CONCEDIDO',
            'CREDITO DISPONIVEL',
            'CREDITO INDISPONIVEL',
            'DESPESAS PRE-EMPENHADAS A EMPENHAR',
            'DESPESAS EMPENHADAS',
            'DESPESAS PAGAS'
        ]
        
        # Colunas exatas no arquivo final (com quebras de linha)
        exact_columns = {
            'A': 'Dotação Atual\n(A)',
            'B': ' Dotação Disponível          \n(B)',
            'C': 'Contingenciado/Bloqueado SOF                            \n( C )',
            'D': 'Pré-Empenhado              (D)',
            'E': 'Empenhado/Descentralizado \n(E)',
            'F': 'Pago              \n (F)'
        }
        
        for filter_value in set(df_data_source_clean['filter_col']):
            if len(filter_value) != 6 or 'total' in filter_value.lower() or filter_value in ['', 'nan', 'none', None] or filter_value not in filter_final_df_Set:
                if len(filter_value) == 6 and 'total' not in filter_value.lower() and filter_value not in ['', 'nan', 'none', None] and filter_value not in filter_final_df_Set:
                    print(f"DEBUG SKIP: Filter '{filter_value}' exists in Source but NOT in Destiny.")
                continue
            
            chunk_source = df_data_source_clean[df_data_source_clean['filter_col'] == filter_value].copy()
            chunk_final = df_final_clean[df_final_clean['filter_col'] == filter_value].copy()
            
            print(f"Processando: {filter_value} ({len(chunk_source)} linha(s))...")
            
            field_names = [f.name for f in fields(ComparisonData)]
            values = chunk_final.iloc[:, 5:12]   # colunas B-F (índices 5 a 11)
            comp_data = ComparisonData(**dict(zip(field_names, values.iloc[0])))
            
            # Condensar linhas da fonte (soma)
            condensed_row = {}
            for col in chunk_source.columns[9:20]:  # colunas 9 a 19 (0-index)
                col_values = chunk_source[col].fillna(0.0)
                if len(col_values) == 0:
                    condensed_row[col] = None
                    continue
                numeric_values = []
                can_sum = True
                for val in col_values:
                    if isinstance(val, (int, float)):
                        numeric_values.append(val)
                    elif isinstance(val, str):
                        try:
                            numeric_values.append(float(val))
                        except ValueError:
                            can_sum = False
                            break
                    else:
                        can_sum = False
                        break
                if can_sum and len(numeric_values) > 0:
                    condensed_row[col] = sum(numeric_values)
                else:
                    condensed_row[col] = col_values.iloc[0]
            
            # Se houver apenas uma linha, usar valores diretamente
            if len(chunk_source) == 1:
                values = [chunk_source[col].iloc[0] for col in source_columns]
                budget_data = BudgetData(**dict(zip(
                    ['dotacao_atual', 'dstq_concedido', 'credito_disp', 
                     'cred_indisp', 'dspz_p_empenhada', 'dsza_empenhada', 'dspz_pagas'],
                    values
                )))
            else:
                budget_data = BudgetData(
                    dotacao_atual=condensed_row.get('DOTACAO ATUALIZADA'),
                    dstq_concedido=condensed_row.get('DESTAQUE CONCEDIDO'),
                    credito_disp=condensed_row.get('CREDITO DISPONIVEL'),
                    cred_indisp=condensed_row.get('CREDITO INDISPONIVEL'),
                    dspz_p_empenhada=condensed_row.get('DESPESAS PRE-EMPENHADAS A EMPENHAR'),
                    dsza_empenhada=condensed_row.get('DESPESAS EMPENHADAS'),
                    dspz_pagas=condensed_row.get('DESPESAS PAGAS')
                )
            
            calculator = BudgetCalculator()
            
            result_b = calculator.item_b(comp_data.dotacao_disponivel, budget_data.credito_disp)
            result_c = calculator.item_c(comp_data.contingenciado, budget_data.cred_indisp, budget_data.dspz_p_empenhada)
            result_d = calculator.item_d(comp_data.pre_empenhado, budget_data.dspz_p_empenhada)
            result_e = calculator.item_e(comp_data.empenhado, budget_data.dstq_concedido, budget_data.dsza_empenhada)
            result_f = calculator.item_f(comp_data.pago, budget_data.dspz_pagas)
            result_a = calculator.item_a(comp_data.dotacao_atual, [result_b, result_c, result_d, result_e])
            
            # Atualizar DataFrame final
            if result_a != comp_data.dotacao_atual:
                df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['A']] = result_a
                print(f"   [A UPDATE] {filter_value}: {comp_data.dotacao_atual} -> {result_a}")
            if result_b != comp_data.dotacao_disponivel:
                df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['B']] = result_b
                print(f"   [B UPDATE] {filter_value}: {comp_data.dotacao_disponivel} -> {result_b}")
            if result_c != comp_data.contingenciado:
                df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['C']] = result_c
                print(f"   [C UPDATE] {filter_value}: {comp_data.contingenciado} -> {result_c}")
            if result_d != comp_data.pre_empenhado:
                df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['D']] = result_d
                print(f"   [D UPDATE] {filter_value}: {comp_data.pre_empenhado} -> {result_d}")
            if result_e != comp_data.empenhado:
                df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['E']] = result_e
                print(f"   [E UPDATE] {filter_value}: {comp_data.empenhado} -> {result_e}")
            if result_f != comp_data.pago:
                df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['F']] = result_f
                print(f"   [F UPDATE] {filter_value}: {comp_data.pago} -> {result_f}")
        
        # Salvar as alterações no novo arquivo (no diretório de trabalho atual)
        output_dir = os.getcwd()
        saved_path, updates_count = save_updates_without_breaking_formulas(
            dest_path_str, exact_columns, df_final_clean, output_dir
        )
        return saved_path, updates_count

class TextRedirector:
    """Classe para redirecionar stdout para o widget Text"""
    def __init__(self, widget):
        self.widget = widget
    def write(self, string):
        self.widget.insert(tk.END, string)
        self.widget.see(tk.END)
        self.widget.update()
    def flush(self):
        pass

if __name__ == "__main__":
    app = Application()
    app.mainloop()