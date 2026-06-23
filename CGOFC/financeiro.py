import os
import re
import sys
import shutil
import warnings

import pandas as pd


from datetime import datetime
from dataclasses import dataclass, fields
from pathlib import Path
from typing import Optional


from openpyxl import load_workbook


warnings.filterwarnings(
    "ignore",
    message=r"^Workbook contains no default style, apply openpyxl's default$",
    category=UserWarning,
)

# Enforce UTF-8 encoding for standard output and error to avoid UnicodeEncodeErrors on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# here only to clean the terminal before running the main logic, can be removed
os.system('cls' if os.name == 'nt' else 'clear')

# ==================================================
# Classe com lógica de negócios para cálculos orçamentários
# ==================================================
class BudgetCalculator:
    """Business logic for budget calculations"""
    
    def __init__(self):
        """Initialize calculator - no persistent state"""
        self.changes_detected = False
    
    def _to_float(self, value) -> float:
        """Convert value to float, or return 0.0 if conversion fails or value is not present"""
        if value is None:
            return 0.0
        if isinstance(value, str) and value.strip() == '':
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _check_change(self, calculated: float, comp_data: float, item: str) -> float:
        """Helper to check if value changed from comparison data"""
        comp_float = self._to_float(comp_data)
        calc_float = self._to_float(calculated)

        if (calc_float is None) and (comp_float is not None and comp_float > 0.0):
            print(f'Calculado inválido para item {item} ({calculated}); mantendo comparação: {comp_float}')
            return comp_float

        elif calc_float == comp_float:
            print(f'Dados do item {item} sem alteração')
            return comp_float
        
        elif calc_float != comp_float:
            print(f'‼️ Dados do item {item} alterados: {comp_float} -> {calc_float}')
            self.changes_detected = True
            return calc_float
    
    # ==================================================
    # Item A — Dotação Atual
    # ==================================================
    def item_a(self, comp_data, args: list) -> float:
        """
        Sum of items B through E
        
        Args:
            comp_data: Reference value to compare against
            *args: Values to sum (typically items B-E)
        """
        # Convert all args to float
        safe_args = [self._to_float(arg) for arg in args]
        result = sum(safe_args)
        return self._check_change(result, comp_data, item='A')
    
    # ==================================================
    # Item B — Dotação Disponível
    # ==================================================
    def item_b(self, comp_data, credito_disp) -> float:
        """Crédito disponível"""
        credito_disp = self._to_float(credito_disp)
        return self._check_change(credito_disp, comp_data, item='B')
    
    # ==================================================
    # Item C — Contingenciado / Bloqueado
    # ==================================================
    def item_c(self, comp_data, cdto_indisponivel, dsza_p_empenhada) -> float:
        """Crédito indisponível"""
        if not self._to_float(dsza_p_empenhada) == 0:
            result = self._to_float(cdto_indisponivel) - self._to_float(dsza_p_empenhada)
            return self._check_change(result, comp_data, item='C')
    
        else:
            result = self._to_float(cdto_indisponivel)
            return self._check_change(result, comp_data, item='C')
        
    # ==================================================
    # Item D — Pré-Empenhado
    # ==================================================
    def item_d(self, comp_data, dspz_p_empenhada) -> float:
        """
        Despesas pré-empenhadas a empenhar ou Crédito indisponível - despesas pré-empenhadas
        """
        result = self._to_float(dspz_p_empenhada)        
        return self._check_change(result, comp_data, item='D')
    
    # ==================================================
    # Item E — Empenhado / Descentralizado
    # ==================================================
    def item_e(self, comp_data, dstq_concedido, dspz_empenhada) -> float:
        """Destaque concedido + Despesas empenhadas"""
        result = self._to_float(dstq_concedido) + self._to_float(dspz_empenhada)
        return self._check_change(result, comp_data, item='E')
    
    # ==================================================
    # Item F — Pago
    # ==================================================
    def item_f(self, comp_data, dspz_pagas) -> float:
        """Despesas pagas"""
        dspz_pagas = self._to_float(dspz_pagas)
        return self._check_change(dspz_pagas, comp_data, item='F')


# ==================================================
# Dataclasse para pegar os valores da planilha de dados sintetizada e passar para o BudgetCalculator 
# ==================================================
@dataclass
class BudgetData:
    dotacao_atual: Optional[float] = None 
    dstq_concedido: Optional[float] = None 
    credito_disp: Optional[float] = None 
    cred_indisp: Optional[float] = None 
    dspz_p_empenhada: Optional[float] = None 
    dsza_empenhada: Optional[float] = None 
    dspz_pagas: Optional[float] = None

# ==================================================
# Dataclasse para pegar os valores da planilha de dados que serve como base de comparação e passar para o BudgetCalculator 
# ==================================================
@dataclass
class ComparisonData:
    """Comparison data for budget calculations"""
    ploa: Optional[float] = None           # PLOA 2026
    dotacao_atual: Optional[float] = None  # Dotação Atual (A)
    dotacao_disponivel: Optional[float] = None  # Dotação Disponível (B)
    contingenciado: Optional[float] = None  # Contingenciado/Bloqueado SOF (C)
    pre_empenhado: Optional[float] = None  # Pré-Empenhado (D)
    empenhado: Optional[float] = None      # Empenhado/Descentralizado (E)
    pago: Optional[float] = None           # Pago (F)

#==================================================
# Define PATHS para as pastas de origem e destino, e lógica para copiar os arquivos mais recentes 
# =================================================
def def_dir_paths(source_pag_total_mesp:bool=False, source_pag_total_orgaos:bool=False, limite_saque:bool=False, dest_dir:bool=False):
    # Define the directory paths
    directory = Path(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Teste001\CGOFC\financeiro")

    # ==================================================
    # Find source file for "Pago total Credor - MESP" (latest by modification date)
    # ==================================================
    if source_pag_total_mesp:
        source_pattern = "Pago total Credor - MESP*.xlsx"
        source_files = list(directory.glob(source_pattern))

        if not source_files:
            raise FileNotFoundError(f"No source files found matching pattern: {source_pattern}")

        # Get the most recently modified file
        source_path = max(source_files, key=lambda f: f.stat().st_mtime)
        print(f"Selected source file: {source_path.name}")
        print(f"Modified: {datetime.fromtimestamp(source_path.stat().st_mtime)}\n")

        return source_path
    
    # ==================================================
    # Find source file for "Pagamentos Totais - Órgãos" (latest by modification date)
    # ==================================================
    elif source_pag_total_orgaos:
        source_pattern = "Pagamentos Totais - *.xlsx"
        source_files = list(directory.glob(source_pattern))

        if not source_files:
            raise FileNotFoundError(f"No source files found matching pattern: {source_pattern}")

        # Get the most recently modified file
        source_path = max(source_files, key=lambda f: f.stat().st_mtime)
        print(f"Selected source file: {source_path.name}")
        print(f"Modified: {datetime.fromtimestamp(source_path.stat().st_mtime)}\n")

        return source_path
    
    # ==================================================
    # Find source file for "Limite de Saque - orgaos" (latest by modification date)
    # ==================================================
    elif limite_saque:
        source_pattern = "Limite de Saque - orgaos*.xlsx"
        source_files = list(directory.glob(source_pattern))

        if not source_files:
            raise FileNotFoundError(f"No source files found matching pattern: {source_pattern}")

        # Get the most recently modified file
        source_path = max(source_files, key=lambda f: f.stat().st_mtime)
        print(f"Selected source file: {source_path.name}")
        print(f"Modified: {datetime.fromtimestamp(source_path.stat().st_mtime)}\n")

        return source_path


    # ==================================================
    # Find destination file (latest by name or modification date)
    # ==================================================
    elif dest_dir:
        dest_pattern = "Demonstrativo Financeiro*.xlsx"
        dest_files = list(directory.glob(dest_pattern))

        if not dest_files:
            # If no file exists, use default name
            dest_path = directory / "Execução Orçamento 05.05.26.xlsx"
            print(f"No existing destination file found. Will create: {dest_path.name}")
        else:
            # Get the most recently modified destination file
            dest_path = max(dest_files, key=lambda f: f.stat().st_mtime)
            print(f"Selected destination file: {dest_path.name}")
            print(f"Modified: {datetime.fromtimestamp(dest_path.stat().st_mtime)}\n\n")

        return dest_path    

#==================================================
# Renomeia os cabeçalhos da planilha final de dados para nomes mais simples e consistentes
# =================================================
def clean_excel_columns(col_name):
    """Clean all column names by removing line breaks and extra spaces"""
    # Replace line breaks with space
    cleaned = col_name.replace('\n', '')
    # Remove extra spaces and trailing/leading whitespace
    cleaned = ' '.join(cleaned.split()).strip()
    # If no mapping, return cleaned name with underscores
    return cleaned
    
#==================================================
# Cria nome do arquivo de destino com a data de hoje, e lógica para evitar sobrescrever arquivos existentes (adiciona contador) 
# =================================================
def get_destination_with_today_date(base_dir: str) -> Path:
    """Create destination path with today's date"""
    directory = Path(base_dir)
    today_date = datetime.now().strftime("%d.%m.%y")
    
    # Option 1: Always create new file with today's date
    dest_path = directory / f"Execução Orçamento {today_date}.xlsx"
    
    # Option 2: If file already exists, add counter
    counter = 1
    original_path = dest_path
    while dest_path.exists():
        dest_path = directory / f"Execução Orçamento {today_date} ({counter}).xlsx"
        counter += 1
    
    if dest_path != original_path:
        print(f"File already exists. Using: {dest_path.name}")
    
    return dest_path

#==================================================
# Lógica para salvar as atualizações na planilha de destino sem quebrar fórmulas ou formatações, usando openpyxl para manipular o arquivo Excel diretamente.
# =================================================
def save_updates_without_breaking_formulas(dest_path: str, exact_columns: dict, df_final_clean: pd.DataFrame) -> Path:
    """Save updates to Excel while preserving all formulas"""
    
    base_dir = dest_path.parent
    file_path = get_destination_with_today_date(base_dir=str(base_dir))
    sheet_name = 'Dados'
    
    # Copy template (has all formatting and formulas)
    shutil.copy(dest_path, file_path)
    
    # Load the copied workbook with data_only=False to preserve formulas
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    # Find the column indices for the exact column names
    header_row = 3
    col_indices = {}
    
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value:
            cell_str = str(cell_value)
            if 'Dotação Atual' in cell_str and '(A)' in cell_str:
                col_indices['A'] = col_idx
            elif 'Dotação Disponível' in cell_str and '(B)' in cell_str:
                col_indices['B'] = col_idx
            elif 'Contingenciado/Bloqueado SOF' in cell_str:
                col_indices['C'] = col_idx
            elif 'Pré-Empenhado' in cell_str and '(D)' in cell_str:
                col_indices['D'] = col_idx
            elif 'Empenhado/Descentralizado' in cell_str and '(E)' in cell_str:
                col_indices['E'] = col_idx
            elif 'Pago' in cell_str and '(F)' in cell_str:
                col_indices['F'] = col_idx  
    
    # Create mapping from filter_col to Excel row
    row_to_filter = {}
    for excel_row in range(header_row + 1, ws.max_row + 1):
        try:
            col0 = ws.cell(row=excel_row, column=2).value
            col1 = ws.cell(row=excel_row, column=3).value
            col2 = ws.cell(row=excel_row, column=4).value
            
            col0_str = re.sub(r'\.0$', '', str(col0)).strip() if col0 and 'nan' not in str(col0).lower() else ''
            
            col1_raw = re.sub(r'\.0$', '', str(col1)).strip() if col1 else ''
            col1_clean = re.sub(r'\s+', '', col1_raw.split('-')[0]) if '-' in col1_raw else col1_raw
            col1_clean = col1_clean if 'nan' not in col1_clean.lower() else ''
            
            col2_raw = re.sub(r'\.0$', '', str(col2)).strip() if col2 else ''
            col2_clean = col2_raw[1] if len(col2_raw) > 1 else col2_raw[:-1]
            col2_clean = col2_clean if 'nan' not in col2_clean.lower() else ''
            
            filter_value = col0_str + col1_clean + col2_clean
            if filter_value and len(filter_value) == 6:
                row_to_filter[excel_row] = filter_value
        except Exception:
            continue
    
    # Update only the cells that changed
    updates_count = 0
    for excel_row, filter_value in row_to_filter.items():
        matching_rows = df_final_clean[df_final_clean['filter_col'] == filter_value]
        if matching_rows.empty:
            continue
        
        row_data = matching_rows.iloc[0]
        
        for col_letter, col_idx in col_indices.items():
            df_col_name = exact_columns[col_letter]
            if df_col_name in row_data and pd.notna(row_data[df_col_name]):
                new_value = row_data[df_col_name]
                current_value = ws.cell(row=excel_row, column=col_idx).value
                
                if current_value != new_value:
                    ws.cell(row=excel_row, column=col_idx, value=new_value)
                    updates_count += 1
                    print(f"Updated {filter_value} - Column {col_letter}: {current_value} -> {new_value}")
    
    # Save the workbook
    wb.save(file_path)
    print(f"\n✅ {updates_count} cells updated")
    print(f"✅ File saved with formulas preserved: {file_path}")
    
    return file_path

#==================================================
# Save a modified DataFrame back to a copied workbook without changing workbook structure
# =================================================
def save_dataframe_preserving_structure(
    df: pd.DataFrame,
    output_file: str,
    sheet_name: str,
    header_row: int = 3,
    preserve_formulas: bool = True,
) -> Path:
    """Save a DataFrame into a workbook copy while preserving the file structure and formulas."""
    wb = load_workbook(output_file, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook: {output_file}")

    print(sheet_name)
    ws = wb[sheet_name]

    workbook_headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        cleaned_header = clean_excel_columns(str(cell_value)) if cell_value is not None else ''
        workbook_headers.append((col_idx, cleaned_header, cell_value))
        print(f"[DEBUG] Workbook header found: col_idx={col_idx}, raw='{cell_value}', cleaned='{cleaned_header}'")

    print(f"[DEBUG] Total workbook headers: {len(workbook_headers)}\n")

    if not workbook_headers:
        raise ValueError(f"No headers found on row {header_row} in sheet '{sheet_name}'")

    df_to_ws_cols = {}
    print(f"[DEBUG] DataFrame columns: {list(df.columns)}\n")

    for idx, df_col in enumerate(df.columns):
        if idx >= len(workbook_headers):
            print(f"[DEBUG] ✗ No workbook header at position {idx} for df_col='{df_col}'")
            continue

        ws_col_idx, ws_cleaned, ws_raw = workbook_headers[idx]
        df_to_ws_cols[df_col] = ws_col_idx
        print(f"[DEBUG] ✓ Matched by position: df_col='{df_col}' -> ws_col={ws_col_idx} (workbook raw='{ws_raw}', cleaned='{ws_cleaned}')")

    print(f"\n[DEBUG] Total matched columns by position: {len(df_to_ws_cols)}\n")

    if not df_to_ws_cols:
        raise ValueError("None of the DataFrame columns could be mapped by position to workbook headers.")

    first_data_row = header_row + 1
    print(f"[DEBUG] header_row={header_row}, first_data_row={first_data_row}")
    print(f"[DEBUG] ws.max_row={ws.max_row}, df.shape={df.shape}\n")
    cell_updates = 0

    for row_offset, row in enumerate(df.itertuples(index=False, name=None)):
        target_row = first_data_row + row_offset
        if target_row > ws.max_row:
            break

        for df_col, ws_col in df_to_ws_cols.items():
            cell = ws.cell(row=target_row, column=ws_col)
            if preserve_formulas and (
                cell.data_type == 'f' or
                (isinstance(cell.value, str) and cell.value.startswith('='))
            ):
                if row_offset < 3:  # Show first 3 rows
                    print(f"[DEBUG] Row {target_row}, Col {ws_col}: Skipped (formula detected)")
                continue

            new_value = row[df.columns.get_loc(df_col)]
            if pd.isna(new_value):
                continue

            if cell.value != new_value:
                cell.value = new_value
                cell_updates += 1
                if row_offset < 3:  # Show first 3 rows
                    print(f"[DEBUG] Row {target_row}, Col {ws_col} ({df_col}): '{cell.value}' -> '{new_value}'")

    wb.save(output_file)
    print(f"\n[DEBUG] Saving workbook: {output_file}")
    print(f"✅ {cell_updates} cells written")
    print(f"✅ File saved preserving workbook structure: {output_file}")


#==================================================
# Seleciona o mês atual.
# =================================================
def gerar_mapeamento_meses() -> str:
    ano_atual = datetime.now().year
    abreviacoes = ["JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT","NOV","DEZ"]

    # Apenas mês atual
    mes_atual = f"{abreviacoes[datetime.now().month-1]}/{ano_atual}"

    return mes_atual

#==================================================
# Cria o filtro na planilha fonte para servir de guia na atualização dos dados 
# =================================================
def read_df_for_specific_sheet(dest_path:str, sheet_name:str, header:int, column_mapping:list) -> pd.DataFrame:
    """Read specific sheet from Excel into DataFrame"""
    mes_atual = gerar_mapeamento_meses()

    try:
        df = pd.read_excel(dest_path, sheet_name=sheet_name, header=header, dtype=str)
        df_filter = df.copy()

        # >>> FOR PAGO LIMITE DE SAQUE <<<
        if header == 5:
            df_filter['filter_col'] = df_filter.iloc[:, column_mapping].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()    
    
        # >>> FOR PAGAMENTOS TOTAIS ORGÃOS <<<
        elif header == 8:
            df_filter['filter_col'] = df_filter.iloc[:, column_mapping].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()
            df_filter[mes_atual] 

        # >>> FOR PAGO TOTAL MESP <<<
        elif header == 11:
            df_filter['filter_col'] = df_filter.iloc[:, column_mapping].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()
            df_filter[mes_atual]

        return df_filter

    except Exception as e:
        print(f"Error reading sheet '{sheet_name}': {type(e).__name__} - {str(e)[:80]}")
        sys.exit(1)

#==================================================
# Helper: make the large debug/update block reusable for debugging
#==================================================
def debug_apply_updates(df_dest: pd.DataFrame, df_source: pd.DataFrame, col_select_source: str, col_select_dest: str, max_detail_print: int = 10):
    """Apply updates from `df_source` into `df_dest` based on 'filter_col' and print debug information.

    Returns a tuple: (updated_df_dest, summary_dict)
    summary_dict contains: updated_count, skipped_count, update_details, total_matches, only_in_dest
    """
    # Basic validation
    if 'filter_col' not in df_dest.columns or 'filter_col' not in df_source.columns:
        raise ValueError("Both dataframes must contain 'filter_col' column")

    df_dest_filter = df_dest.set_index('filter_col')
    df_source_filter = df_source.set_index('filter_col')
    common_filters = df_dest_filter.index.intersection(df_source_filter.index)

    # Debugging prints
    print(f"     ✓ Filter column created on destiny dataframe. Sample values (first 5):")
    for i, val in enumerate(df_dest['filter_col'].head()[:5]):
        print(f"        Row {i}: '{val}'")
    print(f"     ✓ Total unique filter values: {df_dest['filter_col'].nunique()}\n")

    print(f"     ✓ Filter column created on source dataframe. Sample values (first 5):")
    for i, val in enumerate(df_source['filter_col'].head()[:5]):
        print(f"        Row {i}: '{val}'")
    print(f"     ✓ Total unique filter values: {df_source['filter_col'].nunique()}")
    
    print("\n[6] Verifying columns exist in dataframes...")
    if col_select_source in df_source.columns:
        print(f"     ✓ Source column '{col_select_source}' found in df_source")
    else:
        print(f"     ✗ ERROR: Source column '{col_select_source}' NOT found in df_source")
        print(f"       Available columns: {list(df_source.columns)}")

    if col_select_dest in df_dest.columns:
        print(f"     ✓ Destination column '{col_select_dest}' found in df_dest")
    else:
        print(f"     ✗ ERROR: Destination column '{col_select_dest}' NOT found in df_dest")
        print(f"       Available columns: {list(df_dest.columns)}")

    # Show filters that are ONLY in destination (will NOT be updated)
    only_in_dest = df_dest_filter.index.difference(df_source_filter.index)
    print(f"\n     ✓ Filters only in destination (no source match): {len(only_in_dest)}")
    if len(only_in_dest) > 0:
        print(f"       First 5 examples: {list(only_in_dest[:5])}")

    print("\n[9] Performing data updates...")
    updated_count = 0
    skipped_count = 0
    update_details = []

    for filter_value in common_filters:    
        # Skip Total and empty/whitespace filter values to prevent duplicate index matching
        if not filter_value or str(filter_value).strip() == '' or filter_value == 'Total':
            continue
        try:
            # Get source value before update
            source_value = df_source_filter.loc[filter_value, col_select_source]
            if isinstance(source_value, pd.Series):
                # Aggregate: sum all values in the Series
                source_value = source_value.sum()

            # Get destination value before update
            dest_old_value = df_dest_filter.loc[filter_value, col_select_dest]

            # Check if source value is not empty/null
            if pd.notna(source_value) and str(source_value).strip() != '':
                # Perform update
                df_dest_filter.loc[filter_value, col_select_dest] = source_value
                updated_count += 1
                
                # Compare safely in case there are duplicate rows causing dest_old_value to be a Series
                if isinstance(dest_old_value, pd.Series):
                    changed = not (dest_old_value == source_value).all()
                    dest_old_print = f"Series (size {len(dest_old_value)})"
                else:
                    changed = dest_old_value != source_value
                    dest_old_print = dest_old_value

                # Store details for reporting
                update_details.append({
                    'filter': filter_value,
                    'old_value': dest_old_print,
                    'new_value': source_value,
                    'changed': changed
                })
                
                # Print detailed update for first N and every 100th update
                if updated_count <= max_detail_print or updated_count % 100 == 0:
                    print(f"     ✓ Update #{updated_count}: filter='{filter_value}'")
                    print(f"        Old value: '{dest_old_print}'")
                    print(f"        New value: '{source_value}'")
                    if changed:
                        print(f"        Status: CHANGED")
                    else:
                        print(f"        Status: SAME VALUE (no effective change)")
            else:
                skipped_count += 1
                if skipped_count <= 5:  # Show first 5 skips only
                    print(f"     ⚠ Skip #{skipped_count}: filter='{filter_value}' - Source value is empty/NaN")
                
        except Exception as e:
            exc_type, exc_value, exc_tb = sys.exc_info()
            print(f"Error at line: {exc_tb.tb_lineno if exc_tb else 'unknown'}")
            print(f"     ✗ Error updating filter='{filter_value}': {type(e).__name__}: {str(e)[:80]}")

    df_dest = df_dest_filter.reset_index()

    print(f"\n     ✓ Update summary:")
    print(f"        - Successfully processed: {updated_count} rows")
    print(f"        - Skipped (empty source): {skipped_count} rows")
    print(f"        - Total matches found: {len(common_filters)}")

    # Show summary of changes
    if update_details:
        changes_made = [d for d in update_details if d['changed']]
        no_changes = [d for d in update_details if not d['changed']]
        print(f"\n     ✓ Change summary:")
        print(f"        - Values actually changed: {len(changes_made)}")
        print(f"        - Values remained same: {len(no_changes)}")
        
        # Show first 5 actual changes
        if changes_made:
            print(f"\n     ✓ First 5 actual changes made:")
            for i, change in enumerate(changes_made[:5]):
                print(f"        {i+1}. filter='{change['filter']}'")
                print(f"           '{change['old_value']}' → '{change['new_value']}'")

    summary = {
        'updated_count': updated_count,
        'skipped_count': skipped_count,
        'update_details': update_details,
        'total_matches': len(common_filters),
        'only_in_dest': list(only_in_dest)
    }

    return df_dest, summary

#==================================================
# Make sure that all headers are strings and are on the same format (e.g., 'JAN/2026') for easier access later
#==================================================
def change_header(dest_df: pd.DataFrame):
    abreviacoes = ["JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT","NOV","DEZ"]
    ano_atual = datetime.now().year

    old_columns = [
    col.strftime("%b/%Y").upper() if isinstance(col, datetime) else col 
    for col in dest_df.columns
            ]
    new_columns = [f'{month.split('/')[0]}/{ano_atual}' for month in old_columns]

# Direct assignment (no rename needed!)
    dest_df.columns = new_columns
    
    return dest_df

#==================================================
# Lógica do algoritmo para analisar diferenças entre planilhas.
# =================================================
def main():
    # >>> FOR SOURCE SHEETS <<<
    source_pag_total_mesp = def_dir_paths(source_pag_total_mesp=True)
    source_pag_total_orgaos = def_dir_paths(source_pag_total_orgaos=True)
    limite_saque = def_dir_paths(limite_saque=True) 

    # >>> FOR RESULT SHEET <<<
    dest_path = def_dir_paths(dest_dir=True)
    sheet_names = pd.ExcelFile(dest_path).sheet_names

    do_not_change = ['SIMULAÇÃO NOVEMBRO', 'Portaria MF  Nº 515 ', 'Demonstrativo Resumido', 'Fluxo de Caixa']
    
    for sheet in sheet_names:
        if sheet in do_not_change:
            continue
        print(f"Sample data from final sheet: {sheet}")

        df = pd.read_excel(dest_path, sheet_name=sheet, header=2, dtype=str)
        df = change_header(df)
        print(df.columns)
        save_dataframe_preserving_structure(
                                            df=df,
                                            output_file=dest_path,
                                            sheet_name=sheet,
                                            header_row=3,
                                            preserve_formulas=True,
                                        )
    df = pd.read_excel(dest_path, sheet_name=sheet, header=2, dtype=str)    

    print(df.columns)
    sys.exit()

    for i, aba in enumerate(sheet_names):       
        aba = str(aba).strip()
        print(f"Sample data from final sheet: {aba}")
        if aba == 'Fluxo de Caixa':
            column_mapping = [1, 2, 3, 4, 6]
            treasure_flux = read_df_for_specific_sheet(dest_path=limite_saque, sheet_name='Limite de Saque - orgaos', header=5, column_mapping=column_mapping)
            continue

        elif aba == 'Diárias - 180002':
            column_mapping =[2, 3]
            column_mapping_dest_df = [0, 1]
            daily = read_df_for_specific_sheet(dest_path=source_pag_total_mesp,
                                sheet_name='180002SECRETARIA EXECUTIVA, Mo',
                                header=11,
                                column_mapping=column_mapping
                                )
           
            # Reads the xlsx file and change the column names to match the month format (e.g., 'JUN/2026') for easier access
            df_daily = pd.read_excel(dest_path, sheet_name=aba, header=2, dtype=str)
            column_index_map = {}  
            for idx, col in enumerate(df_daily.columns.tolist()):
                if isinstance(col, datetime):
                    # Create string representation for easy access
                    new_name = col.strftime('%b/%Y').upper()  # 'JUN/2026'
                    column_index_map[new_name] = (col, idx)    # Store original datetime object and position
                else:
                    # Keep string columns as-is
                    column_index_map[col] = (col, idx)
            rename_dict = {orig_name: new_name for new_name, (orig_name, idx) in column_index_map.items()}
            df_daily = df_daily.rename(columns=rename_dict)

            # apply filter column creation logic to both dataframes using the same mapping
            df_daily['filter_col'] = df_daily.iloc[:, column_mapping_dest_df].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()

            # Get the column for this month in both dataframes
            col_select_source = gerar_mapeamento_meses()
            col_select_dest = col_select_source  # Both dataframes use the 'MMM/YYYY' format after column renaming

            # First, ensure both dataframes have filter_col as index or column
            df_daily_filter = df_daily.set_index('filter_col')
            daily_filter = daily.set_index('filter_col')

            # Find common filter values
            common_filters = df_daily_filter.index.intersection(daily_filter.index)
            
            # Use helper debug function to perform updates and produce a summary
            #debug_apply_updates(df_dest=df_daily,df_source=daily,col_select_source=col_select_source, col_select_dest=col_select_dest)
            continue

        elif aba == 'Contratos de Repasse - 180006':
            column_mapping =[2, 3]
            column_mapping_dest_df =[0, 1]

            onlendings = read_df_for_specific_sheet(dest_path=source_pag_total_mesp, sheet_name='180006CEFMINISTERIO DO ESPORT', header=11, column_mapping=column_mapping)
            df_onlendings = pd.read_excel(dest_path, sheet_name=aba, header=2, dtype=str)

            df_source = onlendings.copy()
            df_dest = df_onlendings.copy()

            # Reads the xlsx file and change the column names to match the month format (e.g., 'JUN/2026') for easier access
            column_index_map = {}  
            for idx, col in enumerate(df_dest.columns.tolist()):
                print(f"Original column: {col} (type: {type(col)})")
                if isinstance(col, datetime):
                    # Create string representation for easy access
                    new_name = col.strftime('%b/%Y').upper()  # 'JUN/2026'
                    column_index_map[new_name] = (col, idx)    # Store original datetime object and position
                else:
                    # Keep string columns as-is
                    column_index_map[col] = (col, idx)

            rename_dict = {orig_name: new_name for new_name, (orig_name, idx) in column_index_map.items()}
            df_dest = df_dest.rename(columns=rename_dict)

            # apply filter column creation logic to both dataframes using the same mapping
            df_dest['filter_col'] = df_dest.iloc[:, column_mapping_dest_df].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()

            # Get the column for this month in both dataframes
            col_select_source = gerar_mapeamento_meses()
            col_select_dest = col_select_source  # Both dataframes use the 'MMM/YYYY' format after column renaming

            # First, ensure both dataframes have filter_col as index or column
            df_dest = df_dest.set_index('filter_col')
            df_source_filter = df_source.set_index('filter_col')

            # Find common filter values
            common_filters = df_dest.index.intersection(df_source_filter.index)
            
            # Use helper debug function to perform updates and produce a summary
            debug_apply_updates(df_dest=df_dest,df_source=df_source,col_select_source=col_select_source, col_select_dest=col_select_dest)
        
            continue

        elif aba == 'Bolsa Atleta - 180009':
            column_mapping =[2, 3]
            column_mapping_dest_df =[0, 1]

            grant = read_df_for_specific_sheet(dest_path=source_pag_total_mesp, sheet_name='180009SECRETARIA NACIONAL DE E', header=11, column_mapping=column_mapping)
            df_grant = pd.read_excel(dest_path, sheet_name=aba, header=2, dtype=str)

            df_source = grant.copy()
            df_dest = df_grant.copy()

            # Reads the xlsx file and change the column names to match the month format (e.g., 'JUN/2026') for easier access
            column_index_map = {}  
            for idx, col in enumerate(df_dest.columns.tolist()):
                print(f"Original column: {col} (type: {type(col)})")
                if isinstance(col, datetime):
                    # Create string representation for easy access
                    new_name = col.strftime('%b/%Y').upper()  # 'JUN/2026'
                    column_index_map[new_name] = (col, idx)    # Store original datetime object and position
                else:
                    # Keep string columns as-is
                    column_index_map[col] = (col, idx)

            rename_dict = {orig_name: new_name for new_name, (orig_name, idx) in column_index_map.items()}
            df_dest = df_dest.rename(columns=rename_dict)

            # apply filter column creation logic to both dataframes using the same mapping
            df_dest['filter_col'] = df_dest.iloc[:, column_mapping_dest_df].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()

            # Get the column for this month in both dataframes
            col_select_source = gerar_mapeamento_meses()
            col_select_dest = col_select_source  # Both dataframes use the 'MMM/YYYY' format after column renaming

            # First, ensure both dataframes have filter_col as index or column
            df_dest = df_dest.set_index('filter_col')
            df_source_filter = df_source.set_index('filter_col')

            # Find common filter values
            common_filters = df_dest.index.intersection(df_source_filter.index)
            
            # Use helper debug function to perform updates and produce a summary
            debug_apply_updates(df_dest=df_dest,df_source=df_source,col_select_source=col_select_source, col_select_dest=col_select_dest)
        
            continue

        elif aba == 'ABCD - 180016':
            column_mapping =[]
            abcd = read_df_for_specific_sheet(dest_path=source_pag_total_mesp, sheet_name='180016AUTORIDADE BRASILEIRA DE', header=11, column_mapping=column_mapping)
            continue

        elif aba == 'Termo de Fomento' or aba == 'Convênio':
            column_mapping =[]
            foster_pact = read_df_for_specific_sheet(dest_path=source_pag_total_mesp, sheet_name='550029SECRETARIA NACIONAL DE P', header=11, column_mapping=column_mapping)
            continue
        
        elif aba == 'Tarifas da Caixa':
            column_mapping =[]
            tariff = read_df_for_specific_sheet(dest_path=source_pag_total_mesp, sheet_name='Total, Movim. Líquido - R$ (Ite', column_mapping=column_mapping, header=11)
            continue

        elif aba == 'Valores pagos das PFs de 2026':
            column_mapping = [2, 3, 8, 9]
            column_mapping_dest_df = [2, 3, 4, 5]

            # Data source dataframe
            paied = read_df_for_specific_sheet(dest_path=source_pag_total_orgaos,
                                               sheet_name='Discricionárias - Fontes Tesour',
                                               header=8,
                                               column_mapping=column_mapping
                                               )
            
            # Last updated dataframe
            df_pf_paied = pd.read_excel(dest_path, sheet_name=aba, header=2, dtype=str)
            df_pf_paied['filter_col'] = df_pf_paied.iloc[:, column_mapping_dest_df].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()

            # Get the column for this month in both dataframes
            col_select_source = gerar_mapeamento_meses()
            col_select_dest = f'{gerar_mapeamento_meses().split('/')[0]}/2024'

            # First, ensure both dataframes have filter_col as index or column
            df_pf_paied_filter = df_pf_paied.set_index('filter_col')
            paied_filter = paied.set_index('filter_col')

            # Find common filter values
            common_filters = df_pf_paied_filter.index.intersection(paied_filter.index)
            
            for filter_value in common_filters:
                # For each matching row, update df_pf_paied with paied data
                source_value = paied_filter.loc[filter_value, col_select_source]
                if isinstance(source_value, pd.Series):
                    # Aggregate: sum all values in the Series
                    source_value = source_value.sum()

                df_pf_paied_filter.loc[filter_value, col_select_dest] = source_value
            
            df_pf_paied = df_pf_paied_filter.reset_index()
            continue    
    sys.exit(0)

    # Now create BudgetData dataclass directly from condensed_row
    budget_data = BudgetData(
        dotacao_atual=condensed_row.get('DOTACAO ATUALIZADA'),
        dstq_concedido=condensed_row.get('DESTAQUE CONCEDIDO'),
        credito_disp=condensed_row.get('CREDITO DISPONIVEL'),
        cred_indisp=condensed_row.get('CREDITO INDISPONIVEL'),
        dspz_p_empenhada=condensed_row.get('DESPESAS PRE-EMPENHADAS A EMPENHAR'),
        dsza_empenhada=condensed_row.get('DESPESAS EMPENHADAS'),
        dspz_pagas=condensed_row.get('DESPESAS PAGAS')
    )

    # ==================================================
    # Now use both dataclasses with BudgetCalculator
    # ==================================================
    calculator = BudgetCalculator()

    # Item B - Dotação Disponível
    result_b = calculator.item_b(
        comp_data=comp_data.dotacao_disponivel,  # from ComparisonData
        credito_disp=budget_data.credito_disp     # from BudgetData
    )

    # Item C - Contingenciado/Bloqueado SOF
    result_c = calculator.item_c(
        comp_data=comp_data.contingenciado,
        cdto_indisponivel=budget_data.cred_indisp,
        dsza_p_empenhada=budget_data.dspz_p_empenhada
    )

    # Item D - Pré-Empenhado
    result_d = calculator.item_d(
        comp_data=comp_data.pre_empenhado,
        dspz_p_empenhada=budget_data.dspz_p_empenhada,
    )

    # Item E - Empenhado/Descentralizado
    result_e = calculator.item_e(
        comp_data=comp_data.empenhado,
        dstq_concedido=budget_data.dstq_concedido,
        dspz_empenhada=budget_data.dsza_empenhada
    )

    # Item F - Pago
    result_f = calculator.item_f(
        comp_data=comp_data.pago,
        dspz_pagas=budget_data.dspz_pagas
    )

    result_a = calculator.item_a(
        comp_data=comp_data.dotacao_atual,  # Dotação Atual (A) from ComparisonData
        args=[result_b, result_c, result_d, result_e]  # Results from items B-E
    )

    exact_columns = {
        'A': 'Dotação Atual\n(A)',
        'B': ' Dotação Disponível          \n(B)',
        'C': 'Contingenciado/Bloqueado SOF                            \n( C )',
        'D': 'Pré-Empenhado              (D)',
        'E': 'Empenhado/Descentralizado \n(E)',
        'F': 'Pago              \n (F)'
    }

    # Now update using the exact column names
    if result_a != comp_data.dotacao_atual:
        df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['A']] = result_a
        print(f"Updated {clean_excel_columns(exact_columns['A'])} for {filter_value}: {result_a}")

    if result_b != comp_data.dotacao_disponivel:
        df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['B']] = result_b
        print(f"Updated {clean_excel_columns(exact_columns['B'])} for {filter_value}: {result_b}")

    if result_c != comp_data.contingenciado:
        df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['C']] = result_c
        print(f"Updated {clean_excel_columns(exact_columns['C'])} for {filter_value}: {result_c}")

    if result_d != comp_data.pre_empenhado:
        df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['D']] = result_d
        print(f"Updated {clean_excel_columns(exact_columns['D'])} for {filter_value}: {result_d}")

    if result_e != comp_data.empenhado:
        df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['E']] = result_e
        print(f"Updated {clean_excel_columns(exact_columns['E'])} for {filter_value}: {result_e}")

    if result_f != comp_data.pago:
        df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['F']] = result_f
        print(f"Updated {clean_excel_columns(exact_columns['F'])} for {filter_value}: {result_f}")
                
    print("=" * 80 + "\n")
    
    file_path = save_updates_without_breaking_formulas(dest_path=dest_path, exact_columns=exact_columns, df_final_clean=df_final_clean)

    # Copy to dashboard folder
    dashboard_dir = Path(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\SNEAELIS - Python\SE")
    dashboard_name = f"Planilha alimentada - {file_path.name}"
    try:
        shutil.copy2(file_path, Path.joinpath(dashboard_dir, dashboard_name))
        print(f"✓ Copied updated file to dashboard: {dashboard_dir} \\ {dashboard_name}")
    except Exception as e:
        print(f"✗ Error copying updated file to dashboard: {type(e).__name__}: {str(e)[:100]}")

    print(f"Data successfully written to {file_path}\n and copied to dashboard folder @ {dashboard_dir}.")

if __name__ == "__main__":
    main()