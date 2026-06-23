import os
import re
import sys
import glob
import shutil
import warnings


from datetime import datetime
from dataclasses import dataclass, fields
from pathlib import Path
from typing import Optional

import pandas as pd

from openpyxl import load_workbook


warnings.filterwarnings(
    "ignore",
    message=r"^Workbook contains no default style, apply openpyxl's default$",
    category=UserWarning,
)

# here only to clean the terminal before running the main logic, can be removed
os.system('cls' if os.name == 'nt' else 'clear')

# ==================================================
# Classe com lógica de negócios para cálculos orçamentários
# ==================================================
class BudgetCalculator:
    """Business logic for budget calculations"""
    
    def __init__(self):
        """Initialize calculator - no persistent state"""
        self.changes_detected = False
    
    def _to_float(self, value) -> float:
        """Convert value to float, or return 0.0 if conversion fails or value is not present"""
        if value is None:
            return 0.0
        if isinstance(value, str) and value.strip() == '':
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _check_change(self, calculated: float, comp_data: float, item: str) -> float:
        """Helper to check if value changed from comparison data"""
        comp_float = self._to_float(comp_data)
        calc_float = self._to_float(calculated)

        if (calc_float is None) and (comp_float is not None and comp_float > 0.0):
            print(f'Calculado inválido para item {item} ({calculated}); mantendo comparação: {comp_float}')
            return comp_float

        elif calc_float == comp_float:
            print(f'Dados do item {item} sem alteração')
            return comp_float
        
        elif calc_float != comp_float:
            print(f'‼️ Dados do item {item} alterados: {comp_float} -> {calc_float}')
            self.changes_detected = True
            return calc_float
    
    # ==================================================
    # Item A — Dotação Atual
    # ==================================================
    def item_a(self, comp_data, args: list) -> float:
        """
        Sum of items B through E
        
        Args:
            comp_data: Reference value to compare against
            *args: Values to sum (typically items B-E)
        """
        # Convert all args to float
        safe_args = [self._to_float(arg) for arg in args]
        result = sum(safe_args)
        return self._check_change(result, comp_data, item='A')
    
    # ==================================================
    # Item B — Dotação Disponível
    # ==================================================
    def item_b(self, comp_data, credito_disp) -> float:
        """Crédito disponível"""
        credito_disp = self._to_float(credito_disp)
        return self._check_change(credito_disp, comp_data, item='B')
    
    # ==================================================
    # Item C — Contingenciado / Bloqueado
    # ==================================================
    def item_c(self, comp_data, cdto_indisponivel, dsza_p_empenhada) -> float:
        """Crédito indisponível"""
        if not self._to_float(dsza_p_empenhada) == 0:
            result = self._to_float(cdto_indisponivel) - self._to_float(dsza_p_empenhada)
            return self._check_change(result, comp_data, item='C')
    
        else:
            result = self._to_float(cdto_indisponivel)
            return self._check_change(result, comp_data, item='C')
        
    # ==================================================
    # Item D — Pré-Empenhado
    # ==================================================
    def item_d(self, comp_data, dspz_p_empenhada) -> float:
        """
        Despesas pré-empenhadas a empenhar ou Crédito indisponível - despesas pré-empenhadas
        """
        result = self._to_float(dspz_p_empenhada)        
        return self._check_change(result, comp_data, item='D')
    
    # ==================================================
    # Item E — Empenhado / Descentralizado
    # ==================================================
    def item_e(self, comp_data, dstq_concedido, dspz_empenhada) -> float:
        """Destaque concedido + Despesas empenhadas"""
        result = self._to_float(dstq_concedido) + self._to_float(dspz_empenhada)
        return self._check_change(result, comp_data, item='E')
    
    # ==================================================
    # Item F — Pago
    # ==================================================
    def item_f(self, comp_data, dspz_pagas) -> float:
        """Despesas pagas"""
        dspz_pagas = self._to_float(dspz_pagas)
        return self._check_change(dspz_pagas, comp_data, item='F')

# ==================================================
# Cria o filtro na planilha final para servir de guia na atualização dos dados
# ==================================================
def create_filter_column(df):
    def process_row(row):
        # Column 0
        try:
            col0 = re.sub(r'\.0$', '', str(row.iloc[0])).strip() if 'nan' not in str(row.iloc[0]).lower() else ''
            
        except Exception as e:
            print(f"Error processing col0: {type(e).__name__} - {str(e)[:80]}")
            sys.exit(1)
            col0 = ''

        # Column 1 - get everything before dash
        try:
            col1_raw = re.sub(r'\.0$', '', str(row.iloc[1])).strip()  # Remove spaces
            col1 = re.sub(r'\s+', '', col1_raw.split('-')[0]) if '-' in col1_raw else col1_raw
            col1 = col1 if 'nan' not in col1.lower() else ''
        except Exception as e:
            print(f"Error processing col1: {type(e).__name__} - {str(e)[:80]}")
            sys.exit(1)
            col1 = ''
        
        # Column 2 - get second digit
        try:
            col2_raw = re.sub(r'\.0$', '', str(row.iloc[2])).strip()
            col2 = col2_raw[1] if len(col2_raw) > 1 else col2_raw[:-1]
            col2 = col2 if 'nan' not in col2.lower() else ''
        except Exception as e:
            print(f"Error processing col2: {type(e).__name__} - {str(e)[:80]}")
            col2 = ''
        
        return col0 + col1 + col2
    
    return df.iloc[:, [1, 2, 3]].apply(process_row, axis=1)

# ==================================================
# Dataclasse para pegar os valores da planilha de dados sintetizada e passar para o BudgetCalculator 
# ==================================================
@dataclass
class BudgetData:
    dotacao_atual: Optional[float] = None 
    dstq_concedido: Optional[float] = None 
    credito_disp: Optional[float] = None 
    cred_indisp: Optional[float] = None 
    dspz_p_empenhada: Optional[float] = None 
    dsza_empenhada: Optional[float] = None 
    dspz_pagas: Optional[float] = None

# ==================================================
# Dataclasse para pegar os valores da planilha de dados que serve como base de comparação e passar para o BudgetCalculator 
# ==================================================
@dataclass
class ComparisonData:
    """Comparison data for budget calculations"""
    ploa: Optional[float] = None           # PLOA 2026
    dotacao_atual: Optional[float] = None  # Dotação Atual (A)
    dotacao_disponivel: Optional[float] = None  # Dotação Disponível (B)
    contingenciado: Optional[float] = None  # Contingenciado/Bloqueado SOF (C)
    pre_empenhado: Optional[float] = None  # Pré-Empenhado (D)
    empenhado: Optional[float] = None      # Empenhado/Descentralizado (E)
    pago: Optional[float] = None           # Pago (F)

#==================================================
# Define PATHS para as pastas de origem e destino, e lógica para copiar os arquivos mais recentes 
# =================================================
def def_dir_paths(source_dir:bool=False, dest_dir:bool=False):
    # Define the directory paths
    directory = Path(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Teste001\CGOFC")

    # ==================================================
    # Find source file (latest by modification date)
    # ==================================================
    if source_dir:
        source_pattern = "1.1 - MESP-Ano 2025-UG 180077 - Por resultado primário - C*.xlsx"
        source_files = list(directory.glob(source_pattern))

        if not source_files:
            raise FileNotFoundError(f"No source files found matching pattern: {source_pattern}")

        # Get the most recently modified file
        source_path = max(source_files, key=lambda f: f.stat().st_mtime)
        print(f"Selected source file: {source_path.name}")
        print(f"Modified: {datetime.fromtimestamp(source_path.stat().st_mtime)}\n")

        return source_path

    # ==================================================
    # Find destination file (latest by name or modification date)
    # ==================================================
    if dest_dir:
        dest_pattern = "Execução Orçamento *.xlsx"
        dest_files = list(directory.glob(dest_pattern))

        if not dest_files:
            # If no file exists, use default name
            dest_path = directory / "Execução Orçamento 05.05.26.xlsx"
            print(f"No existing destination file found. Will create: {dest_path.name}")
        else:
            # Get the most recently modified destination file
            dest_path = max(dest_files, key=lambda f: f.stat().st_mtime)
            print(f"Selected destination file: {dest_path.name}")
            print(f"Modified: {datetime.fromtimestamp(dest_path.stat().st_mtime)}\n\n")

        return dest_path    

#==================================================
# Renomeia os cabeçalhos da planilha final de dados para nomes mais simples e consistentes
# =================================================
def clean_excel_columns(col_name):
    """Clean all column names by removing line breaks and extra spaces"""
    # Replace line breaks with space
    cleaned = col_name.replace('\n', '')
    # Remove extra spaces and trailing/leading whitespace
    cleaned = ' '.join(cleaned.split()).strip()
    # If no mapping, return cleaned name with underscores
    return cleaned
    
#==================================================
# Cria nome do arquivo de destino com a data de hoje, e lógica para evitar sobrescrever arquivos existentes (adiciona contador) 
# =================================================
def get_destination_with_today_date(base_dir: str) -> Path:
    """Create destination path with today's date"""
    directory = Path(base_dir)
    today_date = datetime.now().strftime("%d.%m.%y")
    
    # Option 1: Always create new file with today's date
    dest_path = directory / f"Execução Orçamento {today_date}.xlsx"
    
    # Option 2: If file already exists, add counter
    counter = 1
    original_path = dest_path
    while dest_path.exists():
        dest_path = directory / f"Execução Orçamento {today_date} ({counter}).xlsx"
        counter += 1
    
    if dest_path != original_path:
        print(f"File already exists. Using: {dest_path.name}")
    
    return dest_path

#==================================================
# Lógica para salvar as atualizações na planilha de destino sem quebrar fórmulas ou formatações, usando openpyxl para manipular o arquivo Excel diretamente.
# =================================================
def save_updates_without_breaking_formulas(dest_path: str, exact_columns: dict, df_final_clean: pd.DataFrame) -> Path:
    """Save updates to Excel while preserving all formulas"""
    
    base_dir = dest_path.parent
    file_path = get_destination_with_today_date(base_dir=str(base_dir))
    sheet_name = 'Dados'
    
    # Copy template (has all formatting and formulas)
    shutil.copy(dest_path, file_path)
    
    # Load the copied workbook with data_only=False to preserve formulas
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    # Find the column indices for the exact column names
    header_row = 3
    col_indices = {}
    
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value:
            cell_str = str(cell_value)
            if 'Dotação Atual' in cell_str and '(A)' in cell_str:
                col_indices['A'] = col_idx
            elif 'Dotação Disponível' in cell_str and '(B)' in cell_str:
                col_indices['B'] = col_idx
            elif 'Contingenciado/Bloqueado SOF' in cell_str:
                col_indices['C'] = col_idx
            elif 'Pré-Empenhado' in cell_str and '(D)' in cell_str:
                col_indices['D'] = col_idx
            elif 'Empenhado/Descentralizado' in cell_str and '(E)' in cell_str:
                col_indices['E'] = col_idx
            elif 'Pago' in cell_str and '(F)' in cell_str:
                col_indices['F'] = col_idx  
    
    # Create mapping from filter_col to Excel row
    row_to_filter = {}
    for excel_row in range(header_row + 1, ws.max_row + 1):
        try:
            col0 = ws.cell(row=excel_row, column=2).value
            col1 = ws.cell(row=excel_row, column=3).value
            col2 = ws.cell(row=excel_row, column=4).value
            
            col0_str = re.sub(r'\.0$', '', str(col0)).strip() if col0 and 'nan' not in str(col0).lower() else ''
            
            col1_raw = re.sub(r'\.0$', '', str(col1)).strip() if col1 else ''
            col1_clean = re.sub(r'\s+', '', col1_raw.split('-')[0]) if '-' in col1_raw else col1_raw
            col1_clean = col1_clean if 'nan' not in col1_clean.lower() else ''
            
            col2_raw = re.sub(r'\.0$', '', str(col2)).strip() if col2 else ''
            col2_clean = col2_raw[1] if len(col2_raw) > 1 else col2_raw[:-1]
            col2_clean = col2_clean if 'nan' not in col2_clean.lower() else ''
            
            filter_value = col0_str + col1_clean + col2_clean
            if filter_value and len(filter_value) == 6:
                row_to_filter[excel_row] = filter_value
        except Exception:
            continue
    
    # Update only the cells that changed
    updates_count = 0
    for excel_row, filter_value in row_to_filter.items():
        matching_rows = df_final_clean[df_final_clean['filter_col'] == filter_value]
        if matching_rows.empty:
            continue
        
        row_data = matching_rows.iloc[0]
        
        for col_letter, col_idx in col_indices.items():
            df_col_name = exact_columns[col_letter]
            if df_col_name in row_data and pd.notna(row_data[df_col_name]):
                new_value = row_data[df_col_name]
                current_value = ws.cell(row=excel_row, column=col_idx).value
                
                if current_value != new_value:
                    ws.cell(row=excel_row, column=col_idx, value=new_value)
                    updates_count += 1
                    print(f"Updated {filter_value} - Column {col_letter}: {current_value} -> {new_value}")
    
    # Save the workbook
    wb.save(file_path)
    print(f"\n✅ {updates_count} cells updated")
    print(f"✅ File saved with formulas preserved: {file_path}")
    
    return file_path

#==================================================
# Lógica do algoritmo para analisar diferenças entre planilhas.
# =================================================
def main():
    # ==================================================
    # Lógica para processamento de dados e atualização de planilhas
    # ==================================================
    source_path = def_dir_paths(source_dir=True)
    dest_path = def_dir_paths(dest_dir=True)

    df_data_source = pd.read_excel(source_path, header=2)
    df_data_final = pd.read_excel(dest_path, sheet_name='Dados', header=2)
    #df_data_final = clean_excel_columns(df_data_final)

    # Step 1: Create filter column (handling NaN)
    df_data_source_clean = df_data_source.copy()
    # Column mapping using the actual header names (as strings)
    column_mapping = {
    13: 'DOTACAO ATUALIZADA',
    18: 'DESTAQUE CONCEDIDO',
    19: 'CREDITO DISPONIVEL',
    20: 'CREDITO INDISPONIVEL',
    22: 'DESPESAS PRE-EMPENHADAS A EMPENHAR',
    23: 'DESPESAS EMPENHADAS',
    28: 'DESPESAS PAGAS'
                    }
    # Rename the columns
    df_data_source_clean.rename(columns=column_mapping, inplace=True)

    df_data_source_clean['filter_col'] = df_data_source_clean.iloc[:, [0, 2, 5]].fillna('').astype(str).agg(''.join, axis=1).str.replace(' ', '').str.replace(r'\.0$', '', regex=True).str.strip()    
    
    # Destination file dataframe
    df_final_clean = df_data_final.copy()

    df_final_clean['filter_col'] = create_filter_column(df_data_final)


    # Step 2: Get unique filter values as a set
    filter_final_df_Set = set(df_final_clean['filter_col'])
    
    source_columns = [
        'DOTACAO ATUALIZADA',
        'DESTAQUE CONCEDIDO',
        'CREDITO DISPONIVEL',
        'CREDITO INDISPONIVEL',
        'DESPESAS PRE-EMPENHADAS A EMPENHAR',
        'DESPESAS EMPENHADAS',
        'DESPESAS PAGAS'
        ]

    for filter_value in set(df_data_source_clean['filter_col']):
        if len(filter_value) != 6 or 'total' in filter_value.lower() or filter_value in ['', 'nan', 'none', None] or filter_value not in filter_final_df_Set:
            continue
                
        chunk_source  = df_data_source_clean[df_data_source_clean['filter_col'] == filter_value].copy()
        chunk_final = df_final_clean[df_final_clean['filter_col'] == filter_value].copy()

        print("\n" + "=" * 80)
        print(f"   Processing filter value: {filter_value} with {len(chunk_source )} rows   ".center(80))
        print("=" * 80)

        field_names = [f.name for f in fields(ComparisonData)]
        values = chunk_final.iloc[:, 5:12]   # Assuming items B-F are in columns 5-11 (0-indexed)

        comp_data = ComparisonData(**dict(zip(field_names, values.iloc[0])))

        condensed_row = {}
        for col in chunk_source .columns[9:20]:  # Columns 9 to 19 (0-indexed) - adjust if needed
            # Get non-NaN values from the column
            col_values = chunk_source [col].fillna(0.0)
            
            if len(col_values) == 0:
                # All NaN, keep as NaN
                condensed_row[col] = None
                continue
            
            # Try to convert to numeric for summation
            numeric_values = []
            can_sum = True
                
            for val in col_values:
                # Check if value is already numeric
                if isinstance(val, (int, float)):
                    numeric_values.append(val)
                # Try to convert string to float
                elif isinstance(val, str):
                    try:
                        numeric_values.append(float(val))
                    except ValueError:
                        # Can't convert to float, can't sum this column
                        can_sum = False
                        break
                else:
                    # Other types (datetime, etc.) - can't sum
                    can_sum = False
                    break
                
            if can_sum and len(numeric_values) > 0:
                # Sum the numeric values
                condensed_row[col] = sum(numeric_values)
                #print(f"  Column '{col}': summed {len(numeric_values)} values -> {condensed_row[col]}")
            else:
                # Can't sum, take first non-NaN value
                condensed_row[col] = col_values.iloc[0]
                #print(f"  Column '{col}': took first value -> {condensed_row[col]}")
        if len(chunk_source) == 1:
            # Get values from the source columns
            values = [chunk_source[col].iloc[0] for col in source_columns]
            
            # Create BudgetData using zip
            budget_data = BudgetData(**dict(zip(
                ['dotacao_atual', 'dstq_concedido', 'credito_disp', 
                'cred_indisp', 'dspz_p_empenhada', 'dsza_empenhada', 'dspz_pagas'],
                values
            )))

        # Now create BudgetData dataclass directly from condensed_row
        budget_data = BudgetData(
            dotacao_atual=condensed_row.get('DOTACAO ATUALIZADA'),
            dstq_concedido=condensed_row.get('DESTAQUE CONCEDIDO'),
            credito_disp=condensed_row.get('CREDITO DISPONIVEL'),
            cred_indisp=condensed_row.get('CREDITO INDISPONIVEL'),
            dspz_p_empenhada=condensed_row.get('DESPESAS PRE-EMPENHADAS A EMPENHAR'),
            dsza_empenhada=condensed_row.get('DESPESAS EMPENHADAS'),
            dspz_pagas=condensed_row.get('DESPESAS PAGAS')
        )
   
        # ==================================================
        # Now use both dataclasses with BudgetCalculator
        # ==================================================
        calculator = BudgetCalculator()

        # Item B - Dotação Disponível
        result_b = calculator.item_b(
            comp_data=comp_data.dotacao_disponivel,  # from ComparisonData
            credito_disp=budget_data.credito_disp     # from BudgetData
        )

        # Item C - Contingenciado/Bloqueado SOF
        result_c = calculator.item_c(
            comp_data=comp_data.contingenciado,
            cdto_indisponivel=budget_data.cred_indisp,
            dsza_p_empenhada=budget_data.dspz_p_empenhada
        )

        # Item D - Pré-Empenhado
        result_d = calculator.item_d(
            comp_data=comp_data.pre_empenhado,
            dspz_p_empenhada=budget_data.dspz_p_empenhada,
        )

        # Item E - Empenhado/Descentralizado
        result_e = calculator.item_e(
            comp_data=comp_data.empenhado,
            dstq_concedido=budget_data.dstq_concedido,
            dspz_empenhada=budget_data.dsza_empenhada
        )

        # Item F - Pago
        result_f = calculator.item_f(
            comp_data=comp_data.pago,
            dspz_pagas=budget_data.dspz_pagas
        )

        result_a = calculator.item_a(
            comp_data=comp_data.dotacao_atual,  # Dotação Atual (A) from ComparisonData
            args=[result_b, result_c, result_d, result_e]  # Results from items B-E
        )

        exact_columns = {
            'A': 'Dotação Atual\n(A)',
            'B': ' Dotação Disponível          \n(B)',
            'C': 'Contingenciado/Bloqueado SOF                            \n( C )',
            'D': 'Pré-Empenhado              (D)',
            'E': 'Empenhado/Descentralizado \n(E)',
            'F': 'Pago              \n (F)'
        }

        # Now update using the exact column names
        if result_a != comp_data.dotacao_atual:
            df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['A']] = result_a
            print(f"Updated {clean_excel_columns(exact_columns['A'])} for {filter_value}: {result_a}")

        if result_b != comp_data.dotacao_disponivel:
            df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['B']] = result_b
            print(f"Updated {clean_excel_columns(exact_columns['B'])} for {filter_value}: {result_b}")

        if result_c != comp_data.contingenciado:
            df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['C']] = result_c
            print(f"Updated {clean_excel_columns(exact_columns['C'])} for {filter_value}: {result_c}")

        if result_d != comp_data.pre_empenhado:
            df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['D']] = result_d
            print(f"Updated {clean_excel_columns(exact_columns['D'])} for {filter_value}: {result_d}")

        if result_e != comp_data.empenhado:
            df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['E']] = result_e
            print(f"Updated {clean_excel_columns(exact_columns['E'])} for {filter_value}: {result_e}")

        if result_f != comp_data.pago:
            df_final_clean.loc[df_final_clean['filter_col'] == filter_value, exact_columns['F']] = result_f
            print(f"Updated {clean_excel_columns(exact_columns['F'])} for {filter_value}: {result_f}")
                 
        print("=" * 80 + "\n")
    
    file_path = save_updates_without_breaking_formulas(dest_path=dest_path, exact_columns=exact_columns, df_final_clean=df_final_clean)

    # Copy to dashboard folder
    dashboard_dir = Path(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\SNEAELIS - Python\SE")
    dashboard_name = f"Planilha alimentada - {file_path.name}"
    try:
        shutil.copy2(file_path, Path.joinpath(dashboard_dir, dashboard_name))
        print(f"✓ Copied updated file to dashboard: {dashboard_dir} \\ {dashboard_name}")
    except Exception as e:
        print(f"✗ Error copying updated file to dashboard: {type(e).__name__}: {str(e)[:100]}")

    print(f"Data successfully written to {file_path}\n and copied to dashboard folder @ {dashboard_dir}.")

if __name__ == "__main__":
    main()