from pandas import ExcelWriter
import pdfplumber
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError, Error as PlaywrightError
from playwright.sync_api import sync_playwright
import base64
from PyPDF2 import PdfReader, PdfWriter
import pytesseract
from selenium.common import WebDriverException, TimeoutException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from selenium.webdriver import ActionChains
from colorama import Fore, Style
import time
import json
import fitz  # PyMuPDF
import pyautogui
import re
import zipfile
from openpyxl import load_workbook
from copy import copy
from PIL import Image, ImageEnhance, ImageFilter
import shutil
import sys
import pandas as pd
import io
import os

r'''
"C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="%USERPROFILE%\chrome_profile" --disable-features=TabSearch --disable-component-extensions-with-background-pages --no-first-run --force-dark-mode --enable-features=WebContentsForceDark "https://idp.transferegov.sistema.gov.br/idp/"
"C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9224 --user-data-dir="%USERPROFILE%\chrome_profile_2" --disable-features=TabSearch --disable-component-extensions-with-background-pages --no-first-run --force-dark-mode --enable-features=WebContentsForceDark "https://idp.transferegov.sistema.gov.br/idp/"
"C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9226 --user-data-dir="%USERPROFILE%\chrome_profile_3" --disable-features=TabSearch --disable-component-extensions-with-background-pages --no-first-run --force-dark-mode --enable-features=WebContentsForceDark "https://idp.transferegov.sistema.gov.br/idp/"
'''



pytesseract.pytesseract.tesseract_cmd = r"C:\Users\felipe.rsouza\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"



if __name__ == "__main__":
    func = int(input("Choose a function: "))

    # split the nucleai on the dir into subdirs
    if func == 2:
        def separate_nuclei():
            def aid_func(filename):
                pattern = r'Chamadas\s(.*?)(?:\.pdf|_compressed\.pdf|\.xlsx\.pdf|_compressed_compressed\.pdf)'
                match = re.search(pattern, filename, re.IGNORECASE)
                if match:
                    nucleus_name = match.group(1).strip()
                    return nucleus_name.replace('_.', '')
                return None

            dir_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                        r'Social\SNEAELIS - Termo Fomento Inst. Léo Moura\Frequencia 926508')

            for root, dirs, files in os.walk(dir_path):
                if 'Frequencias 910782' in dirs:
                    dirs.remove('Frequencias 910782')
                for filename in files:
                    if not filename.lower().endswith('.pdf'):
                        continue
                    nucleus = aid_func(filename)

                    if nucleus:
                        new_dir_path = os.path.join(root, nucleus)
                        if not os.path.exists(new_dir_path):
                            print(f"Creating directory: {new_dir_path}")
                            os.makedirs(new_dir_path)
                        source_file = os.path.join(root, filename)
                        destination_file = os.path.join(new_dir_path, filename)

                        print(f"Moving file: {filename} to {new_dir_path}")
                        shutil.move(source_file, destination_file)


    # .pdf to .xlsx
    elif func == 3:
        def pdf_to_xlsx():
            def write_log(log_file, message):
                """Write messages to log file"""
                with open(log_file, 'a', encoding='utf-8') as f:
                    f.write(message + '\n')

            def extract_tables_with_pdfplumber(pdf_source):
                """Extract tables using pdfplumber with focus on area below 'Nº'"""
                tables = []

                with pdfplumber.open(pdf_source) as pdf:
                    for page in pdf.pages:
                        # Find the 'Nº' marker position
                        y_start = None
                        for word in page.extract_words():
                            if word['text'].strip().upper() == 'Nº':
                                y_start = word['top']
                                break

                        if y_start is None:
                            continue  # Skip if no Nº found

                        # Define crop area (everything below Nº)
                        crop_area = (0, y_start - 5, page.width, page.height)  # x0, top, x1, bottom

                        # Extract table from this region
                        cropped_page = page.crop(crop_area)
                        table = cropped_page.extract_table()

                        if table and len(table) > 1:  # Ensure we have at least header + one row
                            tables.append(table)

                return tables

            dir_path = (r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência "
                        r"Social\SNEAELIS - Termo Fomento Inst. Léo Moura")

            # Create log file
            log_file = os.path.join(dir_path,
                                    f"conversion_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")

            for root, dirs, files in os.walk(dir_path):
                if 'Frequencias 910782' in dirs:
                    dirs.remove('Frequencias 910782')
                for filename in files:
                    if not filename.lower().endswith('.pdf'):
                        continue

                    pdf_path = os.path.join(root, filename)
                    xlsx_filename = filename.replace('.pdf', '.xlsx')
                    xlsx_path = os.path.join(root, xlsx_filename)

                    log_message = f"\nProcessing {filename}..."
                    print(log_message)
                    write_log(log_file, log_message)

                    try:
                        # Try with original PDF first
                        try:
                            tables = extract_tables_with_pdfplumber(pdf_path)
                            if tables:
                                log_message = "Extracted tables with pdfplumber"
                                print(log_message)
                                write_log(log_file, log_message)
                            else:
                                raise ValueError("No tables found in initial extraction")
                        except Exception as e:
                            log_message = f"Initial extraction failed, trying cleaned version: {str(e)[:200]}"
                            print(log_message)
                            write_log(log_file, log_message)

                            # Clean PDF in memory
                            pdf_reader = PdfReader(pdf_path)
                            pdf_writer = PdfWriter()

                            for page in pdf_reader.pages:
                                if "/Annots" in page:
                                    del page["/Annots"]
                                pdf_writer.add_page(page)

                            # Create in-memory PDF bytes
                            pdf_bytes = io.BytesIO()
                            pdf_writer.write(pdf_bytes)
                            pdf_bytes.seek(0)

                            # Try extraction again with cleaned PDF
                            tables = extract_tables_with_pdfplumber(pdf_bytes)
                            if not tables:
                                raise ValueError("All extraction methods failed")

                        # Save to Excel
                        with pd.ExcelWriter(xlsx_path) as writer:
                            for i, table in enumerate(tables):
                                try:
                                    # Convert table data to DataFrame
                                    df = pd.DataFrame(table[1:], columns=table[0])
                                    df.to_excel(writer, sheet_name=f"Table_{i + 1}", index=False)
                                except Exception as e:
                                    log_message = f"Error saving table {i + 1}: {str(e)[:200]}"
                                    print(log_message)
                                    write_log(log_file, log_message)
                                    continue

                        log_message = f"Successfully converted {filename} to {xlsx_filename}"
                        print(log_message)
                        write_log(log_file, log_message)

                    except Exception as e:
                        log_message = f"Failed to process {filename}: {type(e).__name__} - {str(e)[:200]}"
                        print(log_message)
                        write_log(log_file, log_message)
                        continue


    # Convert to .pdf from .html
    elif func == 4:
        def html_to_pdf(html_file_path, pdf_file_path=None):
            """
            Convert an HTML file to PDF using Chrome browser.

            Args:
                html_file_path (str): Path to the input HTML file
                pdf_file_path (str): Path for the output PDF file (optional)

            Returns:
                bool: True if conversion was successful, False otherwise
            """
            # If no output path is provided, use the same name with .pdf extension
            if pdf_file_path is None:
                pdf_file_path = os.path.splitext(html_file_path)[0] + '.pdf'

            # Set up Chrome options
            chrome_options = Options()
            chrome_options.add_argument('--headless')  # Run in background
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--window-size=1920,1080')  # Set window size for consistent rendering

            try:
                # Initialize the driver
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)

                try:
                    # Convert file path to URL (handle Windows paths)
                    absolute_path = os.path.abspath(html_file_path)
                    # Replace backslashes with forward slashes for Windows
                    absolute_path = absolute_path.replace('\\', '/')
                    file_url = f"file:///{absolute_path}"

                    # Navigate to the HTML file
                    driver.get(file_url)

                    # Wait for page to load (more reliable than implicit wait for PDF generation)
                    time.sleep(2)  # Adjust this based on your content complexity

                    # Print to PDF with better options
                    print_options = {
                        'landscape': False,
                        'displayHeaderFooter': False,
                        'printBackground': True,
                        'preferCSSPageSize': True,
                        'marginTop': 0.5,
                        'marginBottom': 0.5,
                        'marginLeft': 0.5,
                        'marginRight': 0.5,
                        'pageRanges': '',  # Print all pages
                    }

                    # Execute the print command
                    result = driver.execute_cdp_cmd('Page.printToPDF', print_options)

                    # Ensure the output directory exists
                    os.makedirs(os.path.dirname(pdf_file_path), exist_ok=True)

                    # Save the PDF
                    with open(pdf_file_path, 'wb') as f:
                        f.write(base64.b64decode(result['data']))

                    print(f"✓ Successfully converted: {html_file_path} -> {pdf_file_path}")
                    return True

                except Exception as e:
                    print(f"✗ Error processing {html_file_path}: {str(e)}")
                    return False

                finally:
                    driver.quit()

            except Exception as e:
                print(f"✗ Failed to initialize browser for {html_file_path}: {str(e)}")
                return False


        def convert_all_html_files(root_directory='.', output_directory=None, pattern=('.html', '.htm')):
            """
            Recursively find and convert all HTML files in a directory tree.

            Args:
                root_directory (str): Root directory to search for HTML files
                output_directory (str): Custom output directory (optional)
                pattern (tuple): File extensions to search for
            """
            success_count = 0
            total_count = 0

            print(f"Searching for HTML files in: {os.path.abspath(root_directory)}")

            for root, dirs, files in os.walk(root_directory):
                for file in files:
                    if file.lower().endswith(pattern):
                        total_count += 1
                        html_file_path = os.path.join(root, file)

                        # Determine output PDF path
                        if output_directory:
                            # Maintain directory structure in output directory
                            relative_path = os.path.relpath(root, root_directory)
                            pdf_dir = os.path.join(output_directory, relative_path)
                            pdf_file_path = os.path.join(pdf_dir, os.path.splitext(file)[0] + '.pdf')
                        else:
                            # Save in same directory as HTML file
                            pdf_file_path = None

                        # Convert the file
                        if html_to_pdf(html_file_path, pdf_file_path):
                            success_count += 1

            print(f"\n📊 Conversion Summary:")
            print(f"   Total HTML files found: {total_count}")
            print(f"   Successfully converted: {success_count}")
            print(f"   Failed: {total_count - success_count}")

            return success_count, total_count


        root_dir = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                    r'Social\Teste001\Sofia\Pareceres_SEi')
        convert_all_html_files(root_directory=root_dir)

    # Split parecer
    elif func == 5:
        def split_parecer():
            try:
                df = pd.read_excel(
                    r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                    r'Social\Teste001\Sofia\Pareceres_SEi\processos_DB.xlsx', dtype=str)
                for i, r in df.iterrows():
                    partial_txt = r.iloc[1]
                    txt = partial_txt.split('º')[-1].strip()
                    txt = txt.replace(')', '')
                    df.at[i, 1] = txt
                df.to_excel(r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                            r'Social\Teste001\Sofia\Pareceres_SEi\processos_DB.xlsx',
                            index=False, )
            except Exception as e:
                print(f"✗ Failed to initialize browser for {type(e).__name__}: {str(e)[:100]}")
                return False


    # confere os pads com time.sleep
    elif func == 7:
        class Robo:
            def webdriver_element_wait(self, xpath: str):
                """
                        Espera até que um elemento web esteja clicável, usando um tempo limite máximo de 3 segundos.

                        Args:
                            xpath: O seletor XPath do elemento.

                        Returns:
                            O elemento web clicável, ou lança uma exceção TimeoutException se o tempo limite for atingido.

                        Raises:
                            TimeoutException: Se o elemento não estiver clicável dentro do tempo limite.
                        """
                # Cria uma instância de WebDriverWait com o driver e o tempo limite e espera o elemento ser clicável
                try:
                    return WebDriverWait(self.driver, 8).until(
                        EC.element_to_be_clickable((By.XPATH, xpath)))
                except Exception as e:
                    raise e

            # Chama a função do webdriver com wait element to be clickable
            def __init__(self):
                """
                Inicializa o objeto Robo, configurando e iniciando o driver do Chrome.
                """
                try:
                    # Configuração do registro
                    # Inicia as opções do Chrome
                    self.chrome_options = webdriver.ChromeOptions()
                    # Endereço de depuração para conexão com o Chrome
                    self.chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
                    # Inicializa o driver do Chrome com as opções e o gerenciador de drivers
                    self.driver = webdriver.Chrome(
                        service=Service(ChromeDriverManager().install()),
                        options=self.chrome_options)
                    self.driver.switch_to.window(self.driver.window_handles[0])

                    print("✅ Conectado ao navegador existente com sucesso.")
                except WebDriverException as e:
                    # Imprime mensagem de erro se a conexão falhar
                    print(f"❌ Erro ao conectar ao navegador existente: {e}")
                    # Define o driver como None em caso de falha na conexão
                    self.driver = None

            # Navega até a página de busca da proposta
            def consulta_proposta(self):
                """
                       Navega pelas abas do sistema até a página de busca de processos.

                       Esta função clica nas abas principal e secundária para acessar a página
                       onde é possível realizar a busca de processos.
                       """
                # Reseta para página inicial
                try:
                    reset = self.webdriver_element_wait('//*[@id="header"]')
                    if reset:
                        img = reset.find_element(By.XPATH, '//*[@id="logo"]/a/img')
                        action = ActionChains(self.driver)
                        action.move_to_element(img).perform()
                        reset.find_element(By.TAG_NAME, 'a').click()
                        # print(Fore.MAGENTA + "\n✅ Processo resetado com sucesso !")
                except NoSuchElementException:
                    print('Já está na página inicial do transferegov discricionárias.')
                except Exception as e:
                    print(Fore.RED + f'🔄❌ Falha ao resetar.\nErro: {type(e).__name__}\n{str(e)[:50]}')

                # [0] Excução; [1] Consultar Proposta
                xpaths = ['//*[@id="menuPrincipal"]/div[1]/div[3]',
                          '//*[@id="contentMenu"]/div[1]/ul/li[2]/a'
                          ]
                try:
                    for idx in range(len(xpaths)):
                        self.webdriver_element_wait(xpaths[idx]).click()
                    # print(f"{Fore.MAGENTA}✅ Sucesso em acessar a página de busca de processo{Style.RESET_ALL}")
                except Exception as e:
                    print(Fore.RED + f'🔴📄 Instrumento indisponível. \nErro: {e}')
                    sys.exit(1)

            def campo_pesquisa(self, numero_processo):
                try:
                    # Seleciona campo de consulta/pesquisa, insere o número de proposta/instrumento e da ENTER
                    campo_pesquisa = self.webdriver_element_wait('//*[@id="consultarNumeroProposta"]')
                    campo_pesquisa.clear()
                    campo_pesquisa.send_keys(numero_processo)
                    campo_pesquisa.send_keys(Keys.ENTER)

                    # Acessa o item proposta/instrumento
                    acessa_item = self.webdriver_element_wait('//*[@id="tbodyrow"]/tr/td[1]/div/a')
                    acessa_item.click()
                except Exception as e:
                    print(
                        f' Falha ao inserir número de processo no campo de pesquisa. Erro: {type(e).__name__}')


            def nav_plano_act_det(self):
                time.sleep(1)
                try:
                    print(f'🚢 Navegando para o plano de ação detalhado:'.center(50, '-'), '\n')
                    # Aba Plano de trabalho
                    self.webdriver_element_wait('//*[@id="div_997366806"]').click()

                    # Seleciona Plano de Aplicação Detalhado
                    self.webdriver_element_wait('//*[@id="menu_link_997366806_836661414"]/div').click()

                except Exception as e:
                    print(
                        f"❌ Ocorreu um erro ao executar ao pesquisar Plano de Ação Detalhado: {type(e).__name__}"
                        f".\n Erro {str(e)[:50]}")


            # Loop para adicionar PAD da proposta
            def loop_de_pesquisa(self, df=None, numero_processo: str=None, tipo_desp: str=None,
                                 cod_natur_desp: str=None, cnpj_xlsx: str=None,
                                 caminho_arquivo_fonte: str=None):

                # Inicia o processo de consulta do instrumento
                try:

                    # Pesquisa pelo processo
                    self.campo_pesquisa(numero_processo=numero_processo)

                    # Executa pesquisa de anexos
                    self.nav_plano_act_det()
                    tab_path = '/html/body/div[3]/div[14]/div[4]/div/div/div/form/div[1]'
                    WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, tab_path)))
                    #time.sleep(15)
                    self.consulta_proposta()
                    return True

                except TimeoutException as t:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f'TIMEOUT {str(t)[:50]}')
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    self.consulta_proposta()
                except Exception as erro:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f'❌ Falha ao tentar incluir documentos. Erro:{type(erro).__name__}\n'
                          f'{str(erro)[:100]}...')
                    self.consulta_proposta()

            # Finds which locator to use
            def find_button_with_retry(self):
                # Define all possible locator strategies and values
                locators = [
                    (By.ID, 'form_submit'),
                    (By.NAME, 'detalharEsclarecimentoConvenioDadosDoEsclarecimentoVoltarForm'),
                    (By.XPATH, '//input[@value="Voltar"]'),
                    (By.XPATH, '//td[@class="FormLinhaBotoes"]/input'),
                    (By.CLASS_NAME, 'FormLinhaBotoes'),  # Will need additional find after
                    (By.XPATH, '//input[contains(@onclick, "setaAcao")]'),
                    (By.XPATH, '//*[@id="form_submit"]')
                ]

                for locator in locators:
                    try:
                        # print(f'Botão localizado com o seletor {locator[0]}')
                        return self.driver.find_element(*locator)
                    except Exception as e:
                        print(f"Falha com localizador {locator}: {str(e)[:80]}")
                        continue

                raise NoSuchElementException("Could not find button using any locator strategy")

        def main() -> None:
            l = ['15657/2024']

            try:
                # Instancia um objeto da classe Robo
                robo = Robo()
                # Extrai dados de colunas específicas do Excel
            except Exception as e:
                print(f"\n‼️ Erro fatal ao iniciar o robô: {e}")
                sys.exit("Parando o programa.")

            # inicia consulta e leva até a página de busca do processo
            robo.consulta_proposta()

            for v in l:
                try:
                    robo.loop_de_pesquisa(numero_processo=v)
                except KeyboardInterrupt:
                    print("Script stopped by user (Ctrl+C). Exiting cleanly.")
                    sys.exit(0)  # Exit gracefully
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(
                        f"❌ Falha ao executar script. Erro: {type(e).__name__}\n{str(e)[:100]}")
                    sys.exit(0)  # Exit gracefully

        main()

    # Check if directory is empty
    elif func == 8:
        root_dir = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Automações SNEAELIS\CGAC_2024')
        empty_dris = []  # Use a list to store the folder paths

        for root, dirs, files in os.walk(root_dir):
            for dir in dirs:
                try:
                    with os.scandir(os.path.join(root, dir)) as entries:
                        if next(entries, None) is None:
                            empty_dris.append(os.path.basename(dir).split('_')[0])
                except:
                    print(f"Warning: Could not access directory.")
        print(f'Number of empty dirs: {len(set(empty_dris))}.\nName of the empty dirs: {set(empty_dris)}')


    # Fake long funtion
    elif func == 9:
        class BreakInnerLoop(Exception):
            pass


        class PWRobo:
            def __init__(self, cdp_url: str = "http://localhost:9222"):
                self.playwright = sync_playwright().start()
                self.browser = self.playwright.chromium.connect_over_cdp(cdp_url)
                self.context = self.browser.contexts[0] if self.browser.contexts else self.browser.new_context()
                self.page = self.context.pages[0] if self.context.pages else self.context.new_page()

                # Enable resource blocking for faster performance
                # self.block_rss()

                print(f"✅ Connected to existing Chrome instance via Playwright. Connected to page: {self.page.url}")

            def block_rss(self):
                """Block images, stylesheets, and fonts to make browsing faster"""

                def route_handler(route, request):
                    if request.resource_type in ["image", "stylesheet", "font"]:
                        route.abort()
                    else:
                        route.continue_()

                self.page.route('**/*', route_handler)
                print("✅ Resource blocking enabled for faster performance")

            def consulta_proposta(self):
                """Navigates through the system tabs to the process search page."""
                time.sleep(1.5)
                try:
                    self.page.locator('xpath=//*[@id="logo"]/a').click(timeout=800)
                except PlaywrightTimeoutError:
                    print('Already on the initial page of transferegov discricionárias.')
                except PlaywrightError as e:
                    print(Fore.RED + f'🔄❌ Failed to reset.\nError: {type(e).__name__}\n{str(e)[:100]}')

                xpaths = ['xpath=//*[@id="menuPrincipal"]/div[1]/div[3]',
                          'xpath=//*[@id="contentMenu"]/div[1]/ul/li[2]/a']
                try:
                    for xpath in xpaths:
                        self.page.locator(xpath).click(timeout=10000)
                except PlaywrightError as e:
                    print(Fore.RED + f'🔴📄 Instrument unavailable. \nError: {e}')
                    sys.exit(1)

            def campo_pesquisa(self, numero_processo):
                try:
                    campo_pesquisa_locator = self.page.locator('xpath=//*[@id="consultarNumeroProposta"]')
                    campo_pesquisa_locator.fill(numero_processo)
                    campo_pesquisa_locator.press('Enter')
                    try:
                        acessa_item_locator = self.page.locator('xpath=//*[@id="tbodyrow"]/tr/td[1]/div/a')
                        acessa_item_locator.click(timeout=8000)
                    except PlaywrightTimeoutError:
                        print(f' Process number: {numero_processo}, not found.')
                        raise BreakInnerLoop
                    except PlaywrightError as e:
                        print(f' Process number: {numero_processo}, not found. Error: {type(e).__name__}')
                        raise BreakInnerLoop
                except PlaywrightError as e:
                    print(f' Failed to insert process number in the search field. Error: {type(e).__name__}')

            def busca_endereco(self):
                time.sleep(1)

                self.page.wait_for_timeout(500)
                try:
                    print(f'🔍 Starting data search'.center(50, '-'), '\n')

                    # Click the detail button
                    self.page.locator('input#form_submit[value="Detalhar"]').click()

                    # Wait for the page to load
                    self.page.wait_for_timeout(2000)

                    # Initialize with empty values
                    phone = ""
                    email = ""

                    # Check if elements exist before trying to extract text
                    phone_locator = self.page.locator('#txtTelefone')
                    email_locator = self.page.locator('#txtEmail')

                    phone_exists = phone_locator.count() > 0
                    email_exists = email_locator.count() > 0

                    print(f"🔍 Element check - Phone field exists: {phone_exists}, Email field exists: {email_exists}")

                    # Extract phone if available
                    if phone_exists:
                        try:
                            phone_locator.wait_for(timeout=3000)
                            phone = phone_locator.inner_text(timeout=2000).strip()
                            print(f"📞 Phone found: '{phone}'")
                        except PlaywrightTimeoutError:
                            print("⚠️ Phone element exists but content not loaded or empty")
                        except Exception as e:
                            print(f"⚠️ Error extracting phone: {e}")
                    else:
                        print("📞 Phone field not found on page")

                    # Extract email if available
                    if email_exists:
                        try:
                            email_locator.wait_for(timeout=3000)
                            email = email_locator.inner_text(timeout=2000).strip()
                            print(f"📧 Email found: '{email}'")
                        except PlaywrightTimeoutError:
                            print("⚠️ Email element exists but content not loaded or empty")
                        except Exception as e:
                            print(f"⚠️ Error extracting email: {e}")
                    else:
                        print("📧 Email field not found on page")

                    # Check if we got any data
                    if not phone and not email:
                        print("ℹ️ No contact information found on this page")
                    elif phone and not email:
                        print("ℹ️ Only phone number found")
                    elif email and not phone:
                        print("ℹ️ Only email found")
                    else:
                        print("✅ Both phone and email found")

                    # Back button
                    try:
                        self.page.locator('xpath=//*[@id="lnkConsultaAnterior"]').click(timeout=8000)
                    except:
                        print("⚠️ Could not click back button, navigating manually")
                        # Alternative navigation if back button fails
                        self.page.go_back()

                    return phone, email

                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f"❌ Unexpected error in busca_endereco: {type(e).__name__} - {str(e)[:100]}")
                    # Ensure we navigate back even on error
                    self.page.locator('xpath=//*[@id="lnkConsultaAnterior"]').click(timeout=2000)
                    return "", ""

            def mark_as_done(self, df, numero_processo, phone, email):
                """Safely mark row as done in the DataFrame"""
                time.sleep(1)
                try:
                    # Find the row index where the process number matches
                    mask = df.iloc[:, 4] == numero_processo

                    if mask.any():
                        # Get the actual index position(s)
                        idx_positions = df.index[mask]
                        for idx in idx_positions:
                            df.at[idx, 'Telefone'] = phone
                            df.at[idx, 'Email'] = email

                        print(f"✅ Updated process {numero_processo}: Phone={phone}, Email={email}")
                        return True
                    else:
                        print(f"❌ Process number {numero_processo} not found in DataFrame")
                        return False

                except Exception as err:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f"❌ Error in mark_as_done:{type(err).__name__}\n{str(err)[:100]}\n")
                    return False

            def loop_de_pesquisa(self, df, numero_processo: str):
                time.sleep(1)

                print(f'🔍 Starting data extraction loop'.center(50, '-'), '\n')

                try:
                    self.campo_pesquisa(numero_processo=numero_processo)
                    phone, email = self.busca_endereco()

                    if phone is not None and email is not None:
                        success = self.mark_as_done(df=df, numero_processo=numero_processo, phone=phone, email=email)
                        if success:
                            self.page.locator('xpath=//*[@id="breadcrumbs"]/a[2]').click(timeout=3000)
                            return True
                    return False

                except PlaywrightTimeoutError as t:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f'TIMEOUT {str(t)[:50]}')
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    self.consulta_proposta()
                    return False
                except Exception as erro:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f'❌ Failed to try to include documents. Error:{type(erro).__name__}\n'
                          f'{str(erro)[:100]}...')
                    self.consulta_proposta()
                    raise BreakInnerLoop

            # --- Fixed Save Function ---
            @staticmethod
            def save_to_excel(df, caminho_arquivo_fonte, sheet_name='Sheet1'):
                time.sleep(1.5)

                try:
                    # Create backup of original file
                    backup_path = caminho_arquivo_fonte.replace('.xlsx', '_backup.xlsx')
                    if os.path.exists(caminho_arquivo_fonte):
                        import shutil
                        shutil.copy2(caminho_arquivo_fonte, backup_path)
                        print(f"✅ Backup created: {backup_path}")

                    # Save using 'w' mode to overwrite the entire file
                    with ExcelWriter(caminho_arquivo_fonte, engine='openpyxl', mode='w') as writer:
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                    print(f"✅ Successfully saved {len(df)} rows to Excel.")
                    return True

                except PermissionError:
                    print(f"❌ Permission denied: Please close the Excel file '{caminho_arquivo_fonte}' and try again.")
                    return False
                except Exception as erro:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    line_number = exc_tb.tb_lineno
                    if exc_tb.tb_next:
                        line_number = exc_tb.tb_next.tb_lineno
                    print(f"Error occurred at line: {line_number}")
                    print(f'❌ Failed to save Excel file. Error:{type(erro).__name__}\n{str(erro)}')
                    return False

            @staticmethod
            def extrair_dados_excel(caminho_arquivo_fonte):
                try:
                    complete_data_frame = pd.read_excel(caminho_arquivo_fonte, dtype=str, sheet_name=None)
                    sheet_names_list = list(complete_data_frame.keys())

                    print("✅ Sheet Names Found:")
                    for name in sheet_names_list:
                        print(f"- {name}")

                    data_frame = complete_data_frame['Sheet1']
                    print(f"✅ Loaded {len(data_frame)} rows from Excel.")
                    return data_frame

                except Exception as e:
                    print(f"🤷‍♂️❌ Error reading the excel file: {os.path.basename(caminho_arquivo_fonte)}.\n"
                          f"Error name: {type(e).__name__}\nError: {str(e)}")

            @staticmethod
            def fix_prop_num(numero_proposta):
                time.sleep(0.75)

                if pd.isna(numero_proposta):
                    return False

                numero_proposta = str(numero_proposta)

                pattern = r'^\d{5}/\d{4}'

                if '_' in numero_proposta:
                    numero_proposta_fixed = numero_proposta.replace('_', '/')
                    return numero_proposta_fixed

                if re.findall(pattern, numero_proposta):
                    return numero_proposta
                else:
                    return False


        def main() -> None:
            dir_path = (
                r'C:\Users\felipe.rsouza\Documents\fake_func\Convênios - Pendências Celebração (24.11) - Copia.xlsx')

            try:
                robo = PWRobo()
            except Exception as e:
                print(f"\n‼️ Fatal error starting the robot: {e}")
                sys.exit("Stopping the program.")

            # Load DataFrame
            df_total = robo.extrair_dados_excel(caminho_arquivo_fonte=dir_path)
            if df_total is None:
                print("❌ Failed to load Excel file. Exiting.")
                return

            # Make sure the required columns exist
            if 'Telefone' not in df_total.columns:
                df_total['Telefone'] = ''
            if 'Email' not in df_total.columns:
                df_total['Email'] = ''

            mask = (df_total['Telefone'].isna() | (df_total['Telefone'] == '')) & (df_total['Email'].isna() | (df_total['Email'] == ''))

            df = df_total[mask]

            skipped_count = len(df_total) - len(df)
            print(f"⏭️ Skipping {skipped_count} rows - already processed (both phone and email exist)")

            robo.consulta_proposta()

            successful_updates = 0

            for idx, row in df.iterrows():
                time.sleep(1.5)

                numero_processo_temp = row.iloc[4]
                numero_processo = robo.fix_prop_num(numero_processo_temp)

                if not numero_processo:
                    continue

                print(f"\n{'⚡' * 3}🚀 EXECUTING PROPOSAL: {numero_processo} 🚀{'⚡' * 3}".center(70, '='), '\n')
                print(f"Current progress {idx / len(df):.2%}.")

                try:
                    success = robo.loop_de_pesquisa(df=df, numero_processo=numero_processo)
                    if success:
                        successful_updates += 1

                    # Save every 10 rows or at the end
                    if (idx + 1) % 10 == 0 or idx == len(df) - 1:
                        print(f"💾 Saving progress... ({idx + 1}/{len(df)} rows processed)")
                        if robo.save_to_excel(df=df, caminho_arquivo_fonte=dir_path):
                            print(f"✅ Successfully saved {successful_updates} updates so far, {skipped_count} skipped")

                except BreakInnerLoop:
                    print("⚠️ Stopping this unique_values loop early.")
                    # Save progress before breaking
                    robo.save_to_excel(df=df, caminho_arquivo_fonte=dir_path)
                    break
                except KeyboardInterrupt:
                    print("Script stopped by user (Ctrl+C). Saving progress...")
                    robo.save_to_excel(df=df, caminho_arquivo_fonte=dir_path)
                    sys.exit(0)
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f"❌ Failed to execute script. Error: {type(e).__name__}\n{str(e)}")
                    # Save progress on error
                    robo.save_to_excel(df=df, caminho_arquivo_fonte=dir_path)

            print(f"\n🎉 Processing complete!")
            print(f"📊 Summary: {successful_updates} updated, {skipped_count} skipped, {len(df)} total")


        if __name__ == "__main__":
            main()


    # busca dados de quantos arquivos são esperados em cada proposta
    elif func == 10:
        class Robo:
            def __init__(self):
                """
                Inicializa o objeto Robo, configurando e iniciando o driver do Chrome.
                """
                try:
                    # Configuração do registro
                    # self.arquivo_registro = ''
                    # Inicia as opções do Chrome
                    self.chrome_options = webdriver.ChromeOptions()
                    # Endereço de depuração para conexão com o Chrome
                    self.chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
                    # Inicializa o driver do Chrome com as opções e o gerenciador de drivers
                    self.driver = webdriver.Chrome(
                        service=Service(ChromeDriverManager().install()),
                        options=self.chrome_options)
                    self.driver.switch_to.window(self.driver.window_handles[0])

                    print("✅ Conectado ao navegador existente com sucesso.")
                except WebDriverException as e:
                    # Imprime mensagem de erro se a conexão falhar
                    print(f"❌ Erro ao conectar ao navegador existente: {e}")
                    # Define o driver como None em caso de falha na conexão
                    self.driver = None

            # Chama a função do webdriver com wait element to be clickable
            def webdriver_element_wait(self, xpath: str, num_element: int = 1):
                """
                        Espera até que um elemento web esteja clicável, usando um tempo limite máximo de 3 segundos.

                        Args:
                            xpath: O seletor XPath do elemento.
                            num_element: Identificador do npúmero de elementos que se espera serem retornados

                        Returns:
                            O elemento web clicável, ou lança uma exceção TimeoutException se o tempo limite for atingido.

                        Raises:
                            TimeoutException: Se o elemento não estiver clicável dentro do tempo limite.
                        """
                # Cria uma instância de WebDriverWait com o driver e o tempo limite e espera o elemento ser clicável
                if num_element == 1:
                    return WebDriverWait(self.driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                else:
                    return WebDriverWait(self.driver, 3).until(
                        lambda driver: [elem for elem in driver.find_elements(By.XPATH, xpath)]
                    )

            # Navega até a página de busca do instrumento ou proposta
            def consulta_instrumento(self):
                """
                       Navega pelas abas do sistema até a página de busca de processos.

                       Esta função clica nas abas principal e secundária para acessar a página
                       onde é possível realizar a busca de processos.
                       """
                # Reseta para página inicial
                try:
                    reset = WebDriverWait(self.driver, 3).until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="header"]')))
                    if reset:
                        img = reset.find_element(By.XPATH, '//*[@id="logo"]/a/img')
                        action = ActionChains(self.driver)
                        action.move_to_element(img).perform()
                        reset.find_element(By.TAG_NAME, 'a').click()
                        print(Fore.MAGENTA + "\n✅ Processo resetado com sucesso !")
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(Fore.RED + f'🔄❌ Falha ao resetar.\nErro: {type(e).__name__}\n{str(e)[:100]}')

                # [0] Excução; [1] Consultar Pré-Instrumento/Instrumento
                xpaths = ['//*[@id="menuPrincipal"]/div[1]/div[4]',
                          '//*[@id="contentMenu"]/div[1]/ul/li[6]/a'
                          ]
                try:
                    for idx in range(len(xpaths)):
                        self.webdriver_element_wait(xpaths[idx]).click()
                    print(f"{Fore.GREEN}✅ Sucesso em acessar a página de busca de processo{Style.RESET_ALL}")
                except NoSuchElementException:
                    pass
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(Fore.RED + f'🔴📄 Instrumento indisponível. \nErro: {e}{Style.RESET_ALL}')
                    sys.exit(1)

            def campo_pesquisa(self, numero_processo):
                try:
                    # Seleciona campo de consulta/pesquisa, insere o número de proposta/instrumento e da ENTER
                    campo_pesquisa = self.webdriver_element_wait('//*[@id="consultarNumeroConvenio"]')
                    campo_pesquisa.clear()
                    campo_pesquisa.send_keys(numero_processo)
                    campo_pesquisa.send_keys(Keys.ENTER)

                    # Acessa o item proposta/instrumento
                    acessa_item = self.webdriver_element_wait('/html/body/div[3]/div[14]/div[3]/div['
                                                              '3]/table/tbody/tr/td/div/a')
                    acessa_item.click()
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(
                        f' Falha ao inserir número de processo no campo de pesquisa. Erro: {type(e).__name__}')

            def busca_convenio(self):
                print('\n🔁🔍 Executando loop de pesquisa de convênio')
                # Seleciona aba primária, após acessar processo/instrumento. Aba Projeto Básico/Termo de referência
                termo_referencia = self.webdriver_element_wait('/html[1]/body[1]/div[3]/div[14]/'
                                                               'div[1]/div[1]/div[1]/a[4]/div[1]/span[1]/span[1]')
                termo_referencia.click()

                # Aba Projeto Básico/Termo de referência
                aba_termo_referencia = self.webdriver_element_wait('/html[1]/body[1]/div[3]/div[14]/div[1]'
                                                                   '/div[1]/div[2]/a[9]/div[1]/span[1]/span[1]')
                aba_termo_referencia.click()

            def acessa_anexos(self):
                try:
                    # Executa pesquisa de anexos
                    print('🔁📎 Executando pesquisa de anexos')

                    # Aba Plano de trabalho
                    self.webdriver_element_wait('//*[@id="div_997366806"]').click()

                    # Seleciona aba anexos
                    self.webdriver_element_wait('//*[@id="menu_link_997366806_1965609892"]/div').click()

                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(f"Ocorreu um erro ao executar ao pesquisar anexos: {type(e).__name__}"
                          f".\n Erro {e[:50]}")

            # Pesquisa o termo de fomento listado na planilha e executa download e transferência caso exista algúm.
            def loop_de_pesquisa(self, numero_processo: str):
                def conta_paginas(tabela):
                    # Diz quantas páginas tem
                    try:
                        paginas = tabela.find_element(By.TAG_NAME, 'span').text
                        paginas = paginas.split('(')[1]
                        paginas = int(paginas.split('i')[0].strip())
                        return paginas
                    except Exception:
                        print('No pages found')
                        return 0

                self.campo_pesquisa(numero_processo=numero_processo)

                paginas = 0
                try:
                    self.acessa_anexos()

                    # Seleciona lista de anexos execução e manda baixar os arquivos
                    try:

                        print('\n🔁📎 Executando pesquisa de anexos execução')

                        time.sleep(0.3)
                        botao_lista_execucao = self.webdriver_element_wait('//tbody//tr//input[2]')

                        if botao_lista_execucao.is_displayed() or botao_lista_execucao.is_enabled():
                            try:
                                # Seleciona lista de anexos execução e acessa a mesma
                                botao_lista_execucao.click()
                            except TimeoutException:
                                print(f"Timeout waiting for element: {'//tbody//tr//input[2]'}")
                            except Exception as e:  # Catch other potential exceptions
                                exc_type, exc_value, exc_tb = sys.exc_info()
                                print(
                                    Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                                print(f"🤷‍♂️❌ Erro ao tentar entar na lista de anexos execução: {e}")

                            # Encontra a tabela de anexos
                            tabela = WebDriverWait(self.driver, 10).until(
                                EC.presence_of_element_located((By.ID, 'listaAnexos')))
                            # Diz quantas páginas tem
                            paginas = conta_paginas(tabela)


                        else:
                            # Volta para a aba de consulta (começo do loop) caso não tenha lista de execução
                            self.webdriver_element_wait('/html[1]/body[1]/div[3]/div[2]/div[6]/a[2]').click()

                        print(
                            f"\n{Fore.GREEN}✅ Loop de pesquisa concluído para o processo:"
                            f" {numero_processo}{Style.RESET_ALL}\n")

                        return paginas

                    except Exception as er:
                        exc_type, exc_value, exc_tb = sys.exc_info()
                        print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                        print(f'❌ Falha ao acessar documentos de execução.'
                              f'\nErro:{type(er).__name__}, {str(er)[:50]}')
                        self.consulta_instrumento()

                except TimeoutException as t:
                    print(f'TIMEOUT {t[:50]}')
                    self.consulta_instrumento()
                except Exception as erro:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(
                        f'❌ Falha ao acessar documentos de execução. Erro:{type(erro).__name__}\n{str(erro)[:50]}')
                    self.consulta_instrumento()

            # Limpa o nome do arquivo
            @staticmethod
            def clean_filename(filename):
                # Remove or replace invalid characters for Windows filenames
                return re.sub(r'[\\/*?:"<>|]', "_", filename)

            # Salva a tela de esclarecimento detalhado.
            def loop_esclarecimento(self, numero_processo: str):
                # Xpath [0] Acomp. e Fiscalização // [1] Esclarecimento
                loop_list = ['//*[@id="menuInterno"]/div/div[7]',
                             '//*[@id="contentMenuInterno"]/div/div[1]/ul/li[3]/a']

                conta_resposta = 0

                for j in loop_list:
                    try:
                        self.webdriver_element_wait(j).click()
                    except TimeoutException as e:
                        print(f'❌ Falha ao acessar documentos de esclarecimento. Erro:{type(e).__name__}')

                print(f'🖨️🖼️ Imprimindo tela do processo {numero_processo}.')

                try:
                    # Encontra a tabela de anexos
                    tabela = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.ID, 'esclarecimentos')))
                    # Diz quantas páginas tem
                    paginas = self.conta_paginas(tabela)

                    try:
                        for pagina in range(1, paginas + 1):
                            if pagina > 1:
                                element = WebDriverWait(self.driver, 10).until(
                                    EC.element_to_be_clickable((By.LINK_TEXT, f'{pagina}')))
                                element.click()
                                print(f'\nAcessando página {pagina}\n ')

                                # Encontra a tabela na página atual
                            tabela = WebDriverWait(self.driver, 3).until(
                                EC.presence_of_element_located((By.ID, 'esclarecimentos')))
                            # Encontra todas as linhas da tabela atual
                            linhas = tabela.find_elements(By.TAG_NAME, 'tr')

                            previous_text = ''

                            for indice in range(1, len(linhas)):
                                indice = int(indice)
                                try:
                                    # Refresh table and rows reference periodically
                                    if indice >= 1:
                                        tabela = WebDriverWait(self.driver, 3).until(
                                            EC.presence_of_element_located((By.ID, 'esclarecimentos')))
                                        linhas = tabela.find_elements(By.TAG_NAME, 'tr')

                                    linha = linhas[indice]
                                    if linha.text.split('\n')[3] == previous_text:
                                        continue

                                    botao_detalhar = linha.find_element(By.CLASS_NAME, 'buttonLink')
                                    botao_detalhar.click()
                                except Exception as error:
                                    exc_type, exc_value, exc_tb = sys.exc_info()
                                    print(
                                        Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                                    print(
                                        f"❌ Erro ao processar a linha nº{indice},"
                                        f" erro: {type(error).__name__}\n{str(error)[:80]}")
                                    break

                                try:
                                    time.sleep(0.5)
                                    # Encontra a tabela na página atual
                                    tabela_resp = self.driver.find_element(
                                        By.XPATH,
                                        '/html/body/div[3]/div[14]/div[3]/div/div/form/table/tbody/tr['
                                        '18]/td/div[1]')
                                    # Encontra todas as linhas da tabela atual
                                    linhas_resp = tabela_resp.find_elements(By.XPATH, './/tbody/tr')

                                    print(len(linhas_resp))
                                    conta_resposta += len(linhas_resp)

                                except NoSuchElementException:
                                    pass
                                except TimeoutException:
                                    pass
                                except Exception as err:
                                    exc_type, exc_value, exc_tb = sys.exc_info()
                                    print(
                                        Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}"
                                                   f"{Style.RESET_ALL}")
                                    print(Fore.RED +
                                        f'❌ Erro de download: Termo {str(err)[:100]}.'
                                        f' Err: {type(err).__name__}{Style.RESET_ALL}')

                                self.driver.execute_script(
                                    "window.scrollTo(0, document.body.scrollHeight);")
                                pyautogui.hotkey('alt', 'left')

                    except Exception as error:
                        exc_type, exc_value, exc_tb = sys.exc_info()
                        print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                        print(f"❌ Erro ao buscar nova página de esclarecimento: {error}.\nErr:"
                              f" {type(error).__name__}")
                except Exception as error:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(f"❌ Erro ao econtrar elementos na página de esclarecimentos: {error}. Err:"
                          f" {type(error).__name__}")

                self.consulta_instrumento()

                print(
                    f"\n{Fore.GREEN}✅ Loop de esclarecimento concluído para o processo:"
                    f" {numero_processo}{Style.RESET_ALL}\n")

                return conta_resposta

            # Finds which locator to use
            def find_button_with_retry(self):
                # Define all possible locator strategies and values
                locators = [
                    (By.ID, 'form_submit'),
                    (By.NAME, 'detalharEsclarecimentoConvenioDadosDoEsclarecimentoVoltarForm'),
                    (By.XPATH, '//input[@value="Voltar"]'),
                    (By.XPATH, '//td[@class="FormLinhaBotoes"]/input'),
                    (By.CLASS_NAME, 'FormLinhaBotoes'),  # Will need additional find after
                    (By.XPATH, '//input[contains(@onclick, "setaAcao")]'),
                    (By.XPATH, '//*[@id="form_submit"]')
                ]

                for locator in locators:
                    try:
                        print(f'Botão localizado com o seletor {locator[0]}')
                        return self.driver.find_element(*locator)
                    except Exception as e:
                        exc_type, exc_value, exc_tb = sys.exc_info()
                        print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                        print(f"Falha com localizador {locator}: {str(e)[:100]}")
                        continue

                raise NoSuchElementException("Could not find button using any locator strategy")

            # Conta quantas páginas tem para iterar sobre
            @staticmethod
            def conta_paginas(tabela):
                # Diz quantas páginas tem
                try:
                    paginas = tabela.find_element(By.TAG_NAME, 'span').text
                    paginas = paginas.split('(')[0]
                    paginas.strip()
                    paginas = int(paginas[-3:])
                    return paginas
                except NoSuchElementException:
                    paginas = 1
                    return paginas

            @staticmethod
            def converter_data(data: str) -> datetime | None:
                """
                Função auxiliar para converter uma string de data em um objeto datetime,
                tratando diferentes formatos.

                Args:
                    data: String de data a ser convertida.

                Returns:
                    Um objeto datetime correspondente à data fornecida ou None se o formato for inválido.
                """
                formatos_tentativa = [
                    '%Y-%m-%d %H:%M:%S',  # Formato com hora
                    '%Y-%m-%d',  # Formato sem hora
                    '%Y-%m-%dT%H:%M:%S.%fZ',  # Formato ISO 8601 com microssegundos e 'Z'
                    '%d/%m/%Y'  # Formato de data BR
                ]

                for formato in formatos_tentativa:
                    try:
                        return datetime.strptime(data, formato)
                    except ValueError:
                        pass  # Tenta o próximo formato

                print(f"⚠️📆 Formato de data inválido: {data}")
                return None

            @staticmethod
            def data_hoje():
                # pega a data do dia que está executando o programa
                agora = datetime.now()
                data_hoje = datetime(agora.year, agora.month, agora.day)
                return data_hoje

            @staticmethod
            def extrair_dados_excel(caminho_arquivo_fonte, busca_id: str, tipo_instrumento_id: str,
                                    situacional_id: str) -> tuple:
                """
                   Lê os contatos de uma planilha Excel e executa ações baseadas nos dados extraídos.

                   Args:
                       caminho_arquivo_fonte (str): Caminho do arquivo Excel que será lido.
                       busca_id (str): Nome da coluna que contém os números de processo.
                       tipo_instrumento_id (str): Nome da coluna que contém o tipo de processo.
                   """

                pd.set_option('future.no_silent_downcasting', True)
                dados_processo = pd.read_excel(caminho_arquivo_fonte, dtype=str)
                dados_processo = dados_processo.fillna('')
                dados_processo = dados_processo.replace({u'\xa0': ''})
                dados_processo = dados_processo.infer_objects(copy=False)

                # Cria um lista para cada coluna do arquivo xlsx
                numero_processo = list()
                tipo_instrumento = list()
                situacional = list()

                try:
                    # Itera a planilha e armazena os dados em listas
                    for indice, linha in dados_processo.iterrows():  # Assume que a primeira linha e um cabeçalho
                        numero_processo.append(linha[busca_id])  # Busca o número do processo
                        tipo_instrumento.append(linha[tipo_instrumento_id])  # Busca o tipo de instrumento
                        situacional.append(linha[situacional_id])  # Busca a situação do processo
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(f"❌ Erro de leitura encontrado, erro: {e}")

                return numero_processo, tipo_instrumento, situacional

            # Limpa os dados que vem da planilha
            def limpa_dados(self, lista: list):
                """
                Limpa os elementos de uma lista, removendo quebras de linha, espaços extras e
                substituindo valores NaN por strings vazias.

                :param lista: Lista contendo os elementos a serem processados.
                :return: Uma nova lista com os elementos limpos.
                """
                lista_limpa = []
                for i in lista:
                    if pd.isna(i):
                        lista_limpa.append('')
                        continue
                    # Remove quebra de linha "\n"
                    i_limpo = i.replace('\n', '').replace('\r', '').strip()
                    # Anexa à lista limpa
                    lista_limpa.append(i_limpo)
                return lista_limpa

            # Salva o progresso em um arquivo json
            def salva_progresso(self, processo_visitado: str, indice: int, cont_arq_anexo: int,
                                cont_escl:int, arquivo_log):
                """
                Salva o progresso atual em um arquivo JSON.

                :param arquivo_log: Endereço do arquivo JSON
                :param processo_visitado: Lista de processos concluídos.
                :param arquivos_baixados: Lista de arquivos baixados.
                :param indice: Diz qual linha o programa iterou por último.
                """
                # Carrega os dados antigos

                try:
                    dados_log = self.carrega_progresso(arquivo_log=arquivo_log)

                    if dados_log:
                        # Carrega os dados novos
                        dados_log[processo_visitado] = {
                            "arquivos_anexo": cont_arq_anexo,
                            "arquivos_esclarecimento": cont_escl,
                            "indice": indice
                        }


                        with open(arquivo_log, 'w', encoding='utf-8') as arq:
                            json.dump(dados_log, arq, indent=4, ensure_ascii=False)
                        print(f"💾 Processo salvo: {processo_visitado} (Índice: {indice})")
                        print(f"   Arquivos anexo: {cont_arq_anexo}, Esclarecimento: {cont_escl}")
                    else:
                        print(f'Falha ao carregar os arquivos. Retornado {type(dados_log)}')
                        sys.exit(0)
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(f"\n❌ Erro ao salvar progresso, tipo: {type(e).__name__}\n{str(e)[:100]}")


            # Carrega os dados do arquivo JSON que sereve como Cartão de Memória
            def carrega_progresso(self, arquivo_log: str):
                """
                    Carrega o progresso do arquivo JSON.

                    :param arquivo_log: Endereço do arquivo JSON

                    :return: Um dicionário contendo os dados de progresso.
                             Se o arquivo não existir, retorna valores padrão.
                """
                try:
                    if not os.path.exists(arquivo_log):
                        print("⚠️  Arquivo de log não existe, retornando dict vazio")
                        return {}

                    with open(arquivo_log, 'r', encoding='utf-8') as arq:
                        print('Success reading file !')
                        return json.load(arq)
                except json.JSONDecodeError as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")
                    print(f"JSON decode error: {e}")
                    print(f"Content that failed to parse: {type(e).__name__}\n{str(e)[:100]}")
                    return None

                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")
                    print(f"Other error tipo: {type(e).__name__}\n{str(e)[:100]}")
                    return None


            # Reseta o arquivo JSON
            def reset(self, arquivo_log: str):
                """
                Reseta o progresso atual em um arquivo JSON.

                :param arquivo_log: Endereço do arquivo JSON

                """
                dados_vazios = {
                    "processo_visitado": [],
                    "indice": 0
                }
                try:
                    # Salva os dados vazios no arquivo JSON
                    with open(arquivo_log, 'w', encoding='utf-8') as arq:
                        json.dump(dados_vazios, arq, indent=4)
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")
                    print(f"Other error tipo: {type(e).__name__}\n{str(e)[:100]}")
                    return None
            # Verifica as condições para mandar um e-mail para o técnico
            def condicao_email(self, numero_processo: str, caminho_pasta: str):
                """
                Verifica se há arquivos modificados na data de hoje dentro de um diretório específico.

                :param numero_processo: Número do processo relacionado aos arquivos.
                :param caminho_pasta: Caminho da pasta onde os arquivos estão armazenados.
                :return: Uma tupla contendo (numero_processo, caminho_pasta, lista de arquivos modificados hoje).
                         Caso não haja arquivos modificados hoje, retorna uma lista vazia.
                """

                # lista que guarda os arquivos novos, na função ENVIAR_EMAIL_TECNICO recebe o nome de lista_documentos
                docs_atuais = []
                try:
                    # Data de hoje
                    hoje = self.data_hoje()
                    # Itera os arquivos da pasta para buscar a data de modificação individual
                    for arq_nome in os.listdir(caminho_pasta):
                        arq_caminho = os.path.join(caminho_pasta, arq_nome)
                        # Pula diretórios
                        if os.path.isfile(arq_caminho):
                            data_mod = datetime.fromtimestamp(os.path.getmtime(arq_caminho))
                            # Compara as datas de modificação dos arquivos
                            if data_mod >= hoje:
                                docs_atuais.append(arq_nome)
                    if docs_atuais:
                        print(f"📂✨ Documentos novos encontrados para o processo {numero_processo}")
                        return numero_processo, caminho_pasta, docs_atuais

                    else:

                        print(f"⚠️Nenhum documento novo encontrado para o processo {numero_processo}.\n")
                        return []
                except Exception:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f"⚠️Nenhum documento novo encontrado para o processo {numero_processo}.\n")
                    return []


        def main() -> None:
            # Caminho do arquivo .xlsx que contem os dados necessários para rodar o robô
            caminho_arquivo_fonte = (
                r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                r'Social\Automações SNEAELIS\Analise_Custos_Exec_Print\source\RTMA Passivo 2024 - '
                r'PROJETOS E PROGRAMAS (3) Atualizado em Julho de.2025.xlsx ')

            arquivo_log = (
                r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                r'Social\Automações SNEAELIS\Analise_Custos_Exec_Print\source\verify_log.json')

            try:
                # Instancia um objeto da classe Robo
                robo = Robo()
                # Extrai dados de colunas específicas do Excel
            except Exception as e:
                exc_type, exc_value, exc_tb = sys.exc_info()
                print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                print(f"\n‼️ Erro fatal ao iniciar o robô: {e}")
                sys.exit("Parando o programa.")

            numero_processo, tipo_instrumento, situacional = robo.extrair_dados_excel(
                caminho_arquivo_fonte=caminho_arquivo_fonte,
                busca_id='Instrumento nº',
                tipo_instrumento_id='Regime Jurídico do Instrumento (Modalidade)',
                situacional_id='SITUACIONAL'
            )

            max_linha = len(numero_processo)

            # Pós-processamento dos dados para não haver erros na execução do programa
            numero_processo = robo.limpa_dados(numero_processo)

            # Inicia o processo de consulta do instrumento
            robo.consulta_instrumento()

            robo.reset(arquivo_log=arquivo_log)

            inicio_range = 0

            for indice in range(inicio_range, max_linha):
                if situacional[indice] != '':
                    continue
                try:
                    # Executa pesquisa dos termos e salva os resultados na pasta "caminho_pasta"
                    pagina = robo.loop_de_pesquisa(numero_processo=numero_processo[indice])

                    conta_resposta = robo.loop_esclarecimento(numero_processo=numero_processo[indice])

                    robo.salva_progresso(indice=indice,
                                         processo_visitado=numero_processo[indice],
                                         cont_arq_anexo=pagina,
                                         cont_escl=conta_resposta,
                                         arquivo_log=arquivo_log
                                         )

                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(Fore.RED + f"Error occurred at line: {exc_tb.tb_lineno}{Style.RESET_ALL}")

                    print(f"\n❌ Erro ao processar o índice {indice} ({numero_processo[indice]}): {e}")
                    robo.consulta_instrumento()
                    continue  # Continua para o próximo processo

        main()


    # Check files on analise_custos_print
    elif func == 11:
        def conta_arquivo_zip(zip_path: str) -> int:
            try:
                with zipfile.ZipFile(zip_path, 'r') as zf:
                    return  len([f for f in zf.namelist() if not f.endswith('/')])
            except Exception:
                print(f'No files foun at {os.path.basename(zip_path)}')
                return 0

        def conta_arquivo_esclarecimento(dir_path: str) -> tuple:
            pattern = re.compile(r'\d{2}_\d{2}_\d{4}_\d{1,2}_\d{4}')
            counter_zip = 0
            counter_esc = 0

            for file in os.listdir(dir_path):
                if file.lower().endswith(".zip") and file.split('.')[0] == os.path.basename(dir_path):
                   counter_zip = conta_arquivo_zip(os.path.join(dir_path, file))
                   continue
                elif pattern.search(file):
                    continue

                counter_esc += 1

            return counter_esc, counter_zip

        def compara_com_json(dir_path: str, json_path: str, json_out_path: str):
            key = os.path.basename(dir_path).split('_')[0]

            with open(json_path, 'r', encoding='utf-8') as f:
                dados = json.load(f)

            if key not in dados:
                print(f"⚠️ Chave {key} não encontrada no JSON.")
                return

            arq_esclarecimento, arq_anexo = conta_arquivo_esclarecimento(dir_path)

            esperado_anexo = dados[key].get('arquivos_anexo', 0)
            esperado_esclarecimento = dados[key].get('arquivos_esclarecimento', 0)

            # Comparação
            has_diff = False

            if arq_anexo != esperado_anexo or arq_esclarecimento != esperado_esclarecimento:
                print(f"📂 Processo {key}")
                print(f"   Anexo → encontrado: {arq_anexo}, esperado: {esperado_anexo}")
                print(
                    f"   Esclarecimento → encontrado: {arq_esclarecimento},"
                    f" esperado: {esperado_esclarecimento}")
                has_diff = True

            else:
                print(f"✅ Processo {key}: sem diferenças")

            resultado = {
                "processo": key,
                "encontrado": {
                    "arquivos_anexo": arq_anexo,
                    "arquivos_esclarecimento": arq_esclarecimento,
                },
                "esperado": {
                    "arquivos_anexo": esperado_anexo,
                    "arquivos_esclarecimento": esperado_esclarecimento,
                }
            }
            if os.path.exists(json_out_path):
                with open(json_out_path, 'r', encoding='utf-8') as f_load:
                    try:
                        dados = json.load(f_load)
                    except json.JSONDecodeError:
                        dados = []
            else:
                dados = []

            if not isinstance(dados, list):
                dados = [dados]

            dados.append(resultado)

            with open(json_out_path, 'w', encoding='utf-8') as f_out:
                json.dump(dados, f_out, indent=4, ensure_ascii=False)


            return has_diff


        def main():
            json_out_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                r'Social\Automações '
             r'SNEAELIS\Analise_Custos_Exec_Print\source\comparison_log.json')

            dir_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                r'Social\Automações '
             r'SNEAELIS\Analise_Custos_Exec_Print')

            json_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                r'Social\Automações '
             r'SNEAELIS\Analise_Custos_Exec_Print\source\verify_log.json')

            xlsx_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                r'Social\Automações '
             r'SNEAELIS\Analise_Custos_Exec_Print\source\RTMA Passivo 2024 - PROJETOS E PROGRAMAS (3) '
                         r'Atualizado em Julho de.2025.xlsx')

            count_diff = 0
            to_redo_list = []

            df = pd.read_excel(xlsx_path, dtype=str)
            isnt_num = list()
            missing_instruments = list()

            for indice, linha in df.iterrows():  # Assume que a primeira linha e um cabeçalho
                if (linha['SITUACIONAL']) != '':
                    isnt_num.append(linha['Instrumento nº'])


            for root, _, files in os.walk(dir_path):
                for d in _:
                    if compara_com_json(dir_path=os.path.join(root, d),
                                     json_path=json_path,
                                     json_out_path=json_out_path):
                        count_diff += 1
                        to_redo_list.append(d)
                    if d not in isnt_num:
                        missing_instruments.append(d)

            print(f"\n🔎 Total de diretórios com diferenças: {count_diff}\n\nList to redo:"
                  f"{to_redo_list}\n\nDid not iterate: {missing_instruments}, {len(missing_instruments)}")


        main()


    # Compare xlsx files(DFP directories)
    elif func == 12:
        def find_matching_values(old_xlsx, new_xlsx, col1, col2):
            df1 = pd.read_excel(old_xlsx)
            df2 = pd.read_excel(new_xlsx)

            values1 = set(df1[col1].dropna().astype(str))
            values2 = set(df2[col2].dropna().astype(str))

            # Find matches
            matches = values1.intersection(values2)
            only_in_old_xlsx = values1 - values2
            only_in_new_xlsx = values2 - values1

            print(f"Total matches: {len(matches)}")
            print(f"Only in {os.path.basename(old_xlsx)}: {len(only_in_old_xlsx)}")
            print(f"Only in {os.path.basename(new_xlsx)}: {len(only_in_new_xlsx)}")

            # Create a results DataFrame
            results = pd.DataFrame({
                'Matches': list(matches),
                'Only_in_File1': list(only_in_old_xlsx) + [''] * (len(matches) - len(only_in_old_xlsx)),
                'Only_in_File2': list(only_in_new_xlsx) + [''] * (len(matches) - len(only_in_new_xlsx))
            })

            return results


        old_xlsx = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                    r'Social\Teste001\Processos Gerar Parecer TF.xlsm')
        new_xlsx = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                    r'Social\Teste001\fabi_DFP\Propostas Para Diligências Padrão.xlsm')

        # Usage
        results = find_matching_values(old_xlsx, new_xlsx, 'Nº Proposta', 'Nº Proposta')
        print(results)


    # Create directories
    elif func == 13:
        root_dir = r'C:\Users\felipe.rsouza\Documents\fabi'
            # The directory whose contents are to be copied
        std_dir_mapping = {
            'Termo de Fomento': (
                r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                r'Social\SNEAELIS - Propostas 1\Termo de Fomento\25448_2025 - 71000057438202574\Portarias e '
                r'Parecer Referencial CONJUR'),
            'Convênio': (
                r'C:\Users\felipe.rsouza\Documents\Convênio\29378-2025 - 71000062161202500\Portarias e Parecer '
                r'Referencial CONJUR')
        }

        target_folder_name = 'Portarias e Parecer Referencial CONJUR'

        # Process each directory type separately
        for dir_type, std_source in std_dir_mapping.items():
            type_dir = os.path.join(root_dir, dir_type)

            if not os.path.exists(type_dir):
                print(f"Directory not found: {type_dir}")
                continue

            # Iterate through all subdirectories of the type directory
            for main_dir in os.listdir(type_dir):
                main_dir_path = os.path.join(type_dir, main_dir)

                if os.path.isdir(main_dir_path):
                    target_folder_path = os.path.join(main_dir_path, target_folder_name)

                    if os.path.exists(target_folder_path):
                        # Clear the target folder
                        for filename in os.listdir(target_folder_path):
                            file_path = os.path.join(target_folder_path, filename)
                            try:
                                if os.path.isfile(file_path) or os.path.islink(file_path):
                                    os.unlink(file_path)
                                elif os.path.isdir(file_path):
                                    shutil.rmtree(file_path)
                            except Exception as e:
                                print(f'Failed to delete {file_path}. Reason: {e}')

                        # Copy new contents
                        try:
                            shutil.copytree(std_source, target_folder_path, dirs_exist_ok=True)
                            print(f"Successfully updated: {target_folder_path}")
                        except Exception as e:
                            print(f"Failed to copy files to {target_folder_path}. Reason: {e}")
                    else:
                        print(f"Target folder not found: {target_folder_path}")

        print("Folder substitution completed successfully.")


    # Compress pdf
    elif func == 14:
        def compress_pdf_fitz(dir_path, dpi=150, quality=80):
            """
            Compress PDF using PyMuPDF - no poppler required
            """

            for root, dirs, files in os.walk(dir_path):
                for filename in files:
                    if filename.endswith('.pdf'):
                        input_path = os.path.join(root, filename)
                        print(f"\n{'⚡' * 3}🚀 EXECUTING FILE: {filename} 🚀{'⚡' * 3}".center(70, '=')
                              , '\n')

                    file_name_output = os.path.basename(filename) + '_compressed' + '.pdf'

                    output_path = os.path.join(root, file_name_output)
            # Open the original PDF
            doc = fitz.open(input_path)

            # Create a new PDF for compressed output
            writer = fitz.open()

            for page_num in range(len(doc)):
                page = doc[page_num]

                # Render page as image with specified DPI
                mat = fitz.Matrix(dpi / 72, dpi / 72)  # Convert DPI to matrix
                pix = page.get_pixmap(matrix=mat)

                # Convert to bytes and then to PIL Image for compression
                img_data = pix.tobytes("ppm")
                pil_img = Image.open(io.BytesIO(img_data))

                # Convert to RGB if necessary
                if pil_img.mode != 'RGB':
                    pil_img = pil_img.convert('RGB')

                # Compress image
                output_buffer = io.BytesIO()
                pil_img.save(output_buffer, format='JPEG', quality=quality, optimize=True)
                compressed_img_data = output_buffer.getvalue()

                # Create new page with compressed image
                new_page = writer.new_page(width=page.rect.width, height=page.rect.height)
                new_page.insert_image(new_page.rect, stream=compressed_img_data)

            # Save with compression options
            writer.save(output_path, garbage=4, deflate=True, clean=True)

            doc.close()
            writer.close()

            # Print compression results
            original_size = os.path.getsize(input_path)
            compressed_size = os.path.getsize(output_path)
            reduction = (1 - compressed_size / original_size) * 100

            print(f"Original: {original_size / 1024 / 1024:.2f} MB")
            print(f"Compressed: {compressed_size / 1024 / 1024:.2f} MB")
            print(f"Reduction: {reduction:.1f}%")


        dir_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                    r'Social\Teste001\pdf_converter')

        compress_pdf_fitz(dir_path=dir_path,
                            dpi=120,
                            quality=80)

    # PDF slicer
    elif func == 15:
        from  PyPDF2 import PdfReader, PdfWriter

        reader = PdfReader(r"C:\Users\felipe.rsouza\Downloads\02. RG_E_RESIDENCIA.pdf")

        pages_len = len(reader.pages)

        writer = PdfWriter()

        writer.add_page(reader.pages[pages_len - 1])

        with open(r"C:\Users\felipe.rsouza\Documents\rg.pdf", 'wb') as out_file:
            writer.write(out_file)

            print(f"Last page extracted and saved to {out_file}")
        print(f"Total pages in original: {pages_len}")

    # Checks folder names with excel column
    elif func == 17:
        root_path = r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Automações SNEAELIS\CGAC_2024'
        xlsx_path = r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Automações SNEAELIS\CGAC_2024\CGAC_2024_dataSource_filtered.xlsx'

        df = pd.read_excel(xlsx_path, dtype=str)

        folder_names = [
            name.split('_')[0] for name in os.listdir(root_path)
            if os.path.isdir(os.path.join(root_path, name))
        ]

        # Convert DataFrame column to a set for fast lookup
        df_values = set(df['Nº CONVÊNIO'].astype(str))

        # Compare and find matches
        matches = [folder for folder in folder_names if folder in df_values]

        print(matches)

    # Text cleaner for monitor.py
    elif func == 18:
        def text_cleaner(text: str):
            text = re.sub(r'\([^)]*\)', '', text)# Remove (content)
            text = text.replace('.', '')

            text = re.sub(r'\s+', ' ', text).strip()
            text = text.upper()
            return text


        def create_matrix(file_path_org, file_path_dst, unique_values, process_col='Processo'):
            df_original = pd.read_excel(file_path_org, dtype=str).fillna('').astype(str)

            df_result = pd.DataFrame({process_col: df_original[process_col].unique()})

            for value in unique_values:
                df_result[value] = ''

            for idx, row in df_original.iterrows():
                process_num = row[process_col]

                if pd.isna(row[process_col]) or process_num == '':
                    continue

                clear_row = []
                for value in row.values:
                    clear_row.append(text_cleaner(str(value)))


                for value in unique_values:
                    if str(value).upper() in clear_row:
                        df_result.loc[df_result[process_col] == process_num, value] = 'X'

            print(len(df_result))
            df_result.to_excel(file_path_dst, index=False)


        file_path_org = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                         r'Social\Teste001\Monitoramento processos SEi\monitoramento_sei.xlsx')
        file_path_dst = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                         r'Social\Teste001\Monitoramento processos SEi\data_set - Copia.xlsx')

        unique_values = ['DAPC', 'DFP', 'GAB-AssTécnica', 'CGC', 'CGFP', 'Processo não possui andamentos '
                                                                         'abertos', 'DIE', 'DEAELIS',
                         'CGAP-TEF', 'CGEALIS', 'CGIIE', 'CGAP-CVTD', 'CGAP', 'GAB-COAAD', 'CGPIE',
                         'GAB-CMOF']

        create_matrix(file_path_org, file_path_dst, unique_values, process_col='Processo')

    elif func == 19:
        def copy_style(source_cell, target_cell):
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)


        origin_file_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                      r'Social\Teste001\Monitoramento processos SEi\Controle SNEAELIS - 2025 - Copia.xlsx')

        data_set_file_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e '
                              r'Assistência '
                     r'Social\Teste001\Monitoramento processos SEi\data_set - Copia.xlsx')

        destiny_file_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                     r'Social\Teste001\Monitoramento processos SEi\propostas_fluxos_20251113_120435 (1) - '
                             r'Copia.xlsx')

        origin_df = pd.read_excel(origin_file_path)
        data_set_df  = pd.read_excel(data_set_file_path)

        columns_to_append = [col for col in data_set_df.columns if col != 'Processo']

        process_to_protocol = origin_df[['Processo', 'Nº Proposta']].drop_duplicates()

        data_set_protocol_map = pd.merge(
            data_set_df,
            process_to_protocol,
            on='Processo',
            how='left'
        )
        print(f"data_set_with_protocol shape: {data_set_protocol_map.shape}\n")

        wb = load_workbook(destiny_file_path)
        ws = wb.active

        last_col = ws.max_column

        new_cols = [col for col in data_set_protocol_map.columns if col != 'Nº Proposta']

        header_row = 1
        header_cells = [cell.value for cell in ws[header_row]]

        try:
            protocol_col_idx = header_cells.index('Número da Proposta') + 1
            print(f"Protocol column found at index: {protocol_col_idx}")
        except:
            print("ERROR: 'Número da Proposta' column not found!")
            protocol_col_idx = 1

        for i, col_name in enumerate(new_cols, start=1):
            new_col_idx = last_col + i

            ws.cell(row=1, column=new_col_idx).value = col_name
            copy_style(ws.cell(row=1, column=last_col), ws.cell(row=1, column=new_col_idx))

            for row_idx in range(2, ws.max_row + 1):
                protocol_num = ws.cell(row=row_idx, column=protocol_col_idx).value
                matching_data = data_set_protocol_map[data_set_protocol_map['Nº Proposta'] == protocol_num]

                if not matching_data.empty:
                    value =matching_data.iloc[0][col_name]
                    ws.cell(row=row_idx, column=new_col_idx).value = value

        s = time.time()
        wb.save(destiny_file_path)
        d = time.time()
        elapsed = datetime.timedelta(seconds=d - s)

        # Remove microseconds for cleaner output
        elapsed_str = str(elapsed).split('.')[0]
        print(f"Save time: {elapsed_str}")

    # Test selectors
    elif func == 20:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC


        def setup_driver_with_existing_session():
            """
            Connect to existing Chrome session with remote debugging
            """
            chrome_options = Options()
            chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

            try:
                driver = webdriver.Chrome(options=chrome_options)
                print("✅ Successfully connected to existing Chrome session")
                return driver
            except Exception as e:
                print(f"❌ Failed to connect to Chrome session: {e}")
                return None


        def extract_total_value(driver):
            """
            Extract the TOTAL GERAL value using multiple selector strategies
            Returns the value and shows which selectors worked
            """

            selectors = {# Direct selectors for the specific cell
                "css_total_geral_cell": "tr.odd:last-child .valorTotal",
                "xpath_total_geral_cell": "//tr[contains(@class, 'odd') and .//div[text()='TOTAL GERAL']]//div[@class='valorTotal']",

                # Table structure based selectors
                "css_table_last_row": "#valoresTotais #tbodyrow tr:last-child .valorTotal",
                "xpath_table_last_row": "//div[@id='valoresTotais']//tbody[@id='tbodyrow']//tr[last()]//div[@class='valorTotal']",

                # Text content based selectors
                "xpath_by_text": "//div[text()='TOTAL GERAL']/following-sibling::td//div[@class='valorTotal']",
                "css_by_parent_text": "tr:has(div.descricao:contains('TOTAL GERAL')) .valorTotal",

                # ID and class combination selectors
                "css_id_combination": "#valoresTotais #tbodyrow .odd .valorTotal:last-child",
                "xpath_id_combination": "//div[@id='valoresTotais']//tr[@class='odd' and position()=last()]//div[@class='valorTotal']",

                # Multiple class selectors
                "css_multi_class": "tr.odd .valorTotal",
                "xpath_multi_class": "//tr[contains(@class, 'odd')]//div[@class='valorTotal']",

                # Additional robust selectors
                "css_specific_total": "tr:last-child .valorTotal",
                "xpath_specific_total": "//tr[.//div[contains(text(), 'TOTAL GERAL')]]//div[@class='valorTotal']", }

            working_selectors = {}
            total_value = None

            print("🔍 Testing multiple selectors for TOTAL GERAL:")
            print("=" * 50)

            for selector_name, selector in selectors.items():
                try:
                    if selector_name.startswith("css"):
                        elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    else:  # xpath
                        elements = driver.find_elements(By.XPATH, selector)

                    if elements:
                        # For selectors that return multiple elements, find the one with TOTAL GERAL
                        target_element = None
                        for element in elements:
                            if selector_name in ["css_multi_class", "xpath_multi_class"]:
                                # Check if this row contains "TOTAL GERAL"
                                row_text = element.find_element(By.XPATH, "./ancestor::tr[1]").text
                                if "TOTAL GERAL" in row_text:
                                    target_element = element
                                    break
                            else:
                                target_element = element
                                break

                        if target_element:
                            value = target_element.text.strip()
                            working_selectors[selector_name] = value

                            # Store the first successful value
                            if total_value is None:
                                total_value = value

                            print(f"✅ {selector_name}: {value}")
                        else:
                            print(f"⚠️ {selector_name}: Found elements but no target")
                    else:
                        print(f"❌ {selector_name}: No elements found")

                except Exception as e:
                    print(f"❌ {selector_name}: Failed - {str(e)[:50]}...")

            print(f"\n🎯 Total working selectors: {len(working_selectors)}/{len(selectors)}")
            print(f"💰 Extracted value: {total_value}")

            return total_value, working_selectors


        def extract_all_table_data(driver):
            """
            Extract all data from the table for comprehensive analysis
            """
            table_data = {}

            try:
                # Extract all rows
                rows = driver.find_elements(By.CSS_SELECTOR, "#valoresTotais #tbodyrow tr")

                print("📋 Extracting all table data:")
                print("=" * 50)

                for row in rows:
                    try:
                        # Get category name
                        category_element = row.find_element(By.CSS_SELECTOR, ".descricao")
                        category = category_element.text.strip()

                        # Get all values in the row
                        values = {'valorTotal': row.find_element(By.CSS_SELECTOR, ".valorTotal").text.strip(),
                            'valorContrapartida': row.find_element(By.CSS_SELECTOR, ".valorContrapartida").text.strip(),
                            'valorBensServicos': row.find_element(By.CSS_SELECTOR, ".valorBensServicos").text.strip(),
                            'valorAplicacao': row.find_element(By.CSS_SELECTOR, ".valorAplicacao").text.strip()}

                        table_data[category] = values
                        print(f"📊 {category}: {values}")

                    except Exception as e:
                        print(f"⚠️ Error extracting row: {e}")

                return table_data

            except Exception as e:
                print(f"❌ Error finding table: {e}")
                return {}


        def get_current_page_info(driver):
            """
            Get information about the current page
            """
            try:
                current_url = driver.current_url
                page_title = driver.title
                print(f"🌐 Current URL: {current_url}")
                print(f"📄 Page Title: {page_title}")
                return current_url, page_title
            except Exception as e:
                print(f"❌ Error getting page info: {e}")
                return None, None


        def main():
            """
            Main function to control existing Chrome session
            """
            # Connect to existing Chrome session
            driver = setup_driver_with_existing_session()

            if not driver:
                print("❌ Could not connect to Chrome. Make sure you ran Chrome with:")
                print(
                    '"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="%USERPROFILE%\\chrome_profile"')
                return

            try:
                # Get current page information
                print("\n" + "=" * 60)
                current_url, page_title = get_current_page_info(driver)
                print("=" * 60)

                # Check if we're on the right page and table exists
                try:
                    # Wait a bit for potential dynamic content
                    driver.implicitly_wait(5)

                    # Test multiple selectors
                    total_value, working_selectors = extract_total_value(driver)

                    # Extract all table data
                    all_data = extract_all_table_data(driver)

                    if total_value:
                        print(f"\n🎉 Success! Extracted TOTAL GERAL: {total_value}")
                    else:
                        print(f"\n⚠️ Could not extract TOTAL GERAL. The table might not be visible or loaded.")

                except Exception as e:
                    print(f"❌ Error during extraction: {e}")
                    print("💡 Make sure you're on the page with the 'valoresTotais' table")

            except Exception as e:
                print(f"❌ Unexpected error: {e}")

            finally:
                # Note: We don't quit the driver since we're controlling an existing session
                print(
                    "\n🔧 Chrome session remains open. Close it manually when done.")  # If you want to close the browser, uncomment the line below:  # driver.quit()


        if __name__ == "__main__":
            main()

    # capture_position
    elif func == 21:
        def capture_position(element_name="element"):
            """Capture position with countdown"""
            print(f"Get ready to position mouse on {element_name}...")
            for i in range(7, 0, -1):
                print(f"Capturing in {i} seconds...")
                time.sleep(1)

            x, y = pyautogui.position()
            print(f"✓ {element_name}: X={x}, Y={y}")
            return x, y


        # Capture multiple elements
        login_button = capture_position("Login Button")
        #search_box = capture_position("Search Box")
        #submit_btn = capture_position("Submit Button")


    elif func == 22:
        def clear_dir(dir_path):
            try:
                files = [f for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f))]

                if not files:
                    print(f"  No files found in: {dir_path}")
                    return 0

                print(f"\n  Files to delete in '{os.path.basename(dir_path)}':")

                for file in files[:10]:
                    print(f"    - {file}")
                if len(files) > 10:
                    print(f"    ... and {len(files) - 10} more files")

                response = input(f"\n  Delete {len(files)} files from this directory? (y/n): ").strip().lower()

                if response == 'y':
                    del_count = 0
                    for file in files:
                        file_path = os.path.join(dir_path, file)
                        try:
                            os.remove(file_path)
                            del_count += 1
                        except:
                            print(f"    Error deleting {file}")

                    print(f"  ✓ Deleted {del_count} files from: {dir_path}")
                    return del_count

                else:
                    print(f"  ✗ Skipped: {dir_path}")
                    return 0

            except Exception as e:
                print(f"  Error accessing directory {dir_path}: {e}")
                return 0


        def find_and_clear_dirs(main_dir, tgt_name):
            print(f"Searching for directories named '{tgt_name}' in:")
            print(f"  {main_dir}")
            print("-" * 80)

            dir_to_tgt = [os.path.join(root, tgt_name) for root, dirs, files in os.walk(main_dir) if tgt_name in dirs]

            if not dir_to_tgt:
                print(f"No directories named '{tgt_name}' found!")
                return

            print(f"\nFound {len(dir_to_tgt)} directories named '{tgt_name}':")

            for i, dir_path in enumerate(dir_to_tgt, 1):
                parent = os.path.dirname(dir_path)
                grandparent = os.path.dirname(parent)

                print(f"\n{i}. Directory: {dir_path}")
                print(f"   Located in: {os.path.basename(parent)}/")

            response = input(f"\nProceed with cleaning {len(dir_to_tgt)} directories? (y/n): ").strip().lower()

            if response != 'y':
                print("Operation cancelled.")
                return

            total_deleted = 0
            processed_dirs = 0

            for dir_path in dir_to_tgt:
                print(f"\n{'=' * 60}")
                print(f"Processing directory {processed_dirs + 1} of {len(dir_to_tgt)}:")
                print(f"Path: {dir_path}")

                deleted = clear_dir(dir_path=dir_path)
                total_deleted += 1
                processed_dirs += 1

            # Summary
            print("\n" + "=" * 80)
            print("SUMMARY:")
            print(f"  Directories processed: {processed_dirs}")
            print(f"  Total files deleted: {total_deleted}")
            print("=" * 80)

        # Your main directory path
        main_directory = (r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência "
                          r"Social\SNEAELIS - Propostas 1\Termo de Fomento")
        target_dir_name = "Consultas Celebração"

        print("=" * 80)
        print("DIRECTORY CLEANUP TOOL")
        print("=" * 80)

        find_and_clear_dirs(main_dir=main_directory, tgt_name=target_dir_name)


    # copy sharepoint xlsx file to desire directory
    elif func == 23:
        def manual_sync_workflow(destiny_dir: str):
            # Option C: Browser download folder
            download_folder = os.path.expanduser("~/Downloads/")
            # Find the latest Excel file
            excel_files = [f for f in os.listdir(download_folder) if f.endswith('.xlsx')]

            if excel_files:
                latest_file = max(excel_files, key=lambda x: os.path.getmtime(os.path.join(download_folder, x)))
                shutil.move(os.path.join(download_folder, latest_file), destiny_dir)


        dir_path = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência '
                    r'Social\Teste001\fabi_DFP\Consultas.xlsx')

        manual_sync_workflow(destiny_dir=dir_path)

    # better fake function
    elif func == 24:
        class BreakInnerLoop(Exception):
            pass


        class PWRobo:
            def __init__(self, cdp_url: str = "http://localhost:9222"):
                self.playwright = sync_playwright().start()
                self.browser = self.playwright.chromium.connect_over_cdp(cdp_url)
                self.context = self.browser.contexts[0] if self.browser.contexts else self.browser.new_context()
                self.page = self.context.pages[0] if self.context.pages else self.context.new_page()

                # Enable resource blocking for faster performance
                # self.block_rss()

                print(f"✅ Connected to existing Chrome instance via Playwright. Connected to page: {self.page.url}")

            def block_rss(self):
                """Block images, stylesheets, and fonts to make browsing faster"""

                def route_handler(route, request):
                    if request.resource_type in ["image", "stylesheet", "font"]:
                        route.abort()
                    else:
                        route.continue_()

                self.page.route('**/*', route_handler)
                print("✅ Resource blocking enabled for faster performance")

            def consulta_proposta(self):
                """Navigates through the system tabs to the process search page."""
                try:
                    self.page.locator('xpath=//*[@id="logo"]/a').click(timeout=800)
                except PlaywrightTimeoutError:
                    print('Already on the initial page of transferegov discricionárias.')
                except PlaywrightError as e:
                    print(Fore.RED + f'🔄❌ Failed to reset.\nError: {type(e).__name__}\n{str(e)[:100]}')

                xpaths = ['xpath=//*[@id="menuPrincipal"]/div[1]/div[3]',
                          'xpath=//*[@id="contentMenu"]/div[1]/ul/li[2]/a']
                try:
                    for xpath in xpaths:
                        self.page.locator(xpath).click(timeout=10000)
                except PlaywrightError as e:
                    print(Fore.RED + f'🔴📄 Instrument unavailable. \nError: {e}')
                    sys.exit(1)

            def campo_pesquisa(self, numero_processo):
                try:
                    campo_pesquisa_locator = self.page.locator('xpath=//*[@id="consultarNumeroProposta"]')
                    campo_pesquisa_locator.fill(numero_processo)
                    campo_pesquisa_locator.press('Enter')
                    try:
                        acessa_item_locator = self.page.locator('xpath=//*[@id="tbodyrow"]/tr/td[1]/div/a')
                        acessa_item_locator.click(timeout=8000)
                    except PlaywrightTimeoutError:
                        print(f' Process number: {numero_processo}, not found.')
                        raise BreakInnerLoop
                    except PlaywrightError as e:
                        print(f' Process number: {numero_processo}, not found. Error: {type(e).__name__}')
                        raise BreakInnerLoop
                except PlaywrightError as e:
                    print(f' Failed to insert process number in the search field. Error: {type(e).__name__}')

            def busca_endereco(self):
                time.sleep(1)

                self.page.wait_for_timeout(500)
                try:
                    print(f'🔍 Starting data search'.center(50, '-'), '\n')

                    # Click the detail button
                    self.page.locator('input#form_submit[value="Detalhar"]').click()

                    # Wait for the page to load
                    self.page.wait_for_timeout(2000)

                    # Initialize with empty values
                    phone = ""
                    email = ""

                    # Check if elements exist before trying to extract text
                    phone_locator = self.page.locator('#txtTelefone')
                    email_locator = self.page.locator('#txtEmail')

                    phone_exists = phone_locator.count() > 0
                    email_exists = email_locator.count() > 0

                    print(f"🔍 Element check - Phone field exists: {phone_exists}, Email field exists: {email_exists}")

                    # Extract phone if available
                    if phone_exists:
                        try:
                            phone_locator.wait_for(timeout=3000)
                            phone = phone_locator.inner_text(timeout=2000).strip()
                            print(f"📞 Phone found: '{phone}'")
                        except PlaywrightTimeoutError:
                            print("⚠️ Phone element exists but content not loaded or empty")
                        except Exception as e:
                            print(f"⚠️ Error extracting phone: {e}")
                    else:
                        print("📞 Phone field not found on page")

                    # Extract email if available
                    if email_exists:
                        try:
                            email_locator.wait_for(timeout=3000)
                            email = email_locator.inner_text(timeout=2000).strip()
                            print(f"📧 Email found: '{email}'")
                        except PlaywrightTimeoutError:
                            print("⚠️ Email element exists but content not loaded or empty")
                        except Exception as e:
                            print(f"⚠️ Error extracting email: {e}")
                    else:
                        print("📧 Email field not found on page")

                    # Check if we got any data
                    if not phone and not email:
                        print("ℹ️ No contact information found on this page")
                    elif phone and not email:
                        print("ℹ️ Only phone number found")
                    elif email and not phone:
                        print("ℹ️ Only email found")
                    else:
                        print("✅ Both phone and email found")

                    # Back button
                    try:
                        self.page.locator('xpath=//*[@id="lnkConsultaAnterior"]').click(timeout=8000)
                    except:
                        print("⚠️ Could not click back button, navigating manually")
                        # Alternative navigation if back button fails
                        self.page.go_back()

                    return phone, email

                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f"❌ Unexpected error in busca_endereco: {type(e).__name__} - {str(e)[:100]}")
                    # Ensure we navigate back even on error
                    self.page.locator('xpath=//*[@id="lnkConsultaAnterior"]').click(timeout=2000)
                    return "", ""

            def mark_as_done(self, df, numero_processo, phone, email):
                """Safely mark row as done in the DataFrame"""
                time.sleep(1)
                try:
                    # Find the row index where the process number matches
                    mask = df.iloc[:, 4] == numero_processo

                    if mask.any():
                        # Get the actual index position(s)
                        idx_positions = df.index[mask]
                        for idx in idx_positions:
                            df.at[idx, 'Telefone'] = phone
                            df.at[idx, 'Email'] = email

                        print(f"✅ Updated process {numero_processo}: Phone={phone}, Email={email}")
                        return True
                    else:
                        print(f"❌ Process number {numero_processo} not found in DataFrame")
                        return False

                except Exception as err:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f"❌ Error in mark_as_done:{type(err).__name__}\n{str(err)[:100]}\n")
                    return False

            def loop_de_pesquisa(self, df, numero_processo: str):

                print(f'🔍 Starting data extraction loop'.center(50, '-'), '\n')
                xpath = [
                    '//*[@id="menu_link_330887298_1035188630"]',
                    '//*[@id="menu_link_330887298_-442619135"]',
                    '//*[@id="grupo_abas"]/a[3]',
                    '//*[@id="menu_link_2144784112_100344749"]',
                    '//*[@id="menu_link_2144784112_2061164"]',
                    '//*[@id="dados-cauc"]/siconv-historico/div/div[4]/button[2]',
                    '//*[@id="breadcrumbs"]/a[2]'
                ]
                try:
                    self.campo_pesquisa(numero_processo=numero_processo)
                    for x in xpath:
                        time.sleep(2.5)
                        self.page.locator(x).click(timeout=10000)

                    print(f"{'✨' * 3}📦 DADOS COLETADOS COM SUCESSO 📦{'✨' * 3}")
                except PlaywrightTimeoutError as t:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f'TIMEOUT {str(t)[:50]}')
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    self.consulta_proposta()
                    return False
                except Exception as erro:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f'❌ Failed to try to include documents. Error:{type(erro).__name__}\n'
                          f'{str(erro)[:100]}...')
                    self.consulta_proposta()
                    raise BreakInnerLoop

            # --- Fixed Save Function ---
            @staticmethod
            def save_to_excel(df, caminho_arquivo_fonte, sheet_name='Sheet1'):
                time.sleep(1.5)

                try:
                    # Create backup of original file
                    backup_path = caminho_arquivo_fonte.replace('.xlsx', '_backup.xlsx')
                    if os.path.exists(caminho_arquivo_fonte):
                        import shutil
                        shutil.copy2(caminho_arquivo_fonte, backup_path)
                        print(f"✅ Backup created: {backup_path}")

                    # Save using 'w' mode to overwrite the entire file
                    with ExcelWriter(caminho_arquivo_fonte, engine='openpyxl', mode='w') as writer:
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                    print(f"✅ Successfully saved {len(df)} rows to Excel.")
                    return True

                except PermissionError:
                    print(f"❌ Permission denied: Please close the Excel file '{caminho_arquivo_fonte}' and try again.")
                    return False
                except Exception as erro:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    line_number = exc_tb.tb_lineno
                    if exc_tb.tb_next:
                        line_number = exc_tb.tb_next.tb_lineno
                    print(f"Error occurred at line: {line_number}")
                    print(f'❌ Failed to save Excel file. Error:{type(erro).__name__}\n{str(erro)}')
                    return False

            @staticmethod
            def extrair_dados_excel(caminho_arquivo_fonte):
                try:
                    complete_data_frame = pd.read_excel(caminho_arquivo_fonte, dtype=str, sheet_name=None)
                    sheet_names_list = list(complete_data_frame.keys())

                    print("✅ Sheet Names Found:")
                    for name in sheet_names_list:
                        print(f"- {name}")

                    data_frame = complete_data_frame['Sheet1']
                    print(f"✅ Loaded {len(data_frame)} rows from Excel.")
                    return data_frame

                except Exception as e:
                    print(f"🤷‍♂️❌ Error reading the excel file: {os.path.basename(caminho_arquivo_fonte)}.\n"
                          f"Error name: {type(e).__name__}\nError: {str(e)}")

            @staticmethod
            def fix_prop_num(numero_proposta):
                time.sleep(0.75)

                if pd.isna(numero_proposta):
                    return False

                numero_proposta = str(numero_proposta)

                pattern = r'^\d{5}/\d{4}'

                if '_' in numero_proposta:
                    numero_proposta_fixed = numero_proposta.replace('_', '/')
                    return numero_proposta_fixed

                if re.findall(pattern, numero_proposta):
                    return numero_proposta
                else:
                    return False


        def main() -> None:
            dir_path = (
                r'C:\Users\felipe.rsouza\Documents\fake_func\Convênios - Pendências Celebração (24.11) - Copia.xlsx')

            try:
                robo = PWRobo()
            except Exception as e:
                print(f"\n‼️ Fatal error starting the robot: {e}")
                sys.exit("Stopping the program.")

            # Load DataFrame
            df = robo.extrair_dados_excel(caminho_arquivo_fonte=dir_path)
            if df is None:
                print("❌ Failed to load Excel file. Exiting.")
                return

            robo.consulta_proposta()

            for idx, row in df.iterrows():
                numero_processo_temp = row.iloc[4]
                numero_processo = robo.fix_prop_num(numero_processo_temp)

                if not numero_processo:
                    continue
                print()
                print(f"{'⚡' * 3}🚀 EXECUTING PROPOSAL: {numero_processo} 🚀{'⚡' * 3}".center(70, '='), '\n')
                print()
                print(f"Current progress {idx / len(df):.2%}.")

                try:
                    robo.loop_de_pesquisa(df=df, numero_processo=numero_processo)


                except BreakInnerLoop:
                    sys.exit()
                except KeyboardInterrupt:
                    print("Script stopped by user (Ctrl+C). Saving progress...")
                    sys.exit(0)
                except Exception as e:
                    exc_type, exc_value, exc_tb = sys.exc_info()
                    print(f"Error occurred at line: {exc_tb.tb_lineno}")
                    print(f"❌ Failed to execute script. Error: {type(e).__name__}\n{str(e)}")
                    # Save progress on error

            print(f"\n🎉 Processing complete!")


        if __name__ == "__main__":
            main()


    # merge aba_dados resuolts databases
    elif func == 25:
        df1 = pd.read_excel(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\Teste001\resultado_aba_dados_1-2.xlsx")
        df2 = pd.read_excel(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência "
                            r"Social\Teste001\resultado_aba_dados_2-2.xlsx")

        # Step 1: Merge the dataframes
        merged_df = pd.concat([df1, df2], ignore_index=True)

        # Step 3: Remove duplicate rows
        merged_df = merged_df.drop_duplicates()

        # Step 4: Reset index (optional, but nice to have clean index)
        merged_df = merged_df.reset_index(drop=True)

        # Display results
        print(f"Original df1: {len(df1)} rows")
        print(f"Original df2: {len(df2)} rows")
        print(f"After merge: {len(df1) + len(df2)} rows")
        print(f"After removing duplicates: {len(merged_df)} rows")

        merged_df.to_excel(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\SNEAELIS - webscraping\Resultado scraping Aba Dados\resultado_aba_dados.xlsx", index=False)


    # merges processos sei databases
    elif func == 26:
        df1 = pd.read_excel(
        r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\SNEAELIS - "
        r"webscraping\Consulta_SEi\Consultas parciais\consulta_direcao_sei_1-2.xlsx")
        df2 = pd.read_excel(
            r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência Social\SNEAELIS - "
            r"webscraping\Consulta_SEi\Consultas parciais\consulta_direcao_sei_2-2.xlsx")

        pattern = r'^([a-z]+\.[a-z]+)'

        # Step 1: Merge the dataframes
        merged_df = pd.concat([df1, df2], ignore_index=True)

        # Step 2: Remove rows that contain a period (".") in any cell
        specific_column = 'texto_link'

        matches_pattern_mask = merged_df[specific_column].astype(str).str.match(pattern, na=False)

        # Keep only rows that DO NOT have a period
        merged_df = merged_df[~matches_pattern_mask]

        # Step 3: Remove completely empty rows (rows where all values are NaN/empty)
        merged_df = merged_df.dropna(how='all')

        # Step 4: Remove duplicate rows
        merged_df = merged_df.drop_duplicates()

        # Step 5: Reset index (optional, but nice to have clean index)
        merged_df = merged_df.reset_index(drop=True)

        merged_df.to_excel(r"C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e Assistência "
                           r"Social\SNEAELIS - webscraping\Consulta_SEi\consulta_direcao_sei_final.xlsx")
        # Display results
        print(f"Original df1: {len(df1)} rows")
        print(f"Original df2: {len(df2)} rows")
        print(f"After merge: {len(df1) + len(df2)} rows")

        # Count rows with periods
        rows_with_period = matches_pattern_mask.sum()
        print(f"Rows containing pattern: {rows_with_period}")
        print(f"After removing pattern rows: {len(merged_df)} rows")
        print(f"After removing empty rows: {len(merged_df)} rows")
        print(f"After removing duplicates: {len(merged_df)} rows")

