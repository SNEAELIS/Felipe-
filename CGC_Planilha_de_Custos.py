from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import shutil
#import camelot

# Trata a tabela para automatização
def separa_blocos_tabelas(xlsx_entrada: str):
    # Read Excel without converting to strings initially

    df = pd.read_excel(xlsx_entrada, header=None)

    # Find blank rows (all cells NaN or empty strings)
    blank_rows = (df.isnull().all(axis=1) | (df.astype(str) == 'nan').all(axis=1))

    # Find rows containing "Descrição" (case insensitive)
    desc_mask = df.apply(lambda row: row.astype(str).str.contains('Descrição', case=False).any(), axis=1)
    desc_indices = df.index[desc_mask].tolist()

    with pd.ExcelWriter(xlsx_entrada) as writer:
        for i, start_idx in enumerate(desc_indices, 1):
            # Find next blank row after current Descrição
            end_idx = start_idx + 1
            while end_idx < len(df) and not blank_rows.iloc[end_idx]:
                end_idx += 1

            # Extract ONLY this block (original data, no modifications)
            bloco = df.iloc[start_idx:end_idx].copy()
            if len(bloco) < 2:
                continue

            # Get sheet name from row below Descrição (column 2)
            nome_tab = str(bloco.iloc[1, 2]) if len(bloco) > 1 else f'Tab_{i}'
            nome_tab = nome_tab[:28]  # Excel sheet name limit

            bloco = bloco.iloc[2:]

            # Clean ONLY this block's NaN values (from below)
            for col in bloco.columns:
                nan_mask = bloco[col].isna()
                bloco.loc[nan_mask, col] = bloco[col].shift(-1)[nan_mask]

            # Drop duplicated row
            bloco = bloco.drop(bloco.index[1])

            # Replace remaining NaN with empty strings
            bloco = bloco.fillna('')

            try:
                headers = bloco.iloc[0].copy().fillna('').tolist()
                for i in range(len(bloco.columns)):
                    bloco = bloco.rename(columns={i: headers[i]})
                bloco = bloco.drop(bloco.index[0])

                # Write ONLY this block to its own sheet
                bloco.to_excel(writer, sheet_name=nome_tab, index=False, header=True)

            except Exception as e:
                print(e)
        return True

# Aplica a funçao de separar dados
def main_tratar_dados():
    arquivo_fonte = Path(r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento '
                         r'e Assistência Social\Teste001\CUSTOS')
    print(f"Searching in: {arquivo_fonte.resolve()}")

    arq_xlsx = list(x for x in arquivo_fonte.rglob('*') if x.suffix.lower() == '.xlsx' or x.suffix.lower()
                    == '.xls')
    print(f"Achados {len(arq_xlsx)} arquivos Excel.")
    for caminho_arq in arq_xlsx:
        try:
            print(f'Processando: {caminho_arq.name}')

            base, ext = os.path.splitext(caminho_arq)
            if ext == '.xls':
                df = pd.read_excel(caminho_arq, engine='xlrd')
                novo_nome = base + '.xlsx'
                df.to_excel(novo_nome, index=False, engine='openpyxl')
                caminho_arq = novo_nome

            separa_blocos_tabelas(xlsx_entrada=caminho_arq)
            format_valor_columns(file_path=caminho_arq)
        except Exception as e:
            print(f"❌ Falha ao processar o arquivo: {caminho_arq}: \n{e}")

# Muda o formato da célula
def format_valor_columns(file_path):
    """
    Formats all columns containing 'valor' in their header as accounting/currency
    across all sheets in an Excel file.

    Args:
        file_path (str): Path to the Excel file (.xlsx or .xls)
    """
    try:
        # Load the workbook while preserving all sheets
        wb = load_workbook(filename=file_path)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
             # Find columns with 'valor' in header (case-insensitive)
            valor_columns = []
            for col_idx, cell in enumerate(sheet[1], 1): # Check header row (row 1)
                if isinstance(cell.value, str) and 'valor' in cell.value.lower():
                    valor_columns.append(col_idx)

            # Apply accounting format to identified columns
            for col_idx in valor_columns:
                col_letter = get_column_letter(col_idx)
                for cell in sheet[col_letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = 'R$ #,##0.00;[Red]R$ -#,##0.00'

        wb.save(filename=file_path)
        print(f"✅ Successfully formatted 'valor' columns in {file_path}")
    except Exception as e:
        print(f"❌ Error processing {file_path}: {str(e)}")


# Copia as planilhas para uma pasta separada
def copia_xlsx(source_dir: str, dest_dir: str):
    max_filename_length = 50
    allowed_ext = ['.xlsx', '.xls']

    for root, _, files in os.walk(source_dir):
        for file in files:
            if file.lower().endswith(allowed_ext[0]) or file.lower().endswith(allowed_ext[1]):
                try:
                    source_path = os.path.join(root, file)
                    if not os.path.isfile(source_path):
                        continue

                    source_folder_name = os.path.basename(root)
                    base, ext = os.path.splitext(file)
                    if ext.lower() in allowed_ext:
                        if len(base) > max_filename_length:
                            base = base[:max_filename_length]

                        target_subdir = os.path.join(dest_dir, source_folder_name)
                        os.makedirs(target_subdir, exist_ok=True)

                        dest_path = os.path.join(target_subdir, f'{base}{ext}')

                        shutil.copy2(source_path,dest_path)
                        print(f"Copied: {source_path}\n --->\n {dest_path}\n")
                    else:
                        if len(base) > max_filename_length:
                            base = base[:max_filename_length]

                        target_subdir = os.path.join(dest_dir, source_folder_name)
                        os.makedirs(target_subdir, exist_ok=True)

                        ext= '.xlsx'
                        dest_path = os.path.join(target_subdir, f'{base}{ext}')

                        shutil.copy2(source_path,dest_path)
                        print(f"Copied: {source_path}\n --->\n {dest_path}\n")

                except FileNotFoundError as e:
                    if hasattr(e, "winerror") and e.winerror == 206:
                        print("⚠️ Skipped: Path too long (WinError 206)")
                        print(f"Folder: {source_folder_name}\n")
                    continue
                except Exception as e:
                    print(e)

def reorganiza_pasta():
    arquivo_fonte = Path(r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento '
                         r'e Assistência Social\Teste001\CUSTOS')
    arq_xlsx = list(x for x in arquivo_fonte.rglob('*') if x.suffix.lower() == '.xls')

    print(f"🧾 Found {len(arq_xlsx)} '.xls' Excel file(s) to delete:\n")

    for f in arq_xlsx:
        print(' -', f)

    confirmar = input('\nDeseja deletar os arqivos? (s/n): ').strip().lower()
    if confirmar == 's':
        for f in arq_xlsx:
            try:
                f.unlink()
                print(f"🗑️ Deletado: {f}")
            except Exception as e:
                print(f"❌ Falha ao deletar {f}: \n{e}")
    else:
        print('❎ Process de deletar abortado.')


def main_copiar():
    source_dir = (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento '
                r'e Assistência Social\SNEAELIS - Prioridades - Análise de CNPJ')

    dest_dir =  (r'C:\Users\felipe.rsouza\OneDrive - Ministério do Desenvolvimento e'
                  r' Assistência Social\Teste001\CUSTOS')
    copia_xlsx(source_dir=source_dir, dest_dir=dest_dir)

if __name__ == '__main__':
    main_copiar()
    main_tratar_dados()
    reorganiza_pasta()
